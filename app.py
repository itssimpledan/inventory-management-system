"""
Inventory Management System  —  app.py
Run with: python app.py  (or double-click start.bat on Windows)
"""
import sqlite3, os, re
from datetime import datetime, date
from functools import wraps
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, g, send_file)
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = "ims-secret-key-change-in-production-2026"
DB_PATH = os.path.join(os.path.dirname(__file__), "inventory.db")

# ─── FRIENDLY ERROR PAGES ─────────────────────────────────────────────────────
@app.errorhandler(500)
def internal_error(e):
    return render_template("500.html"), 500

@app.errorhandler(404)
def not_found(e):
    return render_template("500.html"), 404

# ─── DB HELPERS ───────────────────────────────────────────────────────────────
def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop("db", None)
    if db: db.close()

# ─── AUTH HELPERS ─────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if session.get("role") not in roles:
                flash("You don't have permission to perform this action.", "danger")
                return redirect(url_for("dashboard"))
            return f(*args, **kwargs)
        return decorated
    return decorator

# ─── FIFO CALCULATION ────────────────────────────────────────────────────────
def calculate_fifo(db, item_id):
    layers = db.execute("""
        SELECT id, layer_id, po_reference, receipt_date, qty_received, unit_cost
        FROM fifo_layers WHERE item_id=? ORDER BY receipt_date ASC, id ASC
    """, (item_id,)).fetchall()

    total_consumed = db.execute("""
        SELECT COALESCE(SUM(qty_out),0) FROM inventory_ledger WHERE item_id=? AND qty_out>0
    """, (item_id,)).fetchone()[0]

    result, cum_qty = [], 0
    for layer in layers:
        qty_rec   = layer["qty_received"]
        cum_qty  += qty_rec
        cum_prev  = cum_qty - qty_rec
        qty_cons  = max(0, min(qty_rec, total_consumed - cum_prev))
        qty_rem   = qty_rec - qty_cons
        result.append({
            "layer_id":      layer["layer_id"],
            "po_reference":  layer["po_reference"],
            "receipt_date":  layer["receipt_date"],
            "qty_received":  qty_rec,
            "unit_cost":     layer["unit_cost"],
            "qty_consumed":  qty_cons,
            "qty_remaining": qty_rem,
            "value_remaining": qty_rem * layer["unit_cost"],
            "cogs_value":    qty_cons * layer["unit_cost"],
            "status": "Closed" if qty_rem == 0 else ("Open" if qty_rem == qty_rec else "Partial"),
        })
    return result

def get_wacc_summary(db):
    items = db.execute("SELECT item_id, name, reorder_point FROM items WHERE status='Active' ORDER BY item_id").fetchall()
    summary = []
    for item in items:
        layers    = calculate_fifo(db, item["item_id"])
        qty       = sum(l["qty_remaining"]   for l in layers)
        value     = sum(l["value_remaining"] for l in layers)
        cogs      = sum(l["cogs_value"]      for l in layers)
        wacc      = value / qty if qty > 0 else 0
        purchased = sum(l["qty_received"]    for l in layers)
        summary.append({
            "item_id":        item["item_id"],
            "name":           item["name"],
            "qty_on_hand":    qty,
            "inventory_value": value,
            "wacc":           wacc,
            "total_cogs":     cogs,
            "total_purchased": purchased,
            "reorder_point":  item["reorder_point"],
            "reorder_alert":  qty <= item["reorder_point"],
        })
    return summary

def next_fifo_cost(db, item_id):
    """Return unit cost of oldest open FIFO layer for an item."""
    for layer in calculate_fifo(db, item_id):
        if layer["qty_remaining"] > 0:
            return layer["unit_cost"]
    return 0.0

def next_po_number(db):
    yr = date.today().year
    last = db.execute(
        "SELECT po_number FROM purchase_orders WHERE po_number LIKE ? ORDER BY po_number DESC LIMIT 1",
        (f"PO-{yr}-%",)
    ).fetchone()
    if last:
        seq = int(last["po_number"].split("-")[-1]) + 1
    else:
        seq = 1
    return f"PO-{yr}-{seq:03d}"

def next_txn_id(db):
    last = db.execute("SELECT txn_id FROM inventory_ledger ORDER BY id DESC LIMIT 1").fetchone()
    if last:
        m = re.search(r"(\d+)$", last["txn_id"])
        seq = int(m.group(1)) + 1 if m else 1
    else:
        seq = 1
    return f"TXN-{seq:03d}"

# ─── AUTH ROUTES ─────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username","").strip().lower()
        password = request.form.get("password","")
        db   = get_db()
        user = db.execute("SELECT * FROM users WHERE username=? AND active=1", (username,)).fetchone()
        if user and check_password_hash(user["password"], password):
            session["user_id"]      = user["id"]
            session["username"]     = user["username"]
            session["full_name"]    = user["full_name"]
            session["role"]         = user["role"]
            session["profile_photo"] = user["profile_photo"] if user["profile_photo"] else None
            return redirect(url_for("dashboard"))
        flash("Invalid username or password.", "danger")
    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# ─── DASHBOARD ───────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    from datetime import date as _d
    db      = get_db()
    summary = get_wacc_summary(db)
    today   = _d.today().isoformat()

    # ── Inventory KPIs ──────────────────────────────────────────────────────
    total_inventory = sum(s["inventory_value"] for s in summary)

    # Top 5 SKUs by inventory value
    top5_skus = sorted(summary, key=lambda x: x["inventory_value"], reverse=True)[:5]

    # Monthly inventory value trend (last 12 complete months + current partial)
    # Compute cumulative value from all ledger history, then slice last 13 months
    all_monthly = db.execute("""
        SELECT strftime('%Y-%m', txn_date) AS month,
               SUM(COALESCE(qty_in,0)  * COALESCE(unit_cost,0)) AS value_in,
               SUM(COALESCE(qty_out,0) * COALESCE(unit_cost,0)) AS value_out
        FROM inventory_ledger
        GROUP BY month
        ORDER BY month
    """).fetchall()
    running = 0.0
    all_cum = []
    for r in all_monthly:
        running += (r["value_in"] or 0) - (r["value_out"] or 0)
        all_cum.append({"month": r["month"], "value": round(running, 2)})
    trend_months = all_cum[-13:] if len(all_cum) >= 13 else all_cum

    # Below historical-average sales run-rate
    # Compute avg monthly units sold per SKU from sales data
    avg_sales = {}
    sales_rows = db.execute("""
        SELECT sl.sku_id,
               SUM(sl.qty_sold) AS total_sold,
               COUNT(DISTINCT su.period) AS num_periods
        FROM sales_lines sl
        JOIN sales_uploads su ON sl.upload_id = su.id
        WHERE sl.sku_id IS NOT NULL AND sl.sku_id != ''
        GROUP BY sl.sku_id
    """).fetchall()
    for row in sales_rows:
        if row["num_periods"] and row["num_periods"] > 0:
            avg_sales[row["sku_id"]] = row["total_sold"] / row["num_periods"]

    below_runrate = []
    for s in summary:
        avg = avg_sales.get(s["item_id"], 0)
        if avg > 0 and s["qty_on_hand"] < avg:
            below_runrate.append({
                "item_id":    s["item_id"],
                "name":       s["name"],
                "qty_on_hand": s["qty_on_hand"],
                "avg_monthly": round(avg, 1),
                "weeks_cover": round(s["qty_on_hand"] / avg * 4.33, 1) if avg > 0 else None,
            })
    below_runrate.sort(key=lambda x: x["weeks_cover"] if x["weeks_cover"] is not None else 999)

    # ── PO KPIs ─────────────────────────────────────────────────────────────
    po_counts = {r["status"]: r["cnt"] for r in db.execute(
        "SELECT status, COUNT(*) cnt FROM purchase_orders GROUP BY status").fetchall()}
    active_pos = sum(v for k,v in po_counts.items()
                     if k not in ("Fully Paid – Received","Cancelled"))

    # Overdue payment: Approved or Deposit Paid, payment_due_date < today, balance > 0
    overdue_payment_pos = db.execute("""
        SELECT po.id, po.po_number, po.status, po.payment_due_date,
               s.name AS supplier_name,
               ROUND(COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0),2) AS total_val,
               COALESCE(po.amount_paid,0) AS amount_paid
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        LEFT JOIN (SELECT po_id, SUM(qty_ordered*unit_price) AS subtotal
                   FROM po_lines GROUP BY po_id) ts ON ts.po_id = po.id
        WHERE po.status IN ('Approved','Deposit Paid')
          AND po.payment_due_date IS NOT NULL
          AND po.payment_due_date < ?
          AND (COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0)
               - COALESCE(po.amount_paid,0)) > 0
        ORDER BY po.payment_due_date ASC
    """, (today,)).fetchall()

    # Overdue delivery: active POs (not received/cancelled) past expected delivery
    overdue_delivery_pos = db.execute("""
        SELECT po.id, po.po_number, po.status, po.expected_delivery_date,
               s.name AS supplier_name,
               (SELECT i2.name FROM po_lines pl2 JOIN items i2 ON pl2.item_id=i2.item_id
                WHERE pl2.po_id=po.id ORDER BY pl2.id LIMIT 1) AS first_item_name,
               (SELECT COUNT(*) FROM po_lines WHERE po_id=po.id) AS line_count
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        WHERE po.status IN ('Approved','Deposit Paid','Fully Paid – Shipped','Partially Received')
          AND po.expected_delivery_date IS NOT NULL
          AND po.expected_delivery_date < ?
        ORDER BY po.expected_delivery_date ASC
    """, (today,)).fetchall()

    # Recent POs (last 5)
    recent_pos = db.execute("""
        SELECT po.*, s.name supplier_name,
            (SELECT i2.name FROM po_lines pl2 JOIN items i2 ON pl2.item_id=i2.item_id
             WHERE pl2.po_id=po.id ORDER BY pl2.id LIMIT 1) AS first_item_name,
            (SELECT COUNT(*) FROM po_lines WHERE po_id=po.id) AS line_count,
            (SELECT COALESCE(SUM(pl.qty_ordered * pl.unit_price), 0)
             FROM po_lines pl WHERE pl.po_id=po.id)
             + COALESCE(po.freight_costs, 0) + COALESCE(po.other_costs, 0) AS total_val
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        ORDER BY po.created_at DESC LIMIT 5
    """).fetchall()

    # Returned POs raised by this planner (for demand planner alert)
    returned_pos = []
    if session.get("role") in ("ops_planner", "admin"):
        returned_pos = db.execute("""
            SELECT po.id, po.po_number, po.updated_at, s.name supplier_name
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            WHERE po.status = 'Returned'
            AND (po.raised_by = ? OR ? = 'admin')
            ORDER BY po.updated_at DESC
        """, (session.get("username"), session.get("role"))).fetchall()

    reorder_count = sum(1 for s in summary if s["reorder_alert"])

    return render_template("dashboard.html",
        summary=summary, po_counts=po_counts, recent_pos=recent_pos,
        total_inventory=total_inventory, reorder_count=reorder_count,
        active_pos=active_pos, returned_pos=returned_pos,
        top5_skus=top5_skus, trend_months=trend_months,
        below_runrate=below_runrate,
        overdue_payment_pos=overdue_payment_pos,
        overdue_delivery_pos=overdue_delivery_pos)

# ── Notification helper ──────────────────────────────────────────────────────
def create_po_notification(db, username, po_id, po_number, message, notif_type='info'):
    """Insert a new unread notification for a specific user."""
    try:
        db.execute("""INSERT INTO po_notifications
            (username, po_id, po_number, message, notif_type)
            VALUES (?,?,?,?,?)""",
            (username, po_id, po_number, message, notif_type))
    except Exception as e:
        print(f"[notification] Warning: {e}")

@app.context_processor
def inject_notifications():
    if 'username' not in session:
        return {'unread_notifications': [], 'unread_notif_count': 0}
    try:
        db = get_db()
        notifs = db.execute("""
            SELECT * FROM po_notifications
            WHERE username=? AND is_read=0
            ORDER BY created_at DESC LIMIT 30
        """, (session['username'],)).fetchall()
        return {'unread_notifications': notifs, 'unread_notif_count': len(notifs)}
    except Exception:
        return {'unread_notifications': [], 'unread_notif_count': 0}

@app.route("/notifications/<int:notif_id>/dismiss", methods=["POST"])
@login_required
def dismiss_notification(notif_id):
    db = get_db()
    db.execute("UPDATE po_notifications SET is_read=1 WHERE id=? AND username=?",
               (notif_id, session['username']))
    db.commit()
    return ('', 204)

@app.route("/notifications/dismiss-all", methods=["POST"])
@login_required
def dismiss_all_notifications():
    db = get_db()
    db.execute("UPDATE po_notifications SET is_read=1 WHERE username=?",
               (session['username'],))
    db.commit()
    return ('', 204)

# ─── PO ROUTES ───────────────────────────────────────────────────────────────
@app.route("/pos")
@login_required
def po_list():
    db = get_db()
    status_filter   = request.args.get("status","")
    item_filter     = request.args.get("item","")
    supplier_filter = request.args.get("supplier","")
    due_from_filter = request.args.get("due_from","")
    due_to_filter   = request.args.get("due_to","")

    query  = """SELECT po.*, s.name supplier_name,
                (SELECT COUNT(*) FROM po_lines WHERE po_id=po.id) AS line_count,
                (SELECT i2.name FROM po_lines pl2 JOIN items i2 ON pl2.item_id=i2.item_id
                 WHERE pl2.po_id=po.id ORDER BY pl2.id LIMIT 1) AS first_item_name,
                (SELECT pl2.item_id FROM po_lines pl2 WHERE pl2.po_id=po.id ORDER BY pl2.id LIMIT 1) AS first_item_id,
                (SELECT i2.name FROM po_lines pl2 JOIN items i2 ON pl2.item_id=i2.item_id
                 WHERE pl2.po_id=po.id ORDER BY pl2.id LIMIT 1 OFFSET 1) AS second_item_name,
                (SELECT i2.name FROM po_lines pl2 JOIN items i2 ON pl2.item_id=i2.item_id
                 WHERE pl2.po_id=po.id ORDER BY pl2.id LIMIT 1 OFFSET 2) AS third_item_name,
                (SELECT COALESCE(SUM(pl.qty_ordered * pl.unit_price), 0)
                 FROM po_lines pl WHERE pl.po_id=po.id)
                 + COALESCE(po.freight_costs, 0) + COALESCE(po.other_costs, 0) AS total_val,
                (SELECT COALESCE(SUM(pl.qty_ordered + COALESCE(pl.free_units,0)), 0)
                 FROM po_lines pl WHERE pl.po_id=po.id) AS total_qty
                FROM purchase_orders po
                JOIN suppliers s ON po.supplier_id = s.supplier_id
                WHERE 1=1"""
    params = []
    if status_filter:
        query += " AND po.status=?"; params.append(status_filter)
    if item_filter:
        query += " AND EXISTS (SELECT 1 FROM po_lines WHERE po_id=po.id AND item_id=?)"
        params.append(item_filter)
    if supplier_filter:
        query += " AND po.supplier_id=?"; params.append(supplier_filter)
    if due_from_filter:
        query += " AND po.payment_due_date >= ?"; params.append(due_from_filter)
    if due_to_filter:
        query += " AND po.payment_due_date <= ?"; params.append(due_to_filter)
    query += " ORDER BY po.po_date DESC"

    pos       = db.execute(query, params).fetchall()
    suppliers = db.execute("SELECT supplier_id, name FROM suppliers WHERE status='Active' ORDER BY name").fetchall()
    items     = db.execute("SELECT item_id, name FROM items WHERE status='Active' ORDER BY item_id").fetchall()
    statuses  = ["Draft","Pending Approval","Ops Approved","Approved","Deposit Paid",
                 "Fully Paid – Shipped","Fully Paid – Received","Partially Received","Cancelled"]
    return render_template("po_list.html", pos=pos, suppliers=suppliers,
                           items=items, statuses=statuses,
                           status_filter=status_filter, item_filter=item_filter,
                           supplier_filter=supplier_filter,
                           due_from_filter=due_from_filter,
                           due_to_filter=due_to_filter,
                           today=date.today().isoformat())

@app.route("/pos/<int:po_id>/quick_pay", methods=["POST"])
@login_required
@roles_required("admin","finance_exec","finance_manager")
def po_quick_pay(po_id):
    db = get_db()
    amount_paid = float(request.form.get("amount_paid") or 0)
    po = db.execute("""
        SELECT po.id, po.status,
               (SELECT COALESCE(SUM(pl.qty_ordered * pl.unit_price),0) FROM po_lines pl WHERE pl.po_id=po.id)
               + COALESCE(po.freight_costs,0) + COALESCE(po.other_costs,0) AS total_val
        FROM purchase_orders po WHERE po.id=?""", (po_id,)).fetchone()
    if not po:
        flash("PO not found.","danger")
        return redirect(url_for("po_list"))

    total_val = po["total_val"] or 0

    # ── Tax invoice gate ──────────────────────────────────────────────────────
    has_tax_invoice = db.execute("""
        SELECT COUNT(*) FROM po_documents
        WHERE po_id=? AND doc_type='Tax Invoice'""", (po_id,)).fetchone()[0]
    if not has_tax_invoice:
        flash("Cannot record payment: a Tax Invoice must be uploaded to this PO before processing payment.", "danger")
        return redirect(url_for("po_list"))

    # ── Validation ────────────────────────────────────────────────────────────
    if amount_paid < 0:
        flash("Payment amount cannot be negative.", "danger")
        return redirect(url_for("po_list"))
    if amount_paid > total_val:
        flash(f"Payment amount (${amount_paid:,.2f}) cannot exceed the PO total of ${total_val:,.2f}.", "danger")
        return redirect(url_for("po_list"))

    new_status = po["status"]
    if amount_paid == 0:
        new_status = "Approved"
    elif amount_paid < total_val:
        new_status = "Deposit Paid"
    else:
        new_status = "Fully Paid – Received" if po["status"] == "Fully Paid – Received" else "Fully Paid – Shipped"

    payment_date = request.form.get("payment_date") or date.today().isoformat()
    username     = session["username"]
    db.execute("""UPDATE purchase_orders
                  SET amount_paid=?, status=?,
                      deposit_date=CASE WHEN ? > 0 AND deposit_date IS NULL THEN ? ELSE deposit_date END,
                      deposit_paid_by=CASE WHEN ? > 0 AND deposit_paid_by IS NULL THEN ? ELSE deposit_paid_by END,
                      deposit_amount_paid=CASE WHEN ? > 0 AND deposit_amount_paid IS NULL OR deposit_amount_paid=0 THEN ? ELSE deposit_amount_paid END,
                      full_payment_date=CASE WHEN ? >= ? AND ? > 0 THEN ? ELSE full_payment_date END,
                      full_paid_by=CASE WHEN ? >= ? AND ? > 0 THEN ? ELSE full_paid_by END,
                      updated_at=datetime('now')
                  WHERE id=?""",
               (amount_paid, new_status,
                amount_paid, payment_date,
                amount_paid, username,
                amount_paid, amount_paid,
                amount_paid, total_val, amount_paid, payment_date,
                amount_paid, total_val, amount_paid, username,
                po_id))
    db.commit()
    flash(f"Payment updated: {new_status}", "success")
    redirect_to = request.form.get("redirect","list")
    if redirect_to == "detail":
        return redirect(url_for("po_detail", po_id=po_id))
    return redirect(url_for("po_list") + "?" + request.query_string.decode())

# ─── PO DOCUMENT UPLOAD / DOWNLOAD / DELETE ───────────────────────────────────

@app.route("/pos/<int:po_id>/upload_doc", methods=["POST"])
@login_required
@roles_required("admin","ops_planner","ops_exec")
def po_upload_doc(po_id):
    db = get_db()
    po = db.execute("SELECT id FROM purchase_orders WHERE id=?", (po_id,)).fetchone()
    if not po:
        flash("PO not found.", "danger")
        return redirect(url_for("po_list"))

    file      = request.files.get("doc_file")
    doc_type  = request.form.get("doc_type", "Other").strip()
    doc_notes = request.form.get("doc_notes", "").strip()

    if not file or file.filename == "":
        flash("No file selected.", "danger")
        return redirect(url_for("po_detail", po_id=po_id))

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_DOC_EXT:
        flash(f"File type '.{ext}' is not allowed. Accepted: {', '.join(sorted(ALLOWED_DOC_EXT))}", "danger")
        return redirect(url_for("po_detail", po_id=po_id))

    # Build a unique stored filename: <po_id>_<timestamp>_<original>
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    safe_name  = secure_filename(file.filename)
    stored_name = f"po{po_id}_{ts}_{safe_name}"
    file.save(os.path.join(PO_DOCS_FOLDER, stored_name))

    db.execute("""INSERT INTO po_documents (po_id, doc_type, filename, original_name, uploaded_by, notes)
                  VALUES (?,?,?,?,?,?)""",
               (po_id, doc_type, stored_name, file.filename,
                session["username"], doc_notes))
    db.commit()
    # Notify finance team when a Tax Invoice is uploaded
    if doc_type == "Tax Invoice":
        po_row = db.execute("SELECT po_number FROM purchase_orders WHERE id=?", (po_id,)).fetchone()
        if po_row:
            fin_users = db.execute(
                "SELECT username FROM users WHERE role IN ('finance_exec','finance_manager','admin') "
                "AND active=1"
            ).fetchall()
            for fu in fin_users:
                if fu["username"] != session["username"]:
                    create_po_notification(db, fu["username"], po_id, po_row["po_number"],
                        f"Tax Invoice uploaded for PO {po_row['po_number']} by {session['username']} — pending payment processing.",
                        'info')
            db.commit()
    flash(f"Document '{file.filename}' uploaded successfully.", "success")
    return redirect(url_for("po_detail", po_id=po_id))


@app.route("/pos/<int:po_id>/docs/<int:doc_id>/download")
@login_required
def po_download_doc(po_id, doc_id):
    db  = get_db()
    doc = db.execute("SELECT * FROM po_documents WHERE id=? AND po_id=?",
                     (doc_id, po_id)).fetchone()
    if not doc:
        flash("Document not found.", "danger")
        return redirect(url_for("po_detail", po_id=po_id))

    file_path = os.path.join(PO_DOCS_FOLDER, doc["filename"])
    if not os.path.exists(file_path):
        flash("File missing from server.", "danger")
        return redirect(url_for("po_detail", po_id=po_id))

    return send_file(file_path, as_attachment=True,
                     download_name=doc["original_name"])


@app.route("/pos/<int:po_id>/docs/<int:doc_id>/delete", methods=["POST"])
@login_required
@roles_required("admin","ops_planner","ops_exec")
def po_delete_doc(po_id, doc_id):
    db  = get_db()
    doc = db.execute("SELECT * FROM po_documents WHERE id=? AND po_id=?",
                     (doc_id, po_id)).fetchone()
    if not doc:
        flash("Document not found.", "danger")
        return redirect(url_for("po_detail", po_id=po_id))

    # Only admins or the original uploader can delete
    if session["role"] != "admin" and doc["uploaded_by"] != session["username"]:
        flash("You can only delete documents you uploaded.", "danger")
        return redirect(url_for("po_detail", po_id=po_id))

    # Remove physical file
    file_path = os.path.join(PO_DOCS_FOLDER, doc["filename"])
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except OSError:
        pass  # log if needed

    db.execute("DELETE FROM po_documents WHERE id=?", (doc_id,))
    db.commit()
    flash(f"Document '{doc['original_name']}' deleted.", "info")
    return redirect(url_for("po_detail", po_id=po_id))


@app.route("/pos/new", methods=["GET","POST"])
@login_required
@roles_required("admin","ops_planner")
def po_new():
    db = get_db()
    if request.method == "POST":
        f = request.form
        import json as _json
        # Use user-supplied PO number or auto-generate
        po_number = f.get("po_number","").strip().upper()
        if not po_number:
            po_number = next_po_number(db)

        item_ids = request.form.getlist("item_id[]")
        qtys     = request.form.getlist("qty_ordered[]")
        prices   = request.form.getlist("unit_price[]")
        frees    = request.form.getlist("free_units[]")
        lines_data = [(item_ids[i].strip(), qtys[i], prices[i],
                       frees[i] if i < len(frees) else "0")
                      for i in range(len(item_ids)) if item_ids[i].strip()]

        # Helper: re-render form with all data intact
        def _render_with_error(msg):
            flash(msg, "danger")
            suppliers_ = db.execute("SELECT supplier_id, name FROM suppliers WHERE status='Active' ORDER BY name").fetchall()
            form_data_ = {k: f.get(k,"") for k in
                ["po_number","po_date","supplier_id","currency","freight_costs",
                 "other_costs","deposit_pct","expected_delivery_date","payment_due_date","notes"]}
            prefill_ = [{"item_id": iid, "qty": q, "price": p, "free": fr}
                        for iid, q, p, fr in lines_data]
            return render_template("po_new.html",
                suppliers=suppliers_, today=date.today().isoformat(),
                suggested_po=po_number,
                form_data=form_data_,
                prefill_lines=_json.dumps(prefill_))

        if db.execute("SELECT id FROM purchase_orders WHERE po_number=?", (po_number,)).fetchone():
            return _render_with_error(f"PO number '{po_number}' already exists. Choose a different one.")

        if not lines_data:
            return _render_with_error("Please add at least one SKU line item.")

        try:
            db.execute("""INSERT INTO purchase_orders
                (po_number,po_date,supplier_id,currency,deposit_pct,
                 freight_costs,other_costs,expected_delivery_date,payment_due_date,
                 status,raised_by,notes)
                VALUES (?,?,?,?,?,?,?,?,?,'Pending Approval',?,?)""",
                (po_number, f["po_date"], f["supplier_id"], f["currency"],
                 float(f.get("deposit_pct") or 0),
                 float(f.get("freight_costs") or 0),
                 float(f.get("other_costs") or 0),
                 f.get("expected_delivery_date") or None,
                 f.get("payment_due_date") or None,
                 session["username"], f.get("notes","")))
            po_db_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            for (iid, qty, price, free) in lines_data:
                db.execute(
                    "INSERT INTO po_lines (po_id,item_id,qty_ordered,unit_price,free_units) VALUES (?,?,?,?,?)",
                    (po_db_id, iid, float(qty or 0), float(price or 0), float(free or 0)))
            db.commit()
            flash(f"PO {po_number} created with {len(lines_data)} line item(s).", "success")
            return redirect(url_for("po_list"))
        except Exception as e:
            return _render_with_error(f"Error creating PO: {e}")

    suppliers = db.execute("SELECT supplier_id, name FROM suppliers WHERE status='Active' ORDER BY name").fetchall()
    today     = date.today().isoformat()
    suggested_po = next_po_number(db)
    return render_template("po_new.html", suppliers=suppliers, today=today,
                           suggested_po=suggested_po, form_data=None, prefill_lines="[]")


# ── PO Edit (for Returned POs) ──────────────────────────────────────────────
@app.route("/pos/<int:po_id>/edit", methods=["GET","POST"])
@login_required
@roles_required("admin","ops_planner")
def po_edit(po_id):
    import json as _json
    db = get_db()
    po = db.execute("SELECT * FROM purchase_orders WHERE id=?", (po_id,)).fetchone()
    if not po:
        flash("PO not found.", "danger")
        return redirect(url_for("po_list"))
    if po["status"] != "Returned":
        flash("Only Returned POs can be edited.", "warning")
        return redirect(url_for("po_detail", po_id=po_id))
    # Only the original requestor (or admin) may amend a returned PO
    if session.get("role") != "admin" and po["raised_by"] != session.get("username"):
        flash("Only the original requestor can amend this PO.", "warning")
        return redirect(url_for("po_detail", po_id=po_id))

    lines_rows = db.execute("SELECT * FROM po_lines WHERE po_id=?", (po_id,)).fetchall()

    if request.method == "POST":
        f = request.form
        item_ids = request.form.getlist("item_id[]")
        qtys     = request.form.getlist("qty_ordered[]")
        prices   = request.form.getlist("unit_price[]")
        frees    = request.form.getlist("free_units[]")
        lines_data = [(item_ids[i].strip(), qtys[i], prices[i],
                       frees[i] if i < len(frees) else "0")
                      for i in range(len(item_ids)) if item_ids[i].strip()]

        def _render_edit_error(msg):
            flash(msg, "danger")
            suppliers_ = db.execute("SELECT supplier_id, name FROM suppliers WHERE status='Active' ORDER BY name").fetchall()
            form_data_ = {k: f.get(k, "") for k in
                ["po_number","po_date","supplier_id","currency","freight_costs",
                 "other_costs","deposit_pct","expected_delivery_date","payment_due_date","notes"]}
            prefill_ = [{"item_id": iid, "qty": q, "price": p, "free": fr}
                        for iid, q, p, fr in lines_data]
            return render_template("po_new.html",
                suppliers=suppliers_, today=date.today().isoformat(),
                suggested_po=po["po_number"],
                form_data=form_data_,
                prefill_lines=_json.dumps(prefill_),
                edit_mode=True, edit_po_id=po_id, edit_po_number=po["po_number"])

        if not lines_data:
            return _render_edit_error("Please add at least one SKU line item.")

        # Validate lines
        for iid, qty, price, free in lines_data:
            try:
                p = float(price or 0)
                fr = float(free or 0)
            except ValueError:
                return _render_edit_error("Invalid price or free units value.")
            if p == 0 and fr == 0:
                return _render_edit_error(f"SKU {iid}: unit price and free units cannot both be 0.")

        try:
            # Preserve existing notes but strip old [Returned …] stamp if present
            existing_notes = f.get("notes", "").strip()
            db.execute("""UPDATE purchase_orders SET
                po_date=?, supplier_id=?, currency=?, deposit_pct=?,
                freight_costs=?, other_costs=?,
                expected_delivery_date=?, payment_due_date=?,
                notes=?, status='Pending Approval', updated_at=datetime('now')
                WHERE id=?""",
                (f["po_date"], f["supplier_id"], f["currency"],
                 float(f.get("deposit_pct") or 0),
                 float(f.get("freight_costs") or 0),
                 float(f.get("other_costs") or 0),
                 f.get("expected_delivery_date") or None,
                 f.get("payment_due_date") or None,
                 existing_notes, po_id))
            # Replace lines
            db.execute("DELETE FROM po_lines WHERE po_id=?", (po_id,))
            for (iid, qty, price, free) in lines_data:
                db.execute(
                    "INSERT INTO po_lines (po_id,item_id,qty_ordered,unit_price,free_units) VALUES (?,?,?,?,?)",
                    (po_id, iid, float(qty or 0), float(price or 0), float(free or 0)))
            db.commit()
            flash(f"PO {po['po_number']} updated and resubmitted for approval.", "success")
            return redirect(url_for("po_detail", po_id=po_id))
        except Exception as e:
            return _render_edit_error(f"Error updating PO: {e}")

    # GET — pre-populate form
    suppliers = db.execute("SELECT supplier_id, name FROM suppliers WHERE status='Active' ORDER BY name").fetchall()
    form_data = {
        "po_number":             po["po_number"],
        "po_date":               po["po_date"],
        "supplier_id":           po["supplier_id"],
        "currency":              po["currency"],
        "freight_costs":         str(po["freight_costs"] or 0),
        "other_costs":           str(po["other_costs"] or 0),
        "deposit_pct":           str(po["deposit_pct"] or 0),
        "expected_delivery_date":po["expected_delivery_date"] or "",
        "payment_due_date":      po["payment_due_date"] or "",
        "notes":                 po["notes"] or "",
    }
    prefill = [{"item_id": r["item_id"], "qty": r["qty_ordered"],
                "price": r["unit_price"], "free": r["free_units"]}
               for r in lines_rows]
    return render_template("po_new.html",
        suppliers=suppliers, today=date.today().isoformat(),
        suggested_po=po["po_number"],
        form_data=form_data,
        prefill_lines=_json.dumps(prefill),
        edit_mode=True, edit_po_id=po_id, edit_po_number=po["po_number"])


@app.route("/pos/<int:po_id>", methods=["GET","POST"])
@login_required
def po_detail(po_id):
    db = get_db()
    po = db.execute("""SELECT po.*, s.name supplier_name,
                              s.payment_terms, s.currency sup_currency
                       FROM purchase_orders po
                       JOIN suppliers s ON po.supplier_id = s.supplier_id
                       WHERE po.id=?""", (po_id,)).fetchone()
    if not po:
        flash("PO not found.", "danger"); return redirect(url_for("po_list"))

    if request.method == "POST":
        action = request.form.get("action","")
        role   = session["role"]

        # ── Two-step PO approval ─────────────────────────────────────────────
        if action == "approve":
            if po["status"] == "Pending Approval" and role in ("admin","ops_manager"):
                db.execute("""UPDATE purchase_orders SET status='Ops Approved',
                    ops_approved_by=?, ops_approval_date=?,
                    updated_at=datetime('now') WHERE id=?""",
                    (session["username"], date.today().isoformat(), po_id))
                db.commit()
                create_po_notification(db, po["raised_by"], po_id, po["po_number"],
                    f"PO {po['po_number']} approved by Ops Manager ({session['username']}) — now awaiting Finance approval.",
                    'info')
                flash("PO approved by Ops Manager. Awaiting Finance approval.", "success")
            elif po["status"] == "Ops Approved" and role in ("admin","finance_manager"):
                db.execute("""UPDATE purchase_orders SET status='Approved',
                    approved_by=?, approval_date=?, updated_at=datetime('now') WHERE id=?""",
                    (session["username"], date.today().isoformat(), po_id))
                db.commit()
                create_po_notification(db, po["raised_by"], po_id, po["po_number"],
                    f"PO {po['po_number']} fully approved by Finance ({session['username']}) — ready for payment.",
                    'success')
                flash("PO fully approved. Ready for payment.", "success")
            else:
                flash("Action not permitted for your role or current PO status.", "danger")

        elif action == "record_payment" and role in ("admin","finance_exec","finance_manager"):
            amount_paid  = float(request.form.get("amount_paid") or 0)
            payment_date = request.form.get("payment_date") or date.today().isoformat()
            # Compute total to determine new status
            total_row = db.execute("""
                SELECT ROUND((COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0)),2) AS total_val
                FROM purchase_orders po
                LEFT JOIN (SELECT po_id,SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
                  ON ts.po_id=po.id WHERE po.id=?""", (po_id,)).fetchone()
            total_val_r = total_row["total_val"] if total_row else 0
            # ── Tax invoice gate ──────────────────────────────────────────────
            has_tax_invoice = db.execute("""
                SELECT COUNT(*) FROM po_documents
                WHERE po_id=? AND doc_type='Tax Invoice'""", (po_id,)).fetchone()[0]
            # ── Validation ────────────────────────────────────────────────────
            if not has_tax_invoice:
                flash("Cannot record payment: a Tax Invoice must be uploaded before processing payment.", "danger")
            elif amount_paid < 0:
                flash("Payment amount cannot be negative.", "danger")
            elif total_val_r and amount_paid > total_val_r:
                flash(f"Payment amount (${amount_paid:,.2f}) cannot exceed the PO total of ${total_val_r:,.2f}.", "danger")
            else:
                username = session["username"]
                if amount_paid == 0:
                    new_status = "Approved"
                elif amount_paid < (total_val_r or 0):
                    new_status = "Deposit Paid"
                else:
                    new_status = "Fully Paid – Shipped" if po["status"] != "Fully Paid – Received" else "Fully Paid – Received"
                db.execute("""UPDATE purchase_orders
                              SET amount_paid=?, status=?,
                                  deposit_date=CASE WHEN ? > 0 AND deposit_date IS NULL THEN ? ELSE deposit_date END,
                                  deposit_paid_by=CASE WHEN ? > 0 AND deposit_paid_by IS NULL THEN ? ELSE deposit_paid_by END,
                                  deposit_amount_paid=CASE WHEN ? > 0 AND (deposit_amount_paid IS NULL OR deposit_amount_paid=0) THEN ? ELSE deposit_amount_paid END,
                                  full_payment_date=CASE WHEN ? >= ? AND ? > 0 THEN ? ELSE full_payment_date END,
                                  full_paid_by=CASE WHEN ? >= ? AND ? > 0 THEN ? ELSE full_paid_by END,
                                  updated_at=datetime('now')
                              WHERE id=?""",
                           (amount_paid, new_status,
                            amount_paid, payment_date,
                            amount_paid, username,
                            amount_paid, amount_paid,
                            amount_paid, total_val_r, amount_paid, payment_date,
                            amount_paid, total_val_r, amount_paid, username,
                            po_id))
                db.commit()
                flash(f"Payment recorded: ${amount_paid:,.2f} — status updated to {new_status}.", "success")

        elif action == "cancel":
            # Managers can cancel any non-cancelled, non-closed PO; planners can only cancel Returned POs
            can_cancel = (
                (role in ("admin","ops_manager","finance_manager") and
                 po["status"] not in ("Cancelled", "Closed – Partial")) or
                (role == "ops_planner" and po["status"] == "Returned")
            )
            if can_cancel:
                cancel_reason = request.form.get("cancel_reason","").strip()
                needs_resolution = (po["amount_paid"] or 0) > 0
                db.execute("""UPDATE purchase_orders SET status='Cancelled',
                    cancelled_by=?, cancelled_at=?, cancel_reason=?,
                    cancellation_resolution=CASE WHEN ? > 0 THEN 'Pending' ELSE NULL END,
                    updated_at=datetime('now') WHERE id=?""",
                    (session["username"], date.today().isoformat(),
                     cancel_reason or None, po["amount_paid"] or 0, po_id))
                db.commit()
                # Notify raised_by that their PO was cancelled
                cancel_note = f" Reason: {cancel_reason}" if cancel_reason else ""
                create_po_notification(db, po["raised_by"], po_id, po["po_number"],
                    f"PO {po['po_number']} was cancelled by {session['username']}.{cancel_note}",
                    'warning')
                # If deposit/partial payment was already made, alert the full finance team
                if (po["amount_paid"] or 0) > 0:
                    fin_users = db.execute(
                        "SELECT username FROM users WHERE role IN ('finance_exec','finance_manager','admin') "
                        "AND active=1"
                    ).fetchall()
                    for fu in fin_users:
                        if fu["username"] != session["username"]:
                            create_po_notification(db, fu["username"], po_id, po["po_number"],
                                f"ATTENTION: PO {po['po_number']} was cancelled by {session['username']} "                                f"but a payment of ${po['amount_paid']:,.2f} has already been recorded. "                                f"Please follow up with the supplier.",
                                'warning')
                    db.commit()
                flash("PO cancelled.", "warning")
            else:
                flash("Action not permitted for your role or current PO status.", "danger")

        elif action == "return_for_amendment" and role in ("admin","ops_manager","finance_manager"):
            allowed = (
                (po["status"] == "Pending Approval" and role in ("admin","ops_manager")) or
                (po["status"] == "Ops Approved"     and role in ("admin","finance_manager"))
            )
            if allowed:
                reason = request.form.get("reject_reason","").strip()
                note   = f"[Returned by {session['username']} on {date.today().isoformat()}]: {reason}"
                existing = po["notes"] or ""
                combined = (existing + "\n" + note).strip() if existing else note
                db.execute("""UPDATE purchase_orders SET status='Returned',
                    notes=?, updated_at=datetime('now') WHERE id=?""",
                    (combined, po_id))
                db.commit()
                reason_txt = f" Reason: {reason}" if reason else ""
                create_po_notification(db, po["raised_by"], po_id, po["po_number"],
                    f"PO {po['po_number']} was returned for amendment by {session['username']}.{reason_txt}",
                    'warning')
                flash("PO returned for amendment. The requestor has been notified.", "warning")
            else:
                flash("Action not permitted for current PO status.", "danger")

        elif action == "resolve_cancellation" and role in ("admin","finance_exec","finance_manager"):
            if po["status"] == "Cancelled" and po.get("cancellation_resolution") == "Pending":
                res_type  = request.form.get("resolution_type","").strip()
                res_notes = request.form.get("resolution_notes","").strip()
                if not res_type:
                    flash("Please select a resolution type.", "danger")
                elif not res_notes:
                    flash("Please provide resolution notes before closing this off.", "danger")
                else:
                    db.execute("""UPDATE purchase_orders SET
                        cancellation_resolution='Resolved',
                        resolution_type=?, resolution_notes=?,
                        resolved_by=?, resolved_at=?,
                        updated_at=datetime('now') WHERE id=?""",
                        (res_type, res_notes, session["username"],
                         date.today().isoformat(), po_id))
                    db.commit()
                    # Notify the original requestor that finance has resolved the cancellation
                    create_po_notification(db, po["raised_by"], po_id, po["po_number"],
                        f"PO {po['po_number']} cancellation resolved by {session['username']} "
                        f"({res_type}). {res_notes}",
                        'info')
                    db.commit()
                    flash(f"Cancellation marked as resolved: {res_type}.", "success")
            else:
                flash("This PO does not require cancellation resolution.", "warning")

        elif action == "close_partial" and role in ("admin","ops_planner","ops_manager"):
            if po["status"] != "Partially Received":
                flash("Only partially received POs can be closed off.", "warning")
            else:
                close_reason = request.form.get("close_reason","").strip()
                if not close_reason:
                    flash("A reason is required to close off a partial PO.", "danger")
                else:
                    db.execute("""UPDATE purchase_orders SET
                        status='Closed – Partial',
                        partial_close_reason=?, partial_closed_by=?, partial_closed_at=?,
                        updated_at=datetime('now') WHERE id=?""",
                        (close_reason, session["username"],
                         date.today().isoformat(), po_id))
                    db.commit()
                    # Notify ops manager and finance manager about the partial close
                    notify_roles = db.execute(
                        "SELECT username FROM users WHERE role IN ('ops_manager','finance_manager','admin') "
                        "AND active=1"
                    ).fetchall()
                    for nu in notify_roles:
                        if nu["username"] != session["username"]:
                            create_po_notification(db, nu["username"], po_id, po["po_number"],
                                f"PO {po['po_number']} has been closed as a partial receipt by "
                                f"{session['username']}. Reason: {close_reason}",
                                'info')
                    db.commit()
                    flash("PO closed off as partial receipt. Ops and Finance managers have been notified.", "success")

        elif action == "resubmit" and role in ("admin","ops_planner"):
            if po["status"] == "Returned":
                if role != "admin" and po["raised_by"] != session.get("username"):
                    flash("Only the original requestor can resubmit this PO.", "danger")
                else:
                    db.execute("""UPDATE purchase_orders SET status='Pending Approval',
                        updated_at=datetime('now') WHERE id=?""", (po_id,))
                    flash("PO resubmitted for approval.", "success")
            else:
                flash("Only returned POs can be resubmitted.", "danger")

        # ── Finance Exec / Admin actions ─────────────────────────────────────
        elif action == "deposit_paid" and role in ("admin","finance_exec"):
            dep_date = request.form.get("deposit_date", date.today().isoformat())
            db.execute("""UPDATE purchase_orders SET status='Deposit Paid',
                deposit_date=?, updated_at=datetime('now') WHERE id=?""",
                (dep_date, po_id))
            flash("Deposit payment recorded.", "success")

        elif action == "fully_paid" and role in ("admin","finance_exec"):
            pay_date = request.form.get("full_payment_date", date.today().isoformat())
            db.execute("""UPDATE purchase_orders SET status='Fully Paid – Shipped',
                full_payment_date=?, updated_at=datetime('now') WHERE id=?""",
                (pay_date, po_id))
            flash("Full payment recorded. Status set to Fully Paid – Shipped.", "success")

        # ── Ops Exec / Ops Planner / Admin — record goods receipt ──────────────
        elif action == "mark_received" and role in ("admin","ops_exec","ops_planner"):
            RECEIVABLE = ("Approved","Deposit Paid","Fully Paid – Shipped","Partially Received")
            if po["status"] not in RECEIVABLE:
                flash("Goods can only be received once the PO is approved and paid.", "danger")
            else:
                rec_date      = request.form.get("received_date", date.today().isoformat())
                po_lines_all  = db.execute("SELECT * FROM po_lines WHERE po_id=?", (po_id,)).fetchall()

                # Server-side over-receiving validation
                over_errors = []
                for line in po_lines_all:
                    new_recv  = float(request.form.get("recv_" + str(line["id"]), 0) or 0)
                    prev_recv = line["qty_received"] or 0
                    max_allow = (line["qty_ordered"] or 0) + (line["free_units"] or 0)
                    if new_recv < 0:
                        over_errors.append(f"SKU {line['item_id']}: quantity cannot be negative.")
                    elif prev_recv + new_recv > max_allow:
                        over_errors.append(
                            f"SKU {line['item_id']}: receiving {int(new_recv)} would exceed "
                            f"PO qty of {int(max_allow)} (already received {int(prev_recv)}, "
                            f"remaining {int(max_allow - prev_recv)})."
                        )
                if over_errors:
                    for err in over_errors:
                        flash(err, "danger")
                    return redirect(url_for("po_detail", po_id=po_id) + "#receive-section")

                total_subtotal = sum((l["qty_ordered"] or 0)*(l["unit_price"] or 0) for l in po_lines_all)
                freight        = po["freight_costs"] or 0
                other_c        = po["other_costs"]   or 0
                total_overhead = freight + other_c
                n_lines        = len(po_lines_all)
                total_newly_recv = 0
                skus_booked      = 0

                for line in po_lines_all:
                    new_recv = float(request.form.get("recv_" + str(line["id"]), 0) or 0)
                    if new_recv < 0:
                        new_recv = 0
                    prev_recv = line["qty_received"] or 0
                    cum_recv  = prev_recv + new_recv
                    db.execute("UPDATE po_lines SET qty_received=? WHERE id=?", (cum_recv, line["id"]))
                    if new_recv <= 0:
                        continue
                    qty   = line["qty_ordered"] or 0
                    price = line["unit_price"]  or 0
                    free  = line["free_units"]  or 0
                    sub   = qty * price
                    tq    = qty + free
                    alloc = (sub / total_subtotal * total_overhead) if total_subtotal > 0 else (total_overhead / n_lines if n_lines else 0)
                    eff   = (sub + alloc) / tq if tq > 0 else price
                    existing = db.execute("SELECT COUNT(*) FROM fifo_layers").fetchone()[0]
                    layer_id = "FIFO-{:04d}".format(existing + 1)
                    db.execute("INSERT INTO fifo_layers (layer_id,item_id,po_reference,receipt_date,qty_received,unit_cost) VALUES (?,?,?,?,?,?)",
                               (layer_id, line["item_id"], po["po_number"], rec_date, new_recv, eff))
                    txn_id = next_txn_id(db)
                    db.execute("INSERT INTO inventory_ledger (txn_id,txn_date,txn_type,reference,item_id,qty_in,qty_out,unit_cost,posted_by) VALUES (?,?,?,?,?,?,0,?,?)",
                               (txn_id, rec_date, "PO Receipt", po["po_number"], line["item_id"], new_recv, eff, session["username"]))
                    total_newly_recv += new_recv
                    skus_booked += 1

                updated_lines  = db.execute("SELECT qty_ordered, free_units, qty_received FROM po_lines WHERE po_id=?", (po_id,)).fetchall()
                total_ordered  = sum((l["qty_ordered"] or 0) + (l["free_units"] or 0) for l in updated_lines)
                total_cum_recv = sum(l["qty_received"] or 0 for l in updated_lines)

                if total_ordered > 0 and total_cum_recv >= total_ordered:
                    new_status = "Fully Paid – Received"
                    msg = "All goods received ({} units, {} SKU(s)) — inventory updated.".format(int(total_cum_recv), skus_booked)
                    flash_cat = "success"
                else:
                    new_status = "Partially Received"
                    remaining  = max(0, total_ordered - total_cum_recv)
                    msg = "Partial receipt: {} units received this time, {} units still outstanding.".format(int(total_newly_recv), int(remaining))
                    flash_cat = "warning"

                db.execute("""UPDATE purchase_orders
                              SET status=?, received_date=?, received_qty=?,
                                  received_by=COALESCE(received_by, ?),
                                  updated_at=datetime('now')
                              WHERE id=?""",
                           (new_status, rec_date, total_cum_recv, session["username"], po_id))
                flash(msg, flash_cat)

        # ── All roles ────────────────────────────────────────────────────────
        elif action == "update_notes":
            notes = request.form.get("notes","")
            db.execute("UPDATE purchase_orders SET notes=?, updated_at=datetime('now') WHERE id=?",
                       (notes, po_id))
            flash("Notes updated.", "info")

        else:
            flash("Action not permitted for your role.", "danger")

        db.commit()
        return redirect(url_for("po_detail", po_id=po_id))

    # Fetch line items and compute totals
    lines_raw = db.execute("""
        SELECT pl.*, i.name item_name, i.unit_of_measure
        FROM po_lines pl JOIN items i ON pl.item_id=i.item_id
        WHERE pl.po_id=? ORDER BY pl.id
    """, (po_id,)).fetchall()
    freight     = po["freight_costs"] or 0
    other_c     = po["other_costs"]   or 0
    total_sub   = sum((l["qty_ordered"] or 0)*(l["unit_price"] or 0) for l in lines_raw)
    n_lines     = len(lines_raw)
    lines = []
    for l in lines_raw:
        qty   = l["qty_ordered"] or 0
        price = l["unit_price"]  or 0
        free  = l["free_units"]  or 0
        sub   = qty * price
        tq    = qty + free
        alloc = (sub/total_sub*(freight+other_c)) if total_sub > 0 else ((freight+other_c)/n_lines if n_lines else 0)
        eff   = (sub + alloc) / tq if tq > 0 else 0
        qty_recv = l["qty_received"] or 0
        lines.append({"item_id": l["item_id"], "item_name": l["item_name"],
                      "unit_of_measure": l["unit_of_measure"],
                      "line_id": l["id"],
                      "qty_ordered": qty, "unit_price": price, "free_units": free,
                      "qty_received": qty_recv,
                      "qty_remaining": max(0, tq - qty_recv),
                      "subtotal": sub, "total_qty_line": tq, "eff_unit_cost": eff})
    total_qty   = sum(l["total_qty_line"] for l in lines)
    total_val   = total_sub + freight + other_c
    eff_cost    = total_val / total_qty if total_qty > 0 else 0
    deposit_amt = total_val * (po["deposit_pct"] or 0)

    # Fetch attached documents
    po_docs = db.execute("""
        SELECT * FROM po_documents WHERE po_id=? ORDER BY uploaded_at DESC
    """, (po_id,)).fetchall()

    has_tax_invoice = any(d["doc_type"] == "Tax Invoice" for d in po_docs)

    return render_template("po_detail.html", po=po, lines=lines,
                           total_sub=total_sub, total_qty=total_qty,
                           total_val=total_val, eff_cost=eff_cost,
                           deposit_amt=deposit_amt,
                           po_docs=po_docs,
                           has_tax_invoice=has_tax_invoice,
                           today_str=date.today().isoformat())

# ─── INVENTORY LEDGER ────────────────────────────────────────────────────────
@app.route("/ledger")
@login_required
def ledger():
    db = get_db()
    from datetime import date as _date
    # Month filter — default to current month
    cur_month   = _date.today().strftime("%Y-%m")
    month_filter = request.args.get("month", cur_month)
    item_filter  = request.args.get("item", "")

    # Available months (union of ledger entries and stockcount uploads)
    months_raw = db.execute("""
        SELECT DISTINCT substr(txn_date,1,7) AS m FROM inventory_ledger WHERE txn_date IS NOT NULL
        UNION
        SELECT DISTINCT period AS m FROM stockcount_uploads WHERE period IS NOT NULL
        ORDER BY m DESC
    """).fetchall()
    months = [r["m"] for r in months_raw if r["m"]]

    # ── Section 1: Month-end physical count ──────────────────────────────────
    count_upload = db.execute("""
        SELECT * FROM stockcount_uploads
        WHERE period=? ORDER BY uploaded_at DESC LIMIT 1
    """, (month_filter,)).fetchone()

    count_rows = []
    total_count_value = 0.0
    if count_upload:
        # WACC per item from FIFO layers
        wacc_summary = {r["item_id"]: r["wacc"] for r in get_wacc_summary(db)}
        count_lines = db.execute("""
            SELECT sl.item_id, sl.physical_qty, COALESCE(i.name,sl.item_id) item_name
            FROM stockcount_lines sl
            LEFT JOIN items i ON sl.item_id=i.item_id
            WHERE sl.upload_id=? ORDER BY sl.item_id
        """, (count_upload["id"],)).fetchall()
        for line in count_lines:
            wacc  = wacc_summary.get(line["item_id"], 0.0)
            value = (line["physical_qty"] or 0) * wacc
            total_count_value += value
            count_rows.append({
                "item_id":    line["item_id"],
                "item_name":  line["item_name"],
                "physical_qty": line["physical_qty"] or 0,
                "wacc":       wacc,
                "value":      value,
            })

    # ── Section 2: Manual movements for the month ────────────────────────────
    query  = """SELECT l.*, COALESCE(i.name, l.item_id) item_name
                FROM inventory_ledger l
                LEFT JOIN items i ON l.item_id=i.item_id
                WHERE l.txn_date LIKE ?"""
    params = [f"{month_filter}%"]
    if item_filter:
        query += " AND l.item_id=?"
        params.append(item_filter)
    query += " ORDER BY l.txn_date DESC, l.id DESC"

    entries = db.execute(query, params).fetchall()
    items   = db.execute("SELECT item_id, name FROM items WHERE status='Active' ORDER BY item_id").fetchall()
    txn_types = [
        "PO Receipt","(+) Inbound",
        "(+) Adjustment – Stockcount","(+) Adjustment – Others",
        "(-) Sale","(-) Adjustment – Stockcount","(-) Adjustment – Others",
        "Sale","Shrinkage","Adjustment – In","Adjustment – Out","Write-off","Opening Balance",
    ]
    return render_template("ledger.html",
        entries=entries, items=items,
        txn_types=txn_types, item_filter=item_filter,
        month_filter=month_filter, months=months,
        count_upload=count_upload, count_rows=count_rows,
        total_count_value=total_count_value)

@app.route("/ledger/new", methods=["GET","POST"])
@login_required
@roles_required("admin","ops_exec")
def ledger_new():
    db = get_db()
    if request.method == "POST":
        f        = request.form
        txn_id   = next_txn_id(db)
        txn_type = f["txn_type"]
        item_id  = f["item_id"]
        qty_in   = float(f.get("qty_in") or 0)
        qty_out  = float(f.get("qty_out") or 0)
        unit_cost= float(f.get("unit_cost") or 0)
        try:
            db.execute("""INSERT INTO inventory_ledger
                (txn_id,txn_date,txn_type,reference,item_id,qty_in,qty_out,
                 unit_cost,posted_by,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (txn_id, f["txn_date"], txn_type, f.get("reference",""),
                 item_id, qty_in, qty_out, unit_cost,
                 session["username"], f.get("notes","")))

            # Auto-create FIFO layer for inbound transactions (qty increases inventory)
            IN_TYPES_FIFO = {"PO Receipt", "(+) Inbound", "(+) Adjustment – Stockcount", "(+) Adjustment – Others"}
            if txn_type in IN_TYPES_FIFO and qty_in > 0:
                existing = db.execute("SELECT COUNT(*) FROM fifo_layers").fetchone()[0]
                layer_id = f"L-{existing+1:03d}"
                db.execute("""INSERT INTO fifo_layers
                    (layer_id,item_id,po_reference,receipt_date,qty_received,unit_cost)
                    VALUES (?,?,?,?,?,?)""",
                    (layer_id, item_id, f.get("reference",""), f["txn_date"], qty_in, unit_cost))

            db.commit()
            flash(f"Transaction {txn_id} recorded successfully.", "success")
            return redirect(url_for("ledger"))
        except Exception as e:
            flash(f"Error recording transaction: {e}", "danger")

    items    = db.execute("SELECT item_id, name FROM items WHERE status='Active' ORDER BY item_id").fetchall()
    # Manual-entry types only (PO Receipt is system-generated via goods receipt flow)
    txn_types = [
        "(+) Inbound",
        "(+) Adjustment – Stockcount",
        "(+) Adjustment – Others",
        "(-) Sale",
        "(-) Adjustment – Stockcount",
        "(-) Adjustment – Others",
    ]
    today    = date.today().isoformat()
    return render_template("ledger_new.html", items=items, txn_types=txn_types, today=today)

# ─── USER MANAGEMENT ────────────────────────────────────────────────────────
@app.route("/settings/users")
@login_required
@roles_required("admin")
def user_list():
    db    = get_db()
    users = db.execute("SELECT * FROM users ORDER BY role, full_name").fetchall()
    return render_template("users.html", users=users)

@app.route("/settings/users/add", methods=["POST"])
@login_required
@roles_required("admin")
def user_add():
    from werkzeug.security import generate_password_hash
    f  = request.form
    db = get_db()
    username = f["username"].strip().lower()
    # Check for duplicate username
    if db.execute("SELECT id FROM users WHERE username=?", (username,)).fetchone():
        flash(f"Username '{username}' already exists. Choose a different one.", "danger")
        return redirect(url_for("user_list"))
    try:
        db.execute(
            "INSERT INTO users (username, password, full_name, role) VALUES (?,?,?,?)",
            (username, generate_password_hash(f["password"]), f["full_name"].strip(), f["role"]))
        db.commit()
        flash(f"User '{f['full_name']}' added successfully.", "success")
    except Exception as e:
        flash(f"Error adding user: {e}", "danger")
    return redirect(url_for("user_list"))

@app.route("/settings/users/edit", methods=["POST"])
@login_required
@roles_required("admin")
def user_edit():
    from werkzeug.security import generate_password_hash
    f            = request.form
    user_id      = int(f["user_id"])
    db           = get_db()
    new_password = f.get("password","").strip()
    new_username = f.get("username","").strip().lower()
    # Duplicate username guard — exclude current user from check
    clash = db.execute(
        "SELECT id FROM users WHERE username=? AND id!=?", (new_username, user_id)).fetchone()
    if clash:
        flash(f"Username '{new_username}' is already taken. Choose a different one.", "danger")
        return redirect(url_for("user_list"))
    try:
        if new_password:
            db.execute(
                "UPDATE users SET full_name=?, role=?, username=?, password=? WHERE id=?",
                (f["full_name"].strip(), f["role"], new_username,
                 generate_password_hash(new_password), user_id))
        else:
            db.execute(
                "UPDATE users SET full_name=?, role=?, username=? WHERE id=?",
                (f["full_name"].strip(), f["role"], new_username, user_id))
        db.commit()
        # Refresh session if editing own account
        if user_id == session["user_id"]:
            session["full_name"] = f["full_name"].strip()
            session["role"]      = f["role"]
            session["username"]  = new_username
        flash("User updated successfully.", "success")
    except Exception as e:
        flash(f"Error updating user: {e}", "danger")
    return redirect(url_for("user_list"))

@app.route("/settings/users/<int:user_id>/toggle", methods=["POST"])
@login_required
@roles_required("admin")
def user_toggle(user_id):
    if user_id == session["user_id"]:
        flash("You cannot deactivate your own account.", "danger")
        return redirect(url_for("user_list"))
    db   = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if user:
        new_status = 0 if user["active"] else 1
        db.execute("UPDATE users SET active=? WHERE id=?", (new_status, user_id))
        db.commit()
        action = "activated" if new_status else "deactivated"
        flash(f"User '{user['full_name']}' {action}.", "success")
    return redirect(url_for("user_list"))

@app.route("/settings/users/<int:user_id>/delete", methods=["POST"])
@login_required
@roles_required("admin")
def user_delete(user_id):
    if user_id == session["user_id"]:
        flash("You cannot delete your own account.", "danger")
        return redirect(url_for("user_list"))
    db   = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        flash("User not found.", "danger")
        return redirect(url_for("user_list"))
    # Prevent deleting the last active admin
    if user["role"] == "admin":
        admin_count = db.execute(
            "SELECT COUNT(*) FROM users WHERE role='admin' AND active=1").fetchone()[0]
        if admin_count <= 1:
            flash("Cannot delete the last active Admin account.", "danger")
            return redirect(url_for("user_list"))
    db.execute("DELETE FROM users WHERE id=?", (user_id,))
    db.commit()
    flash(f"User '{user['full_name']}' deleted.", "success")
    return redirect(url_for("user_list"))


@app.route("/settings/users/<int:user_id>/photo", methods=["POST"])
@login_required
def user_photo_upload(user_id):
    """Any user can update their own photo; admins can update any user's photo."""
    if session["user_id"] != user_id and session["role"] != "admin":
        flash("Permission denied.", "danger")
        return redirect(url_for("user_list"))
    db   = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if not user:
        flash("User not found.", "danger")
        return redirect(url_for("user_list"))
    f = request.files.get("photo")
    if not f or not f.filename:
        flash("No file selected.", "danger")
        return redirect(url_for("user_list"))
    allowed = {"png", "jpg", "jpeg", "gif", "webp"}
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in allowed:
        flash("Invalid file type. Please upload a PNG, JPG, GIF, or WEBP image.", "danger")
        return redirect(url_for("user_list"))
    import os as _osp, time as _t
    upload_dir = _osp.join(_osp.dirname(__file__), "static", "uploads", "profiles")
    _osp.makedirs(upload_dir, exist_ok=True)
    # Remove old photo if present
    if user["profile_photo"]:
        old_path = _osp.join(upload_dir, user["profile_photo"])
        if _osp.exists(old_path):
            _osp.remove(old_path)
    filename = f"user_{user_id}_{int(_t.time())}.{ext}"
    f.save(_osp.join(upload_dir, filename))
    db.execute("UPDATE users SET profile_photo=? WHERE id=?", (filename, user_id))
    db.commit()
    # Refresh session photo if updating own account
    if session["user_id"] == user_id:
        session["profile_photo"] = filename
    flash("Profile photo updated.", "success")
    return redirect(url_for("user_list"))


@app.route("/settings/users/<int:user_id>/photo/delete", methods=["POST"])
@login_required
def user_photo_delete(user_id):
    if session["user_id"] != user_id and session["role"] != "admin":
        flash("Permission denied.", "danger")
        return redirect(url_for("user_list"))
    db   = get_db()
    user = db.execute("SELECT * FROM users WHERE id=?", (user_id,)).fetchone()
    if user and user["profile_photo"]:
        import os as _osp2
        old_path = _osp2.join(_osp2.path.dirname(__file__), "static", "uploads", "profiles", user["profile_photo"])
        if _osp2.path.exists(old_path):
            _osp2.remove(old_path)
        db.execute("UPDATE users SET profile_photo=NULL WHERE id=?", (user_id,))
        db.commit()
        if session["user_id"] == user_id:
            session.pop("profile_photo", None)
        flash("Profile photo removed.", "success")
    return redirect(url_for("user_list"))



# ── APPROVALS DASHBOARD ───────────────────────────────────────────────────────
@app.route("/approvals")
@login_required
@roles_required("admin", "ops_manager", "finance_manager")
def approvals():
    db   = get_db()
    role = session["role"]

    # POs awaiting Ops Manager approval
    pos_pending_ops = []
    if role in ("admin", "ops_manager"):
        pos_pending_ops = db.execute("""
            SELECT po.*, s.name supplier_name,
                (SELECT COUNT(*) FROM po_lines WHERE po_id=po.id) AS line_count,
                (SELECT i2.name FROM po_lines pl2 JOIN items i2 ON pl2.item_id=i2.item_id
                 WHERE pl2.po_id=po.id ORDER BY pl2.id LIMIT 1) AS first_item_name,
                (SELECT COALESCE(SUM(pl.qty_ordered * pl.unit_price), 0)
                 FROM po_lines pl WHERE pl.po_id=po.id)
                 + COALESCE(po.freight_costs, 0) + COALESCE(po.other_costs, 0) AS total_val
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            WHERE po.status = 'Pending Approval'
            ORDER BY po.created_at ASC
        """).fetchall()

    # POs awaiting Finance Manager approval
    pos_pending_fin = []
    if role in ("admin", "finance_manager"):
        pos_pending_fin = db.execute("""
            SELECT po.*, s.name supplier_name,
                (SELECT COUNT(*) FROM po_lines WHERE po_id=po.id) AS line_count,
                (SELECT i2.name FROM po_lines pl2 JOIN items i2 ON pl2.item_id=i2.item_id
                 WHERE pl2.po_id=po.id ORDER BY pl2.id LIMIT 1) AS first_item_name,
                (SELECT COALESCE(SUM(pl.qty_ordered * pl.unit_price), 0)
                 FROM po_lines pl WHERE pl.po_id=po.id)
                 + COALESCE(po.freight_costs, 0) + COALESCE(po.other_costs, 0) AS total_val
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            WHERE po.status = 'Ops Approved'
            ORDER BY po.created_at ASC
        """).fetchall()

    # Supplier requests awaiting Ops Manager approval
    sr_pending_ops = []
    if role in ("admin", "ops_manager"):
        sr_pending_ops = db.execute("""
            SELECT * FROM supplier_requests
            WHERE status = 'Pending'
            ORDER BY created_at ASC
        """).fetchall()

    # Supplier requests awaiting Finance Manager approval
    sr_pending_fin = []
    if role in ("admin", "finance_manager"):
        sr_pending_fin = db.execute("""
            SELECT * FROM supplier_requests
            WHERE status = 'Ops Approved'
            ORDER BY created_at ASC
        """).fetchall()

    total_pending = (len(pos_pending_ops) + len(pos_pending_fin) +
                     len(sr_pending_ops)  + len(sr_pending_fin))

    return render_template("approvals.html",
        pos_pending_ops=pos_pending_ops, pos_pending_fin=pos_pending_fin,
        sr_pending_ops=sr_pending_ops,  sr_pending_fin=sr_pending_fin,
        total_pending=total_pending, role=role)

# ── FINANCE MODULE ────────────────────────────────────────────────────────────
@app.route("/finance")
@login_required
@roles_required("admin", "finance_manager", "finance_exec")
def finance_summary():
    db = get_db()
    # Date range — default to current month
    today     = date.today()
    date_from = request.args.get("date_from", today.replace(day=1).isoformat())
    date_to   = request.args.get("date_to",   today.isoformat())

    # 1. Goods Received this period (Dr Inventory / Cr AP) — one row per line item
    goods_received = db.execute("""
        SELECT po.po_number, po.received_date, po.supplier_id, s.name supplier_name,
               pl.item_id, i.name item_name,
               (pl.qty_ordered + pl.free_units) AS received_qty,
               pl.unit_price,
               po.freight_costs, po.other_costs,
               ROUND(
                 CASE WHEN ts.subtotal > 0 THEN
                   (pl.qty_ordered * pl.unit_price / ts.subtotal)
                   * (ts.subtotal + COALESCE(po.freight_costs,0) + COALESCE(po.other_costs,0))
                 ELSE pl.qty_ordered * pl.unit_price END, 2
               ) AS inventory_value
        FROM purchase_orders po
        JOIN po_lines pl ON pl.po_id = po.id
        JOIN items i ON pl.item_id = i.item_id
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        JOIN (SELECT po_id, SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
          ON ts.po_id = po.id
        WHERE po.received_date BETWEEN ? AND ?
          AND po.status IN ('Fully Paid – Received')
        ORDER BY po.received_date DESC, pl.id
    """, (date_from, date_to)).fetchall()

    # 2. Payments made this period (Dr AP / Cr Bank)
    # Payment rows split by event date so cross-month payments land in the right period.
    # deposit_amount_paid > 0 means we have a clean split; show each event on its own date.
    # deposit_amount_paid = 0 means legacy data — the deposit wasn't captured separately:
    #   • If full_payment_date IS NULL  → only deposit paid, amount = amount_paid (recovered by migration)
    #   • If full_payment_date IS NOT NULL AND deposit_date = full_payment_date → single-event full pay
    #   • If full_payment_date IS NOT NULL AND deposit_date ≠ full_payment_date → legacy two-payment,
    #     deposit amount unknown; suppress deposit row, show full amount on full_payment_date.
    deposits_paid = db.execute("""
        SELECT po.po_number, po.deposit_date AS pay_date,
               CASE WHEN po.deposit_amount_paid > 0 THEN 'Deposit'
                    ELSE 'Deposit' END AS pay_type,
               po.supplier_id, s.name supplier_name, po.currency,
               CASE WHEN po.deposit_amount_paid > 0 THEN po.deposit_amount_paid
                    ELSE po.amount_paid END AS amount
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        WHERE po.deposit_date BETWEEN ? AND ?
          AND po.status IN ('Deposit Paid','Fully Paid – Shipped','Fully Paid – Received','Cancelled')
          AND (
               po.deposit_amount_paid > 0
            OR (po.full_payment_date IS NULL AND po.amount_paid > 0)
            OR (po.full_payment_date IS NOT NULL AND po.deposit_date = po.full_payment_date)
          )
        ORDER BY po.deposit_date DESC
    """, (date_from, date_to)).fetchall()

    # Full payment rows — show the balance owed after the deposit.
    # For single-event full pays (deposit_date = full_payment_date) the deposit row above
    # already covers the whole amount, so exclude those here to avoid double-counting.
    # For legacy two-payment records (deposit_amount_paid = 0, dates differ) we show the
    # full amount_paid here because the deposit portion can't be recovered.
    full_payments = db.execute("""
        SELECT po.po_number, po.full_payment_date AS pay_date,
               CASE WHEN po.deposit_amount_paid > 0 THEN 'Full Payment'
                    ELSE 'Full Payment (combined)' END AS pay_type,
               po.supplier_id, s.name supplier_name, po.currency,
               CASE WHEN po.deposit_amount_paid > 0
                    THEN ROUND(
                           (COALESCE(ts.subtotal,0) + COALESCE(po.freight_costs,0) + COALESCE(po.other_costs,0))
                           - po.deposit_amount_paid, 2)
                    ELSE COALESCE(po.amount_paid, 0)
               END AS amount
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        LEFT JOIN (SELECT po_id, SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
          ON ts.po_id = po.id
        WHERE po.full_payment_date BETWEEN ? AND ?
          AND po.status IN ('Fully Paid – Shipped','Fully Paid – Received','Cancelled')
          AND po.full_payment_date != po.deposit_date
        ORDER BY po.full_payment_date DESC
    """, (date_from, date_to)).fetchall()

    payments = list(deposits_paid) + list(full_payments)
    payments.sort(key=lambda r: r["pay_date"] or "", reverse=True)

    # 3. Inventory movements / COGS this period (Dr COGS / Cr Inventory for outbound)
    movements = db.execute("""
        SELECT l.txn_id, l.txn_date, l.txn_type, l.reference,
               l.item_id, i.name item_name,
               l.qty_in, l.qty_out, l.unit_cost,
               ROUND(l.qty_in  * l.unit_cost, 2) AS value_in,
               ROUND(l.qty_out * l.unit_cost, 2) AS value_out,
               l.posted_by
        FROM inventory_ledger l
        JOIN items i ON l.item_id = i.item_id
        WHERE l.txn_date BETWEEN ? AND ?
        ORDER BY l.txn_date DESC, l.id DESC
    """, (date_from, date_to)).fetchall()

    # 4. Outstanding AP — approved POs not yet fully paid
    outstanding_ap = db.execute("""
        SELECT po.id, po.po_number, po.po_date, po.supplier_id, s.name supplier_name,
               po.status, po.currency, po.payment_due_date,
               COALESCE(po.amount_paid, 0) AS amount_paid,
               ROUND(
                 (COALESCE(ts.subtotal,0) + COALESCE(po.freight_costs,0) + COALESCE(po.other_costs,0))
               , 2) AS total_value
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        LEFT JOIN (SELECT po_id, SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
          ON ts.po_id = po.id
        WHERE po.status IN ('Approved','Deposit Paid','Fully Paid – Shipped','Partially Received')
          AND po.status != 'Closed – Partial'
        ORDER BY po.payment_due_date ASC NULLS LAST, po.po_date ASC
    """).fetchall()

    # Summary totals
    total_received_value = sum(r["inventory_value"] or 0 for r in goods_received)
    total_payments       = sum(r["amount"] or 0 for r in payments)
    COGS_TYPES     = {"Sale","Shrinkage","Write-off","Adjustment – Out",
                      "(-) Sale","(-) Adjustment – Stockcount","(-) Adjustment – Others"}
    RECEIPT_TYPES  = {"PO Receipt","(+) Inbound","(+) Adjustment – Stockcount","(+) Adjustment – Others",
                      "Adjustment – In","Opening Balance"}
    total_cogs           = sum(r["value_out"] or 0 for r in movements if r["txn_type"] in COGS_TYPES)
    total_receipts_value = sum(r["value_in"]  or 0 for r in movements if r["txn_type"] in RECEIPT_TYPES)
    total_ap             = sum(r["total_value"] or 0 for r in outstanding_ap)

    # Payment tracking buckets
    today_str = today.isoformat()
    pay_overdue = db.execute("""
        SELECT po.id, po.po_number, po.payment_due_date, po.supplier_id,
               s.name supplier_name, po.status, po.currency,
               COALESCE(po.amount_paid,0) AS amount_paid,
               ROUND((COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0)),2) AS total_val
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id=s.supplier_id
        LEFT JOIN (SELECT po_id,SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
          ON ts.po_id=po.id
        WHERE po.status IN ('Approved','Deposit Paid')
          AND po.payment_due_date IS NOT NULL
          AND po.payment_due_date < ?
        ORDER BY po.payment_due_date ASC""", (today_str,)).fetchall()

    pay_due_soon = db.execute("""
        SELECT po.id, po.po_number, po.payment_due_date, po.supplier_id,
               s.name supplier_name, po.status, po.currency,
               COALESCE(po.amount_paid,0) AS amount_paid,
               ROUND((COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0)),2) AS total_val
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id=s.supplier_id
        LEFT JOIN (SELECT po_id,SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
          ON ts.po_id=po.id
        WHERE po.status IN ('Approved','Deposit Paid')
          AND po.payment_due_date IS NOT NULL
          AND po.payment_due_date BETWEEN ? AND date(?, '+30 days')
        ORDER BY po.payment_due_date ASC""", (today_str, today_str)).fetchall()

    pay_no_date = db.execute("""
        SELECT po.id, po.po_number, po.payment_due_date, po.supplier_id,
               s.name supplier_name, po.status, po.currency,
               COALESCE(po.amount_paid,0) AS amount_paid,
               ROUND((COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0)),2) AS total_val
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id=s.supplier_id
        LEFT JOIN (SELECT po_id,SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
          ON ts.po_id=po.id
        WHERE po.status IN ('Approved','Deposit Paid')
          AND (po.payment_due_date IS NULL OR po.payment_due_date='')
        ORDER BY po.po_date ASC""").fetchall()

    # 5. Cancelled POs with unresolved payments
    pending_resolution = db.execute("""
        SELECT po.id, po.po_number, po.po_date, po.cancelled_at,
               po.cancelled_by, po.cancel_reason,
               po.amount_paid, po.currency,
               s.name supplier_name
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        WHERE po.status = 'Cancelled'
          AND po.cancellation_resolution = 'Pending'
          AND (po.amount_paid IS NOT NULL AND po.amount_paid > 0)
        ORDER BY po.cancelled_at DESC
    """).fetchall()

    return render_template("finance.html",
        date_from=date_from, date_to=date_to,
        goods_received=goods_received, payments=payments,
        movements=movements, outstanding_ap=outstanding_ap,
        total_received_value=total_received_value,
        total_payments=total_payments,
        total_cogs=total_cogs,
        total_receipts_value=total_receipts_value,
        total_ap=total_ap,
        pay_overdue=pay_overdue, pay_due_soon=pay_due_soon,
        pay_no_date=pay_no_date, today=today_str,
        pending_resolution=pending_resolution)

# ── ITEM / SKU MANAGEMENT ────────────────────────────────────────────────────
ITEM_ROLES = ("admin", "ops_manager", "ops_planner", "ops_exec")

def next_item_id(db):
    last = db.execute(
        "SELECT item_id FROM items ORDER BY item_id DESC LIMIT 1"
    ).fetchone()
    if last:
        try:
            num = int(last["item_id"].replace("I", "")) + 1
        except Exception:
            num = 1
        return f"I{num:03d}"
    return "I001"


@app.route("/items")
@login_required
def item_list():
    db = get_db()
    items = db.execute("""
        SELECT i.*, s.name supplier_name
        FROM items i
        LEFT JOIN suppliers s ON i.preferred_supplier = s.supplier_id
        ORDER BY i.item_id
    """).fetchall()
    suppliers = db.execute(
        "SELECT supplier_id, name FROM suppliers WHERE status='Active' ORDER BY name"
    ).fetchall()
    categories = db.execute(
        "SELECT DISTINCT category FROM items WHERE category IS NOT NULL ORDER BY category"
    ).fetchall()
    role = session["role"]
    can_manage = role in ITEM_ROLES
    return render_template("items.html", items=items, suppliers=suppliers,
                           categories=[r["category"] for r in categories],
                           can_manage=can_manage)


@app.route("/items/add", methods=["POST"])
@login_required
@roles_required(*ITEM_ROLES)
def item_add():
    f  = request.form
    db = get_db()
    # Use user-supplied ID if provided, else auto-generate
    item_id = f.get("item_id", "").strip().upper()
    if not item_id:
        item_id = next_item_id(db)
    # Check uniqueness
    if db.execute("SELECT id FROM items WHERE item_id=?", (item_id,)).fetchone():
        flash(f"Item ID '{item_id}' already exists. Choose a different one.", "danger")
        return redirect(request.form.get("redirect_to", "") or url_for("item_list"))
    supplier_id = f.get("preferred_supplier", "").strip() or None
    unit_price  = float(f.get("unit_price") or 0)
    try:
        db.execute("""INSERT INTO items
            (item_id, name, description, category, unit_of_measure,
             unit_price, reorder_point, lead_days, preferred_supplier, status, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (item_id,
             f["name"].strip(),
             f.get("description", "").strip() or None,
             f.get("category", "").strip() or None,
             f.get("unit_of_measure", "Units").strip(),
             unit_price,
             int(f.get("reorder_point") or 0),
             int(f.get("lead_days") or 14),
             supplier_id,
             "Active",
             f.get("notes", "").strip() or None))
        # Auto-link to supplier; record supplier-specific price
        if supplier_id:
            db.execute(
                """INSERT OR IGNORE INTO supplier_items (supplier_id, item_id, unit_price, created_by)
                   VALUES (?,?,?,?)""",
                (supplier_id, item_id, unit_price if unit_price else None, session["username"]))
        db.commit()
        redirect_to = f.get("redirect_to", "")
        flash(f"Item {item_id} — '{f['name']}' created successfully.", "success")
        if redirect_to:
            return redirect(redirect_to)
        return redirect(url_for("item_list"))
    except Exception as e:
        flash(f"Error creating item: {e}", "danger")
        return redirect(request.form.get("redirect_to", "") or url_for("item_list"))


@app.route("/items/<item_id>/edit", methods=["POST"])
@login_required
@roles_required(*ITEM_ROLES)
def item_edit(item_id):
    f  = request.form
    db = get_db()
    old = db.execute("SELECT * FROM items WHERE item_id=?", (item_id,)).fetchone()
    if not old:
        flash("Item not found.", "danger")
        return redirect(url_for("item_list"))
    new_supplier = f.get("preferred_supplier", "").strip() or None
    unit_price   = float(f.get("unit_price") or 0)
    try:
        db.execute("""UPDATE items SET
            name=?, description=?, category=?, unit_of_measure=?,
            unit_price=?, reorder_point=?, lead_days=?, preferred_supplier=?, notes=?
            WHERE item_id=?""",
            (f["name"].strip(),
             f.get("description", "").strip() or None,
             f.get("category", "").strip() or None,
             f.get("unit_of_measure", "Units").strip(),
             unit_price,
             int(f.get("reorder_point") or 0),
             int(f.get("lead_days") or 14),
             new_supplier,
             f.get("notes", "").strip() or None,
             item_id))
        # Sync supplier_items: remove old link if supplier changed, add/update new
        old_supplier = old["preferred_supplier"]
        if old_supplier and old_supplier != new_supplier:
            db.execute("DELETE FROM supplier_items WHERE supplier_id=? AND item_id=?",
                       (old_supplier, item_id))
        if new_supplier:
            db.execute(
                """INSERT INTO supplier_items (supplier_id, item_id, unit_price, created_by)
                   VALUES (?,?,?,?)
                   ON CONFLICT(supplier_id, item_id) DO UPDATE SET unit_price=excluded.unit_price""",
                (new_supplier, item_id, unit_price if unit_price else None, session["username"]))
        db.commit()
        redirect_to = f.get("redirect_to", "")
        flash(f"Item {item_id} updated.", "success")
        if redirect_to:
            return redirect(redirect_to)
        return redirect(url_for("item_list"))
    except Exception as e:
        flash(f"Error updating item: {e}", "danger")
        return redirect(url_for("item_list"))


@app.route("/items/<item_id>/toggle", methods=["POST"])
@login_required
@roles_required(*ITEM_ROLES)
def item_toggle(item_id):
    db   = get_db()
    item = db.execute("SELECT * FROM items WHERE item_id=?", (item_id,)).fetchone()
    if item:
        new_status = "Inactive" if item["status"] == "Active" else "Active"
        db.execute("UPDATE items SET status=? WHERE item_id=?", (new_status, item_id))
        db.commit()
        flash(f"Item {item_id} {new_status.lower()}.", "success")
    redirect_to = request.form.get("redirect_to", "")
    return redirect(redirect_to or url_for("item_list"))


@app.route("/items/<item_id>/delete", methods=["POST"])
@login_required
@roles_required(*ITEM_ROLES)
def item_delete(item_id):
    db = get_db()
    item = db.execute("SELECT * FROM items WHERE item_id=?", (item_id,)).fetchone()
    if not item:
        flash("Item not found.", "danger")
        return redirect(url_for("item_list"))
    # Block deletion if item appears in any PO or ledger entry
    po_count  = db.execute("SELECT COUNT(*) FROM purchase_orders WHERE item_id=?", (item_id,)).fetchone()[0]
    led_count = db.execute("SELECT COUNT(*) FROM inventory_ledger WHERE item_id=?", (item_id,)).fetchone()[0]
    if po_count > 0 or led_count > 0:
        flash(f"Cannot delete {item_id} — it has {po_count} PO(s) and {led_count} ledger entry(ies). Deactivate it instead.", "danger")
        return redirect(request.form.get("redirect_to", "") or url_for("item_list"))
    db.execute("DELETE FROM supplier_items WHERE item_id=?", (item_id,))
    db.execute("DELETE FROM items WHERE item_id=?", (item_id,))
    db.commit()
    flash(f"Item {item_id} — '{item['name']}' deleted.", "success")
    redirect_to = request.form.get("redirect_to", "")
    return redirect(redirect_to or url_for("item_list"))

# ── MASTER DATA CLEAR ────────────────────────────────────────────────────────
@app.route("/items/clear", methods=["POST"])
@login_required
@roles_required("admin")
def items_clear():
    """Delete ALL items, bundles, and bundle components (admin only).
    Blocked if any items have PO or ledger history."""
    db = get_db()
    # Check for any items with PO or ledger history
    po_count  = db.execute("SELECT COUNT(*) FROM po_lines").fetchone()[0]
    led_count = db.execute("SELECT COUNT(*) FROM inventory_ledger").fetchone()[0]
    if po_count > 0 or led_count > 0:
        flash(f"Cannot clear items — there are {po_count} PO line(s) and "
              f"{led_count} ledger entry(ies) referencing existing items. "
              "Clear PO and ledger data first, or deactivate items instead.", "danger")
        return redirect(url_for("item_list"))
    db.execute("DELETE FROM bundle_components")
    db.execute("DELETE FROM bundles")
    db.execute("DELETE FROM supplier_items")
    db.execute("DELETE FROM items")
    db.commit()
    flash("All items, bundles, and supplier-item links have been cleared.", "success")
    return redirect(url_for("master_data_import"))


# ── MASTER DATA IMPORT ───────────────────────────────────────────────────────
@app.route("/master-data/import", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def master_data_import():
    if request.method == "GET":
        return render_template("master_data_import.html")

    # ── POST: parse the uploaded Excel file ──────────────────────────────────
    f = request.files.get("master_file")
    if not f or not f.filename:
        flash("No file selected.", "danger")
        return render_template("master_data_import.html")

    ext = f.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "xls"):
        flash("Only .xlsx / .xls files are accepted.", "danger")
        return render_template("master_data_import.html")

    # openpyxl required
    try:
        import openpyxl as _xl
    except ImportError:
        flash("openpyxl is required. Run: pip install openpyxl then restart.", "danger")
        return render_template("master_data_import.html")

    import io, tempfile, os as _os
    tmp_path = None
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}")
        f.save(tmp.name)
        tmp.close()
        tmp_path = tmp.name

        wb = _xl.load_workbook(tmp_path, data_only=True)
        sheet_names = wb.sheetnames

        items_created   = 0
        items_skipped   = 0
        bundles_created = 0
        bundles_skipped = 0
        errors          = []

        db = get_db()
        user = session.get("username", "system")
        now  = datetime.utcnow().isoformat(timespec="seconds")

        # ── Helper: upsert an item ────────────────────────────────────────────
        def upsert_item(sku_id, sku_name, cost, brand_name):
            nonlocal items_created, items_skipped
            if not sku_id:
                items_skipped += 1
                return
            existing = db.execute("SELECT item_id FROM items WHERE item_id=?",
                                   (sku_id,)).fetchone()
            if existing:
                db.execute(
                    "UPDATE items SET name=?, unit_price=?, brand=?, status='Active' WHERE item_id=?",
                    (sku_name or sku_id, cost or 0, brand_name, sku_id))
                items_created += 1   # counts upserts as created/updated
            else:
                db.execute(
                    """INSERT INTO items
                       (item_id, name, unit_price, brand, status, unit_of_measure, reorder_point)
                       VALUES (?,?,?,?,'Active','EA',0)""",
                    (sku_id, sku_name or sku_id, cost or 0, brand_name))
                items_created += 1

        # ── Parse AS tab (andSons) ────────────────────────────────────────────
        if "AS" in sheet_names:
            ws = wb["AS"]
            rows = list(ws.values)
            # Data starts at row index 7 (row 8 in Excel, 0-indexed = 7)
            for r in rows[7:]:
                if not r or len(r) < 2:
                    items_skipped += 1
                    continue
                sku_id   = str(r[1]).strip() if r[1] is not None else ""
                cost_val = r[4] if len(r) > 4 else None
                remark   = str(r[5]).strip().upper() if len(r) > 5 and r[5] else ""
                if not sku_id or sku_id.upper() in ("NONE", "N/A", "SKU", "PRODUCT SKU"):
                    items_skipped += 1
                    continue
                if remark == "DISCON":
                    items_skipped += 1
                    continue
                try:
                    cost = float(cost_val) if cost_val is not None else 0.0
                except (ValueError, TypeError):
                    cost = 0.0
                upsert_item(sku_id, sku_id, cost, "andSons")
        else:
            errors.append("Tab 'AS' not found in workbook.")

        # ── Parse OVA tab (Ova) ──────────────────────────────────────────────
        if "OVA" in sheet_names:
            ws = wb["OVA"]
            rows = list(ws.values)
            for r in rows[7:]:
                if not r or len(r) < 2:
                    items_skipped += 1
                    continue
                sku_id   = str(r[1]).strip() if r[1] is not None else ""
                cost_val = r[4] if len(r) > 4 else None
                remark   = str(r[5]).strip().upper() if len(r) > 5 and r[5] else ""
                if not sku_id or sku_id.upper() in ("NONE", "N/A", "SKU", "PRODUCT SKU"):
                    items_skipped += 1
                    continue
                if remark == "DISCON":
                    items_skipped += 1
                    continue
                try:
                    cost = float(cost_val) if cost_val is not None else 0.0
                except (ValueError, TypeError):
                    cost = 0.0
                upsert_item(sku_id, sku_id, cost, "Ova")
        else:
            errors.append("Tab 'OVA' not found in workbook.")

        # ── Parse &Sons (Update) tab (Bundles) ────────────────────────────────
        BUNDLE_TAB = "&Sons (Update)"
        if BUNDLE_TAB in sheet_names:
            ws = wb[BUNDLE_TAB]
            rows = list(ws.values)
            # Skip header row (row 0)
            SKU_COLS = [9, 11, 13, 15, 17]   # cols J,L,N,P,R (0-indexed)
            for row_i, r in enumerate(rows[1:], start=2):
                if not r or len(r) < 2:
                    bundles_skipped += 1
                    continue
                raw_id = str(r[0]).strip() if r[0] is not None else ""
                if not raw_id or raw_id.upper() in ("NONE", "N/A"):
                    bundles_skipped += 1
                    continue
                bundle_id = raw_id.split("/")[0].strip()
                if not bundle_id:
                    bundles_skipped += 1
                    continue

                # Skip if no component SKUs present
                component_skus = []
                for ci in SKU_COLS:
                    if ci < len(r) and r[ci] is not None:
                        s = str(r[ci]).strip()
                        if s and s.upper() not in ("NONE", "N/A", ""):
                            component_skus.append(s)
                if not component_skus:
                    bundles_skipped += 1
                    continue

                brand_val = str(r[8]).strip() if len(r) > 8 and r[8] is not None else ""
                desc_val  = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""

                # Upsert bundle
                existing_b = db.execute(
                    "SELECT id FROM bundles WHERE bundle_id=?", (bundle_id,)).fetchone()
                if existing_b:
                    db.execute(
                        "UPDATE bundles SET name=?, description=?, status='Active' WHERE bundle_id=?",
                        (bundle_id, desc_val or bundle_id, bundle_id))
                else:
                    db.execute(
                        """INSERT INTO bundles (bundle_id, name, description, status, created_by, created_at)
                           VALUES (?,?,?,'Active',?,?)""",
                        (bundle_id, bundle_id, desc_val or bundle_id, user, now))
                    bundles_created += 1

                # Upsert components (qty=1 for all, as Excel doesn't store qty)
                for sku in component_skus:
                    try:
                        db.execute(
                            """INSERT OR IGNORE INTO bundle_components (bundle_id, item_id, qty)
                               VALUES (?,?,1)""",
                            (bundle_id, sku))
                    except Exception as ex:
                        errors.append(f"Bundle {bundle_id} component {sku}: {ex}")
        else:
            errors.append(f"Tab '{BUNDLE_TAB}' not found in workbook.")

        db.commit()

    except Exception as e:
        flash(f"Import error: {e}", "danger")
        return render_template("master_data_import.html")
    finally:
        if tmp_path and _os.path.exists(tmp_path):
            _os.unlink(tmp_path)

    return render_template(
        "master_data_import.html",
        result={
            "items_created":   items_created,
            "items_skipped":   items_skipped,
            "bundles_created": bundles_created,
            "bundles_skipped": bundles_skipped,
            "errors":          errors,
        }
    )


# ── SUPPLIER ROUTES ──────────────────────────────────────────────────────────
def next_sr_id(db):
    yr = date.today().year
    last = db.execute(
        "SELECT request_id FROM supplier_requests WHERE request_id LIKE ? ORDER BY request_id DESC LIMIT 1",
        (f"SR-{yr}-%",)
    ).fetchone()
    if last:
        seq = int(last["request_id"].split("-")[-1]) + 1
    else:
        seq = 1
    return f"SR-{yr}-{seq:03d}"


@app.route("/suppliers")
@login_required
def supplier_list():
    db = get_db()
    suppliers = db.execute(
        "SELECT * FROM suppliers ORDER BY supplier_id"
    ).fetchall()
    # Pending requests visible to approvers; all requests shown to requesters
    role = session["role"]
    if role in ("admin", "ops_manager", "finance_manager"):
        pending = db.execute(
            """SELECT * FROM supplier_requests
               WHERE status IN ('Pending','Ops Approved')
               ORDER BY created_at DESC"""
        ).fetchall()
    else:
        pending = db.execute(
            """SELECT * FROM supplier_requests
               WHERE raised_by=? AND status NOT IN ('Approved','Rejected')
               ORDER BY created_at DESC""",
            (session["username"],)
        ).fetchall()
    can_request = role in ("admin", "ops_exec", "ops_planner", "finance_exec")
    can_approve = role in ("admin", "ops_manager", "finance_manager")
    return render_template("suppliers.html",
        suppliers=suppliers, pending=pending,
        can_request=can_request, can_approve=can_approve)


@app.route("/suppliers/request/new", methods=["GET", "POST"])
@login_required
@roles_required("admin", "ops_exec", "ops_planner", "finance_exec")
def supplier_request_new():
    db = get_db()
    if request.method == "POST":
        f = request.form
        request_id   = next_sr_id(db)
        request_type = f.get("request_type", "New")
        supplier_id  = f.get("supplier_id", "").strip() or None
        try:
            prop_supplier_id = f.get("prop_supplier_id","").strip().upper() or None
            # Validate proposed supplier ID uniqueness for New requests
            if request_type == "New" and prop_supplier_id:
                if db.execute("SELECT id FROM suppliers WHERE supplier_id=?", (prop_supplier_id,)).fetchone():
                    flash(f"Supplier ID '{prop_supplier_id}' already exists. Choose a different one.", "danger")
                    return redirect(url_for("supplier_list"))
                if db.execute("SELECT id FROM supplier_requests WHERE prop_supplier_id=? AND status NOT IN ('Rejected')", (prop_supplier_id,)).fetchone():
                    flash(f"Supplier ID '{prop_supplier_id}' is already used in another pending request.", "danger")
                    return redirect(url_for("supplier_list"))
            db.execute("""INSERT INTO supplier_requests
                (request_id, request_type, status, supplier_id, prop_supplier_id,
                 prop_name, prop_contact_name, prop_email, prop_phone,
                 prop_payment_terms, prop_lead_days, prop_currency,
                 prop_bank_name, prop_bank_account, prop_swift_code, prop_notes,
                 reason, raised_by, raised_date)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (request_id, request_type, "Pending", supplier_id, prop_supplier_id,
                 f.get("prop_name","").strip() or None,
                 f.get("prop_contact_name","").strip() or None,
                 f.get("prop_email","").strip() or None,
                 f.get("prop_phone","").strip() or None,
                 f.get("prop_payment_terms","").strip() or None,
                 int(f.get("prop_lead_days") or 14),
                 f.get("prop_currency","USD").strip(),
                 f.get("prop_bank_name","").strip() or None,
                 f.get("prop_bank_account","").strip() or None,
                 f.get("prop_swift_code","").strip() or None,
                 f.get("prop_notes","").strip() or None,
                 f.get("reason","").strip() or None,
                 session["username"], date.today().isoformat()))
            db.commit()
            flash(f"Supplier request {request_id} submitted for approval.", "success")
            return redirect(url_for("supplier_list"))
        except Exception as e:
            flash(f"Error submitting request: {e}", "danger")

    request_type = request.args.get("type", "New")
    supplier_id  = request.args.get("supplier_id", "")
    supplier     = None
    if supplier_id:
        supplier = db.execute(
            "SELECT * FROM suppliers WHERE supplier_id=?", (supplier_id,)
        ).fetchone()
    suppliers = db.execute(
        "SELECT supplier_id, name FROM suppliers WHERE status='Active' ORDER BY name"
    ).fetchall()
    return render_template("supplier_request.html",
        request_type=request_type, supplier=supplier,
        supplier_id=supplier_id, suppliers=suppliers, today=date.today().isoformat())


@app.route("/suppliers/request/<int:req_id>", methods=["GET", "POST"])
@login_required
def supplier_request_detail(req_id):
    db  = get_db()
    req = db.execute("SELECT * FROM supplier_requests WHERE id=?", (req_id,)).fetchone()
    if not req:
        flash("Request not found.", "danger")
        return redirect(url_for("supplier_list"))

    if request.method == "POST":
        action = request.form.get("action", "")
        role   = session["role"]

        if action == "approve" and role in ("admin", "ops_manager", "finance_manager"):
            if req["status"] not in ("Pending", "Ops Approved"):
                flash("Request is not in an approvable state.", "danger")
            else:
                _apply_supplier_request(db, req)
                db.execute("""UPDATE supplier_requests SET status='Approved',
                    fin_approved_by=?, fin_approval_date=? WHERE id=?""",
                    (session["username"], date.today().isoformat(), req_id))
                db.commit()
                flash("Request approved and applied to supplier master data.", "success")

        elif action == "reject" and role in ("admin", "ops_manager", "finance_manager"):
            reason = request.form.get("rejection_reason", "").strip()
            db.execute("""UPDATE supplier_requests SET status='Rejected',
                rejected_by=?, rejection_reason=?, rejected_date=? WHERE id=?""",
                (session["username"], reason, date.today().isoformat(), req_id))
            db.commit()
            flash("Request rejected.", "warning")

        else:
            flash("Action not permitted.", "danger")

        return redirect(url_for("supplier_request_detail", req_id=req_id))

    # Reload after potential POST
    req      = db.execute("SELECT * FROM supplier_requests WHERE id=?", (req_id,)).fetchone()
    supplier = None
    if req["supplier_id"]:
        supplier = db.execute(
            "SELECT * FROM suppliers WHERE supplier_id=?", (req["supplier_id"],)
        ).fetchone()
    role = session["role"]
    can_approve = role in ("admin", "ops_manager", "finance_manager") and req["status"] in ("Pending", "Ops Approved")
    can_reject  = role in ("admin", "ops_manager", "finance_manager") and req["status"] in ("Pending", "Ops Approved")
    # Docs for this request
    req_docs = db.execute(
        "SELECT * FROM supplier_request_docs WHERE request_id=? ORDER BY uploaded_at DESC",
        (req_id,)).fetchall()
    return render_template("supplier_request_detail.html",
        req=req, supplier=supplier,
        can_approve=can_approve,
        can_reject=can_reject,
        req_docs=req_docs)


def _apply_supplier_request(db, req):
    """Apply an approved supplier request to the suppliers master table."""
    rtype = req["request_type"]

    if rtype == "New":
        # Use proposed supplier ID if set, else auto-generate
        if req["prop_supplier_id"]:
            new_sid = req["prop_supplier_id"]
        else:
            last = db.execute(
                "SELECT supplier_id FROM suppliers ORDER BY supplier_id DESC LIMIT 1"
            ).fetchone()
            if last:
                try:
                    num = int(last["supplier_id"].replace("S","")) + 1
                except Exception:
                    num = 1
                new_sid = f"S{num:03d}"
            else:
                new_sid = "S001"
        db.execute("""INSERT INTO suppliers
            (supplier_id, name, contact_name, email, phone, payment_terms,
             lead_days, currency, bank_name, bank_account, swift_code, status, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,'Active',?)""",
            (new_sid,
             req["prop_name"], req["prop_contact_name"], req["prop_email"],
             req["prop_phone"], req["prop_payment_terms"], req["prop_lead_days"] or 14,
             req["prop_currency"] or "USD", req["prop_bank_name"],
             req["prop_bank_account"], req["prop_swift_code"], req["prop_notes"]))
        # Update the request with the newly assigned supplier_id
        db.execute("UPDATE supplier_requests SET supplier_id=? WHERE id=?",
                   (new_sid, req["id"]))

    elif rtype == "Edit":
        sid = req["supplier_id"]
        updates, params = [], []
        fields = [
            ("name", req["prop_name"]),
            ("contact_name", req["prop_contact_name"]),
            ("email", req["prop_email"]),
            ("phone", req["prop_phone"]),
            ("payment_terms", req["prop_payment_terms"]),
            ("lead_days", req["prop_lead_days"]),
            ("currency", req["prop_currency"]),
            ("bank_name", req["prop_bank_name"]),
            ("bank_account", req["prop_bank_account"]),
            ("swift_code", req["prop_swift_code"]),
            ("notes", req["prop_notes"]),
        ]
        for col, val in fields:
            if val is not None:
                updates.append(f"{col}=?")
                params.append(val)
        if updates:
            params.append(sid)
            db.execute(f"UPDATE suppliers SET {', '.join(updates)} WHERE supplier_id=?", params)

    elif rtype == "Deactivate":
        db.execute("UPDATE suppliers SET status='Inactive' WHERE supplier_id=?",
                   (req["supplier_id"],))


@app.route("/suppliers/<supplier_id>")
@login_required
def supplier_detail(supplier_id):
    db = get_db()
    supplier = db.execute("SELECT * FROM suppliers WHERE supplier_id=?", (supplier_id,)).fetchone()
    if not supplier:
        flash("Supplier not found.", "danger")
        return redirect(url_for("supplier_list"))
    linked_items = db.execute("""
        SELECT i.item_id, i.name, i.category, i.unit_of_measure, i.reorder_point, i.status
        FROM supplier_items si
        JOIN items i ON si.item_id = i.item_id
        WHERE si.supplier_id=?
        ORDER BY i.item_id
    """, (supplier_id,)).fetchall()
    linked_ids = {r["item_id"] for r in linked_items}
    all_items = db.execute(
        "SELECT item_id, name, category FROM items WHERE status='Active' ORDER BY item_id"
    ).fetchall()
    available_items = [i for i in all_items if i["item_id"] not in linked_ids]
    role = session["role"]
    can_manage_items = role in ("admin", "ops_manager", "ops_planner", "ops_exec", "finance_exec")
    can_request = role in ("admin", "ops_exec", "ops_planner", "finance_exec")
    # Recent requests for this supplier
    requests = db.execute("""
        SELECT * FROM supplier_requests WHERE supplier_id=?
        ORDER BY created_at DESC LIMIT 5
    """, (supplier_id,)).fetchall()
    return render_template("supplier_detail.html",
        supplier=supplier, linked_items=linked_items,
        available_items=available_items, requests=requests,
        can_manage_items=can_manage_items, can_request=can_request)


@app.route("/suppliers/<supplier_id>/items/add", methods=["POST"])
@login_required
@roles_required("admin", "ops_manager", "ops_planner", "ops_exec", "finance_exec")
def supplier_item_add(supplier_id):
    db = get_db()
    item_id = request.form.get("item_id", "").strip()
    if not item_id:
        flash("Please select an item.", "warning")
        return redirect(url_for("supplier_detail", supplier_id=supplier_id))
    try:
        db.execute(
            "INSERT OR IGNORE INTO supplier_items (supplier_id, item_id, created_by) VALUES (?,?,?)",
            (supplier_id, item_id, session["username"]))
        db.commit()
        item = db.execute("SELECT name FROM items WHERE item_id=?", (item_id,)).fetchone()
        flash(f"'{item['name']}' linked to this supplier.", "success")
    except Exception as e:
        flash(f"Error linking item: {e}", "danger")
    return redirect(url_for("supplier_detail", supplier_id=supplier_id))


@app.route("/suppliers/<supplier_id>/items/remove/<item_id>", methods=["POST"])
@login_required
@roles_required("admin", "ops_manager", "ops_planner", "ops_exec", "finance_exec")
def supplier_item_remove(supplier_id, item_id):
    db = get_db()
    db.execute(
        "DELETE FROM supplier_items WHERE supplier_id=? AND item_id=?",
        (supplier_id, item_id))
    db.commit()
    flash(f"Item {item_id} unlinked from this supplier.", "success")
    return redirect(url_for("supplier_detail", supplier_id=supplier_id))


@app.route("/api/supplier-items/<supplier_id>")
@login_required
def api_supplier_items(supplier_id):
    db = get_db()
    rows = db.execute("""
        SELECT i.item_id, i.name,
               COALESCE(si.unit_price, i.unit_price, 0) AS unit_price
        FROM supplier_items si
        JOIN items i ON si.item_id = i.item_id
        WHERE si.supplier_id=? AND i.status='Active'
        ORDER BY i.item_id
    """, (supplier_id,)).fetchall()
    return jsonify([{"item_id": r["item_id"], "name": r["name"],
                     "unit_price": r["unit_price"]} for r in rows])


# ── PO PDF DOWNLOAD ───────────────────────────────────────────────────────────
@app.route("/pos/<int:po_id>/download")
@login_required
def po_download(po_id):
    import traceback as _tb
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    from flask import Response as FlaskResponse

    db = get_db()

    # ── Fetch PO & guard ─────────────────────────────────────────────────────
    try:
        po = db.execute("""
            SELECT po.*, s.name supplier_name, s.contact_name,
                   s.email sup_email, s.phone sup_phone,
                   s.payment_terms, s.bank_name, s.bank_account, s.swift_code
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            WHERE po.id=?""", (po_id,)).fetchone()
    except Exception as e:
        flash(f"DB error: {e}", "danger")
        return redirect(url_for("po_list"))

    if not po:
        flash("PO not found.", "danger")
        return redirect(url_for("po_list"))

    if po["status"] not in ("Approved", "Deposit Paid",
                             "Fully Paid – Shipped", "Fully Paid – Received",
                             "Partially Received", "Closed – Partial"):
        flash("PO must be approved before downloading.", "warning")
        return redirect(url_for("po_detail", po_id=po_id))

    try:
        lines = db.execute("""
            SELECT pl.qty_ordered, pl.unit_price, pl.free_units,
                   i.item_id, i.name item_name,
                   COALESCE(i.description,'') item_desc,
                   COALESCE(i.unit_of_measure,'Units') unit_of_measure
            FROM po_lines pl
            JOIN items i ON pl.item_id = i.item_id
            WHERE pl.po_id=? ORDER BY pl.id
        """, (po_id,)).fetchall()
    except Exception as e:
        flash(f"DB error fetching lines: {e}", "danger")
        return redirect(url_for("po_detail", po_id=po_id))

    cn_row = db.execute("SELECT value FROM settings WHERE key='company_name'").fetchone()
    company_name = cn_row["value"] if cn_row else "Company"

    # ── Totals ────────────────────────────────────────────────────────────────
    subtotal    = sum((r["qty_ordered"] or 0)*(r["unit_price"] or 0) for r in lines)
    total_qty   = sum((r["qty_ordered"] or 0)+(r["free_units"] or 0) for r in lines)
    freight     = float(po["freight_costs"] or 0)
    other_c     = float(po["other_costs"]   or 0)
    total_val   = subtotal + freight + other_c
    eff_cost    = total_val / total_qty if total_qty else 0
    deposit_pct = float(po["deposit_pct"] or 0)
    deposit_amt = total_val * deposit_pct

    # ── Helper functions ──────────────────────────────────────────────────────
    def fc(v):
        try:    return "${:,.2f}".format(float(v))
        except: return "$0.00"
    def fd(v):
        if not v: return "—"
        try:    return datetime.strptime(str(v), "%Y-%m-%d").strftime("%d %b %Y")
        except: return str(v)
    def fq(v):
        try:    return "{:,.0f}".format(float(v))
        except: return "0"

    styles = getSampleStyleSheet()
    _style_counter = [0]
    def S(base, **kw):
        _style_counter[0] += 1
        s = styles[base].clone(f"{base}_{_style_counter[0]}")
        for k, v in kw.items():
            setattr(s, k, v)
        return s

    NAVY   = colors.HexColor("#1F4E79")
    BLUE   = colors.HexColor("#2E75B6")
    LTBLUE = colors.HexColor("#DEEAF1")
    LGRAY  = colors.HexColor("#F4F6F9")
    MGRAY  = colors.HexColor("#E0E0E0")
    WHITE  = colors.white

    # ── Build PDF ─────────────────────────────────────────────────────────────
    try:
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=20*mm, rightMargin=20*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
        story = []

        # Header
        story.append(Table([[
            Paragraph(f"<b>{company_name}</b>",
                      S("Normal", fontSize=18, fontName="Helvetica-Bold", textColor=NAVY)),
            Paragraph("<b>PURCHASE ORDER</b>",
                      S("Normal", fontSize=18, fontName="Helvetica-Bold",
                        textColor=BLUE, alignment=TA_RIGHT)),
        ]], colWidths=[85*mm, 85*mm]))
        story.append(Table([[
            Paragraph("Inventory Management System",
                      S("Normal", fontSize=8, textColor=colors.HexColor("#555555"))),
            Paragraph(f"<b>{po['po_number']}</b>",
                      S("Normal", fontSize=14, fontName="Helvetica-Bold",
                        textColor=NAVY, alignment=TA_RIGHT)),
        ]], colWidths=[85*mm, 85*mm]))
        story.append(HRFlowable(width="100%", thickness=2, color=NAVY, spaceAfter=6))

        # Meta grid
        b9 = S("Normal", fontSize=9, fontName="Helvetica-Bold")
        n9 = S("Normal", fontSize=9)
        meta = [
            [Paragraph("<b>PO Date</b>",       b9), Paragraph(fd(po["po_date"]),   n9),
             Paragraph("<b>Status</b>",         b9), Paragraph(po["status"],        n9)],
            [Paragraph("<b>Raised By</b>",      b9), Paragraph(po["raised_by"] or "—", n9),
             Paragraph("<b>Approved By</b>",    b9), Paragraph(po["approved_by"] or "—", n9)],
            [Paragraph("<b>Exp. Delivery</b>",  b9), Paragraph(fd(po["expected_delivery_date"]), n9),
             Paragraph("<b>Currency</b>",       b9), Paragraph(po["currency"] or "SGD", n9)],
            [Paragraph("<b>Payment Due</b>",    b9), Paragraph(fd(po["payment_due_date"]), n9),
             Paragraph("<b>Deposit</b>",        b9), Paragraph(f"{int(deposit_pct*100)}%", n9)],
        ]
        mt = Table(meta, colWidths=[30*mm, 55*mm, 30*mm, 55*mm])
        mt.setStyle(TableStyle([
            ("ROWBACKGROUNDS", (0,0),(-1,-1), [LGRAY, WHITE]),
            ("GRID",           (0,0),(-1,-1), 0.5, MGRAY),
            ("LEFTPADDING",    (0,0),(-1,-1), 6),
            ("RIGHTPADDING",   (0,0),(-1,-1), 6),
            ("TOPPADDING",     (0,0),(-1,-1), 4),
            ("BOTTOMPADDING",  (0,0),(-1,-1), 4),
        ]))
        story.append(mt)
        story.append(Spacer(1, 8))

        # Supplier / Deliver-to boxes
        def side_box(rows, hdr_col):
            t = Table([[Paragraph(r, S("Normal", fontSize=8,
                        **({} if i else {"fontName":"Helvetica-Bold","textColor":WHITE})))]
                       for i, r in enumerate(rows)],
                      colWidths=[80*mm])
            t.setStyle(TableStyle([
                ("BACKGROUND",    (0,0),(0,0), hdr_col),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE, LGRAY]),
                ("BOX",           (0,0),(-1,-1), 0.5, MGRAY),
                ("LEFTPADDING",   (0,0),(-1,-1), 6),
                ("TOPPADDING",    (0,0),(-1,-1), 3),
                ("BOTTOMPADDING", (0,0),(-1,-1), 3),
            ]))
            return t

        sup_rows = [
            "SUPPLIER",
            f"<b>{po['supplier_name']}</b>",
            f"Contact: {po['contact_name'] or '—'}",
            f"Email:   {po['sup_email']    or '—'}",
            f"Phone:   {po['sup_phone']    or '—'}",
            f"Terms:   {po['payment_terms'] or '—'}",
        ]
        if po["bank_name"]:
            sup_rows.append(f"Bank: {po['bank_name']}  |  Acct: {po['bank_account'] or '—'}")

        del_rows = [
            "DELIVER TO",
            f"<b>{company_name}</b>",
            "Inventory Management System",
            f"Expected: {fd(po['expected_delivery_date'])}",
            f"Supplier ID: {po['supplier_id']}",
        ]

        at = Table([[side_box(sup_rows, NAVY), side_box(del_rows, BLUE)]],
                   colWidths=[85*mm, 85*mm])
        at.setStyle(TableStyle([
            ("VALIGN",       (0,0),(-1,-1),"TOP"),
            ("LEFTPADDING",  (0,0),(-1,-1),0),
            ("RIGHTPADDING", (0,0),(-1,-1),0),
        ]))
        story.append(at)
        story.append(Spacer(1, 10))

        # Line items table
        story.append(Paragraph("<b>ORDER DETAILS</b>",
                     S("Normal", fontSize=9, fontName="Helvetica-Bold", textColor=NAVY)))
        story.append(Spacer(1, 3))

        hdr_s = lambda t: Paragraph(f"<b>{t}</b>",
                    S("Normal", fontSize=8, fontName="Helvetica-Bold",
                      textColor=WHITE, alignment=TA_RIGHT))
        hdr_l = lambda t: Paragraph(f"<b>{t}</b>",
                    S("Normal", fontSize=8, fontName="Helvetica-Bold", textColor=WHITE))

        line_data = [[
            hdr_l("Item ID"), hdr_l("Description"), hdr_l("UOM"),
            hdr_s("Qty"), hdr_s("Free"), hdr_s("Total Qty"),
            hdr_s("Unit Price"), hdr_s("Subtotal"),
        ]]
        for i, r in enumerate(lines):
            qty   = float(r["qty_ordered"] or 0)
            free  = float(r["free_units"]  or 0)
            price = float(r["unit_price"]  or 0)
            sub   = qty * price
            desc  = r["item_name"] or ""
            if r["item_desc"]:
                desc += " — " + r["item_desc"]
            bg = WHITE if i % 2 == 0 else LGRAY
            c  = S("Normal", fontSize=8, leading=11)
            cr = S("Normal", fontSize=8, leading=11, alignment=TA_RIGHT)
            line_data.append([
                Paragraph(r["item_id"],       c),
                Paragraph(desc,               c),
                Paragraph(r["unit_of_measure"], S("Normal", fontSize=8, alignment=TA_CENTER)),
                Paragraph(fq(qty),            cr),
                Paragraph(fq(free),           cr),
                Paragraph(fq(qty+free),       cr),
                Paragraph(fc(price),          cr),
                Paragraph(fc(sub),            cr),
            ])

        lt = Table(line_data, colWidths=[18*mm, 43*mm, 12*mm, 16*mm, 14*mm, 16*mm, 18*mm, 22*mm])
        lt_style = [
            ("BACKGROUND",    (0,0),(-1,0),  NAVY),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("GRID",          (0,0),(-1,-1), 0.5, MGRAY),
            ("LEFTPADDING",   (0,0),(-1,-1), 5),
            ("RIGHTPADDING",  (0,0),(-1,-1), 5),
            ("TOPPADDING",    (0,0),(-1,-1), 4),
            ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ]
        for i in range(1, len(line_data)):
            lt_style.append(("BACKGROUND",(0,i),(-1,i), WHITE if i%2==1 else LGRAY))
        lt.setStyle(TableStyle(lt_style))
        story.append(lt)
        story.append(Spacer(1, 6))

        # Cost summary
        fw = lambda t, bold=False: Paragraph(t, S("Normal", fontSize=9,
                    fontName="Helvetica-Bold" if bold else "Helvetica"))
        fwr = lambda t, bold=False: Paragraph(t, S("Normal", fontSize=9,
                    fontName="Helvetica-Bold" if bold else "Helvetica",
                    alignment=TA_RIGHT))
        sum_rows2 = [
            (fw("Lines Subtotal"),                               fwr(fc(subtotal)),        LGRAY),
            (fw("Freight / Shipping"),                           fwr(fc(freight)),          WHITE),
            (fw("Other Costs"),                                  fwr(fc(other_c)),          LGRAY),
            (fw(f"Total PO Value ({po['currency'] or 'SGD'})", True), fwr(fc(total_val), True), LTBLUE),
            (fw(f"Deposit ({int(deposit_pct*100)}%)"),           fwr(fc(deposit_amt)),      WHITE),
            (fw("Balance Due"),                                  fwr(fc(total_val-deposit_amt)), LGRAY),
            (fw("Effective Unit Cost (avg)"),                    fwr(f"${eff_cost:.4f}"),   WHITE),
        ]
        st = Table([[a, b] for a, b, _ in sum_rows2], colWidths=[120*mm, 50*mm], hAlign="RIGHT")
        st_style = [("GRID",(0,0),(-1,-1),0.5,MGRAY),
                    ("LEFTPADDING",(0,0),(-1,-1),6),
                    ("RIGHTPADDING",(0,0),(-1,-1),6),
                    ("TOPPADDING",(0,0),(-1,-1),3),
                    ("BOTTOMPADDING",(0,0),(-1,-1),3)]
        for i, (_, _, bg) in enumerate(sum_rows2):
            st_style.append(("BACKGROUND",(0,i),(-1,i),bg))
        st.setStyle(TableStyle(st_style))
        story.append(st)
        story.append(Spacer(1, 10))

        # Notes
        if po["notes"]:
            story.append(HRFlowable(width="100%", thickness=0.5, color=MGRAY, spaceAfter=4))
            story.append(Paragraph("<b>Notes</b>",
                         S("Normal", fontSize=9, fontName="Helvetica-Bold")))
            story.append(Paragraph(str(po["notes"]),
                         S("Normal", fontSize=9, leading=13)))
            story.append(Spacer(1, 6))

        # Footer
        story.append(HRFlowable(width="100%", thickness=1, color=NAVY,
                                spaceBefore=4, spaceAfter=4))
        story.append(Table([[
            Paragraph(f"Computer-generated by {company_name} IMS.",
                      S("Normal", fontSize=8, textColor=colors.HexColor("#555555"))),
            Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}",
                      S("Normal", fontSize=8, textColor=colors.HexColor("#555555"),
                        alignment=TA_RIGHT)),
        ]], colWidths=[120*mm, 50*mm]))

        doc.build(story)
        buf.seek(0)
        fname = f"{po['po_number']}.pdf"
        return FlaskResponse(buf.read(), mimetype="application/pdf",
                             headers={"Content-Disposition":
                                      f'attachment; filename="{fname}"'})

    except Exception as _e:
        flash(f"PDF error ({type(_e).__name__}): {_e}", "danger")
        app.logger.error("po_download error:\n" + _tb.format_exc())
        return redirect(url_for("po_detail", po_id=po_id))


# ── SUPPLIER PO PDF ───────────────────────────────────────────────────────────
@app.route("/pos/<int:po_id>/download/supplier")
@login_required
def po_download_supplier(po_id):
    import traceback as _tb
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
    from flask import Response as FlaskResponse

    db = get_db()

    # All roles except admin may only download the supplier PDF once the PO
    # has cleared both approvals (status = Approved or any later stage).
    role = session.get("role","")
    APPROVED_STATUSES = (
        "Approved", "Deposit Paid", "Fully Paid – Shipped",
        "Fully Paid – Received", "Partially Received"
    )
    if role != "admin":
        po_check = db.execute("SELECT status FROM purchase_orders WHERE id=?", (po_id,)).fetchone()
        if not po_check or po_check["status"] not in APPROVED_STATUSES:
            flash("The supplier PDF is only available once the PO has been fully approved by both Ops and Finance.", "warning")
            return redirect(url_for("po_detail", po_id=po_id))

    try:
        po = db.execute("""
            SELECT po.*, s.name supplier_name, s.contact_name,
                   s.email sup_email, s.phone sup_phone
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.supplier_id
            WHERE po.id=?""", (po_id,)).fetchone()
    except Exception as e:
        flash(f"DB error: {e}", "danger"); return redirect(url_for("po_list"))

    if not po:
        flash("PO not found.", "danger"); return redirect(url_for("po_list"))
    if po["status"] not in ("Approved","Deposit Paid","Fully Paid – Shipped","Fully Paid – Received"):
        flash("PO must be approved before downloading.", "warning")
        return redirect(url_for("po_detail", po_id=po_id))

    try:
        lines = db.execute("""
            SELECT pl.qty_ordered, pl.unit_price, pl.free_units,
                   i.item_id, i.name item_name,
                   COALESCE(i.unit_of_measure,'Units') unit_of_measure
            FROM po_lines pl JOIN items i ON pl.item_id=i.item_id
            WHERE pl.po_id=? ORDER BY pl.id""", (po_id,)).fetchall()
    except Exception as e:
        flash(f"DB error: {e}", "danger"); return redirect(url_for("po_detail", po_id=po_id))

    s_row = {r["key"]: r["value"] for r in db.execute("SELECT key,value FROM settings").fetchall()}
    company_name      = s_row.get("company_name","Company")
    company_phone     = s_row.get("company_phone","")
    company_email     = s_row.get("company_email","")
    warehouse_address = s_row.get("company_warehouse_address","")

    subtotal  = sum((r["qty_ordered"] or 0)*(r["unit_price"] or 0) for r in lines)
    freight   = float(po["freight_costs"] or 0)
    other_c   = float(po["other_costs"]   or 0)
    total_val = subtotal + freight + other_c
    ccy       = po["currency"] or "SGD"

    def fc(v):
        try:    return "${:,.2f}".format(float(v))
        except: return "$0.00"
    def fd(v):
        if not v: return "—"
        try:    return datetime.strptime(str(v),"%Y-%m-%d").strftime("%d %b %Y")
        except: return str(v)
    def fq(v):
        try:    return "{:,.0f}".format(float(v))
        except: return "0"

    styles = getSampleStyleSheet()
    _c = [0]
    def S(base, **kw):
        _c[0] += 1
        s = styles[base].clone(f"s{_c[0]}")
        for k,v in kw.items(): setattr(s,k,v)
        return s

    NAVY  = colors.HexColor("#1F4E79")
    BLUE  = colors.HexColor("#2E75B6")
    LGRAY = colors.HexColor("#F4F6F9")
    MGRAY = colors.HexColor("#E0E0E0")
    LTBLUE= colors.HexColor("#DEEAF1")
    WHITE = colors.white

    try:
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=20*mm, rightMargin=20*mm,
                                topMargin=15*mm, bottomMargin=15*mm)
        story = []

        # Header
        story.append(Table([[
            Paragraph(f"<b>{company_name}</b>",
                      S("Normal",fontSize=20,fontName="Helvetica-Bold",textColor=NAVY)),
            Paragraph("<b>PURCHASE ORDER</b>",
                      S("Normal",fontSize=16,fontName="Helvetica-Bold",textColor=BLUE,alignment=TA_RIGHT)),
        ]], colWidths=[95*mm, 75*mm]))
        contact_parts = []
        if company_phone: contact_parts.append(f"Tel: {company_phone}")
        if company_email: contact_parts.append(company_email)
        if contact_parts:
            story.append(Paragraph(" &nbsp;·&nbsp; ".join(contact_parts),
                         S("Normal",fontSize=8,textColor=colors.HexColor("#666"))))
        story.append(Spacer(1,4))
        story.append(HRFlowable(width="100%",thickness=2,color=NAVY,spaceAfter=6))

        # PO ref (right-aligned)
        ref_rows = [
            [Paragraph("<b>PO Number</b>",S("Normal",fontSize=9,fontName="Helvetica-Bold")),
             Paragraph(f"<b>{po['po_number']}</b>",S("Normal",fontSize=11,fontName="Helvetica-Bold",textColor=NAVY,alignment=TA_RIGHT))],
            [Paragraph("<b>Date</b>",S("Normal",fontSize=9,fontName="Helvetica-Bold")),
             Paragraph(fd(po["po_date"]),S("Normal",fontSize=9,alignment=TA_RIGHT))],
            [Paragraph("<b>Currency</b>",S("Normal",fontSize=9,fontName="Helvetica-Bold")),
             Paragraph(ccy,S("Normal",fontSize=9,alignment=TA_RIGHT))],
        ]
        rt = Table(ref_rows, colWidths=[28*mm,52*mm], hAlign="RIGHT")
        rt.setStyle(TableStyle([
            ("ROWBACKGROUNDS",(0,0),(-1,-1),[LGRAY,WHITE]),
            ("GRID",(0,0),(-1,-1),0.5,MGRAY),
            ("LEFTPADDING",(0,0),(-1,-1),8),
            ("RIGHTPADDING",(0,0),(-1,-1),8),
            ("TOPPADDING",(0,0),(-1,-1),4),
            ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ]))
        story.append(rt)
        story.append(Spacer(1,10))

        # To / Deliver To boxes
        def side_box(title, lines_list, hdr_col):
            rows = [[Paragraph(f"<b>{title}</b>",
                     S("Normal",fontSize=8,fontName="Helvetica-Bold",textColor=WHITE))]]
            for ln in lines_list:
                rows.append([Paragraph(ln,S("Normal",fontSize=9,leading=13))])
            t = Table(rows,colWidths=[80*mm])
            t.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(0,0),hdr_col),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[WHITE,LGRAY]),
                ("BOX",(0,0),(-1,-1),0.5,MGRAY),
                ("LEFTPADDING",(0,0),(-1,-1),8),
                ("TOPPADDING",(0,0),(-1,-1),4),
                ("BOTTOMPADDING",(0,0),(-1,-1),4),
            ]))
            return t

        sup_lines = [f"<b>{po['supplier_name']}</b>"]
        if po["contact_name"]: sup_lines.append(po["contact_name"])
        if po["sup_email"]:    sup_lines.append(po["sup_email"])
        if po["sup_phone"]:    sup_lines.append(po["sup_phone"])

        del_lines = [f"<b>{company_name}</b>"]
        if warehouse_address:
            for ln in warehouse_address.strip().splitlines():
                if ln.strip(): del_lines.append(ln.strip())
        else:
            del_lines.append("(Configure warehouse address in Company Settings)")

        story.append(Table([[side_box("TO",sup_lines,NAVY),
                             side_box("DELIVER TO",del_lines,BLUE)]],
                           colWidths=[85*mm,85*mm]))
        story.append(Spacer(1,12))

        # Line items
        story.append(Paragraph("<b>ORDER DETAILS</b>",
                     S("Normal",fontSize=10,fontName="Helvetica-Bold",textColor=NAVY)))
        story.append(Spacer(1,4))

        def ph(t,align=TA_LEFT):
            return Paragraph(f"<b>{t}</b>",
                   S("Normal",fontSize=9,fontName="Helvetica-Bold",textColor=WHITE,alignment=align))

        line_data = [[ph("Item ID"),ph("Description"),ph("UOM"),
                      ph("Qty Ordered",TA_RIGHT),ph("Free Units",TA_RIGHT),
                      ph("Total Qty",TA_RIGHT),ph("Unit Price",TA_RIGHT),ph("Amount",TA_RIGHT)]]
        for i,r in enumerate(lines):
            qty   = float(r["qty_ordered"] or 0)
            free  = float(r["free_units"]  or 0)
            price = float(r["unit_price"]  or 0)
            c  = S("Normal",fontSize=9,leading=12)
            cr = S("Normal",fontSize=9,leading=12,alignment=TA_RIGHT)
            line_data.append([
                Paragraph(r["item_id"],c),
                Paragraph(r["item_name"] or "—",c),
                Paragraph(r["unit_of_measure"],S("Normal",fontSize=9,alignment=TA_CENTER)),
                Paragraph(fq(qty),cr),Paragraph(fq(free),cr),
                Paragraph(fq(qty+free),cr),Paragraph(fc(price),cr),
                Paragraph(fc(qty*price),cr),
            ])

        lt = Table(line_data,colWidths=[20*mm,47*mm,13*mm,18*mm,16*mm,16*mm,18*mm,21*mm])
        lt_style = [("BACKGROUND",(0,0),(-1,0),NAVY),("FONTSIZE",(0,0),(-1,-1),9),
                    ("GRID",(0,0),(-1,-1),0.5,MGRAY),("LEFTPADDING",(0,0),(-1,-1),5),
                    ("RIGHTPADDING",(0,0),(-1,-1),5),("TOPPADDING",(0,0),(-1,-1),5),
                    ("BOTTOMPADDING",(0,0),(-1,-1),5)]
        for i in range(1,len(line_data)):
            lt_style.append(("BACKGROUND",(0,i),(-1,i),WHITE if i%2==1 else LGRAY))
        lt.setStyle(TableStyle(lt_style))
        story.append(lt)
        story.append(Spacer(1,6))

        # Totals
        sum_rows = [("Subtotal",fc(subtotal),LGRAY)]
        if freight: sum_rows.append(("Freight / Shipping",fc(freight),WHITE))
        if other_c: sum_rows.append(("Other Charges",fc(other_c),LGRAY if not freight else WHITE))
        sum_rows.append((f"<b>TOTAL ({ccy})</b>",f"<b>{fc(total_val)}</b>",LTBLUE))

        st = Table([[Paragraph(lbl,S("Normal",fontSize=9)),
                     Paragraph(val,S("Normal",fontSize=9,alignment=TA_RIGHT))]
                    for lbl,val,_ in sum_rows],
                   colWidths=[120*mm,50*mm],hAlign="RIGHT")
        st_style = [("GRID",(0,0),(-1,-1),0.5,MGRAY),
                    ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
                    ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]
        for i,(_,_,bg) in enumerate(sum_rows):
            st_style.append(("BACKGROUND",(0,i),(-1,i),bg))
        st.setStyle(TableStyle(st_style))
        story.append(st)
        story.append(Spacer(1,12))

        # Notes (strip internal return stamps)
        raw = po["notes"] or ""
        clean = "\n".join(l for l in raw.splitlines()
                          if not l.strip().startswith("[Returned by")).strip()
        if clean:
            story.append(HRFlowable(width="100%",thickness=0.5,color=MGRAY,spaceAfter=4))
            story.append(Paragraph("<b>Notes</b>",
                         S("Normal",fontSize=10,fontName="Helvetica-Bold",textColor=NAVY)))
            story.append(Spacer(1,4))
            story.append(Paragraph(clean,S("Normal",fontSize=9,leading=13)))
            story.append(Spacer(1,8))

        # Footer
        story.append(HRFlowable(width="100%",thickness=1,color=NAVY,spaceBefore=6,spaceAfter=4))
        story.append(Table([[
            Paragraph(f"Issued by {company_name}. Please quote <b>{po['po_number']}</b> on all invoices.",
                      S("Normal",fontSize=7,textColor=colors.HexColor("#555"))),
            Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y')}",
                      S("Normal",fontSize=7,textColor=colors.HexColor("#555"),alignment=TA_RIGHT)),
        ]],colWidths=[130*mm,40*mm]))

        doc.build(story)
        buf.seek(0)
        fname = f"{po['po_number']}_Supplier.pdf"
        return FlaskResponse(buf.read(),mimetype="application/pdf",
                             headers={"Content-Disposition":f'attachment; filename="{fname}"'})
    except Exception as _e:
        flash(f"Supplier PDF error ({type(_e).__name__}): {_e}","danger")
        app.logger.error("po_download_supplier:\n"+_tb.format_exc())
        return redirect(url_for("po_detail",po_id=po_id))

# ─── API ENDPOINTS (JSON) ────────────────────────────────────────────────────
@app.route("/api/fifo-cost/<item_id>")
@login_required
def api_fifo_cost(item_id):
    db   = get_db()
    cost = next_fifo_cost(db, item_id)
    return jsonify({"item_id": item_id, "fifo_cost": cost})

@app.route("/api/wacc")
@login_required
def api_wacc():
    db = get_db()
    return jsonify(get_wacc_summary(db))

# ─── CONTEXT PROCESSORS ──────────────────────────────────────────────────────
@app.context_processor
def inject_globals():
    pending_count = 0
    if session.get("role") in ("admin", "ops_manager", "finance_manager"):
        try:
            db = get_db()
            role = session["role"]
            if role in ("admin", "ops_manager"):
                pending_count += db.execute(
                    "SELECT COUNT(*) FROM purchase_orders WHERE status='Pending Approval'"
                ).fetchone()[0]
                pending_count += db.execute(
                    "SELECT COUNT(*) FROM supplier_requests WHERE status='Pending'"
                ).fetchone()[0]
            if role in ("admin", "finance_manager"):
                pending_count += db.execute(
                    "SELECT COUNT(*) FROM purchase_orders WHERE status='Ops Approved'"
                ).fetchone()[0]
                pending_count += db.execute(
                    "SELECT COUNT(*) FROM supplier_requests WHERE status='Ops Approved'"
                ).fetchone()[0]
        except Exception:
            pass
    # Inject company branding
    company_name = "IMS"
    company_logo = None
    try:
        db2 = get_db()
        r = db2.execute("SELECT value FROM settings WHERE key='company_name'").fetchone()
        if r: company_name = r["value"]
        r = db2.execute("SELECT value FROM settings WHERE key='company_logo'").fetchone()
        if r: company_logo = r["value"]
    except Exception:
        pass
    return {
        "now": datetime.now(),
        "pending_count": pending_count,
        "fmt_currency": lambda v: f"${v:,.2f}" if v is not None else "—",
        "fmt_qty":      lambda v: f"{v:,.0f}" if v is not None else "—",
        "company_name": company_name,
        "company_logo": company_logo,
    }

# ─── TEMPLATE FILTERS ────────────────────────────────────────────────────────
@app.template_filter("currency")
def currency_filter(v):
    try:    return f"${float(v):,.2f}"
    except: return "—"

@app.template_filter("qty")
def qty_filter(v):
    try:    return f"{float(v):,.0f}"
    except: return "—"

@app.template_filter("dateformat")
def date_filter(v):
    if not v: return "—"
    try:
        return datetime.strptime(v, "%Y-%m-%d").strftime("%d %b %Y")
    except:
        return v


# ── BUNDLES ───────────────────────────────────────────────────────────────────
BUNDLE_ROLES = ("admin","ops_manager","ops_planner","ops_exec","finance_manager","finance_exec")

def next_bundle_id(db):
    last = db.execute("SELECT bundle_id FROM bundles ORDER BY id DESC LIMIT 1").fetchone()
    if last:
        try: num = int(last["bundle_id"].replace("B","")) + 1
        except: num = 1
    else:
        num = 1
    return f"B{num:03d}"

@app.route("/bundles")
@login_required
@roles_required(*BUNDLE_ROLES)
def bundle_list():
    db      = get_db()
    bundles = db.execute("""
        SELECT b.*,
            (SELECT COUNT(*) FROM bundle_components WHERE bundle_id=b.bundle_id) AS component_count
        FROM bundles b ORDER BY b.bundle_id
    """).fetchall()
    items_rows = db.execute("SELECT item_id, name, unit_of_measure FROM items WHERE status='Active' ORDER BY item_id").fetchall()
    items      = [dict(i) for i in items_rows]   # plain dicts so tojson works in template
    next_id = next_bundle_id(db)
    # compute COGS for each bundle
    bundle_costs = {}
    for b in bundles:
        comps = db.execute("""
            SELECT bc.item_id, bc.qty,
                   COALESCE(i.name, '(SKU not in Items)') item_name,
                   COALESCE(i.unit_of_measure, '-') unit_of_measure
            FROM bundle_components bc
            LEFT JOIN items i ON bc.item_id=i.item_id
            WHERE bc.bundle_id=?
        """, (b["bundle_id"],)).fetchall()
        cost = sum(c["qty"] * next_fifo_cost(db, c["item_id"]) for c in comps)
        bundle_costs[b["bundle_id"]] = {"cost": cost, "components": [dict(c) for c in comps]}
    return render_template("bundles.html", bundles=bundles, items=items,
                           next_id=next_id, bundle_costs=bundle_costs,
                           role=session["role"])

@app.route("/bundles/add", methods=["POST"])
@login_required
@roles_required("admin","ops_manager","ops_planner")
def bundle_add():
    f  = request.form
    db = get_db()
    bid = f.get("bundle_id","").strip().upper() or next_bundle_id(db)
    if db.execute("SELECT id FROM bundles WHERE bundle_id=?", (bid,)).fetchone():
        flash(f"Bundle ID {bid} already exists.", "danger")
        return redirect(url_for("bundle_list"))
    db.execute("INSERT INTO bundles (bundle_id,name,description,created_by) VALUES (?,?,?,?)",
               (bid, f["name"].strip(), f.get("description","").strip(), session["username"]))
    item_ids = f.getlist("comp_item_id[]")
    qtys     = f.getlist("comp_qty[]")
    for iid, qty in zip(item_ids, qtys):
        if iid.strip():
            db.execute("INSERT OR IGNORE INTO bundle_components (bundle_id,item_id,qty) VALUES (?,?,?)",
                       (bid, iid.strip(), float(qty or 1)))
    db.commit()
    flash(f"Bundle {bid} created.", "success")
    return redirect(url_for("bundle_list"))

@app.route("/bundles/<bundle_id>/edit", methods=["POST"])
@login_required
@roles_required("admin","ops_manager","ops_planner")
def bundle_edit(bundle_id):
    f  = request.form
    db = get_db()
    db.execute("UPDATE bundles SET name=?, description=?, status=? WHERE bundle_id=?",
               (f["name"].strip(), f.get("description","").strip(), f.get("status","Active"), bundle_id))
    db.execute("DELETE FROM bundle_components WHERE bundle_id=?", (bundle_id,))
    item_ids = f.getlist("comp_item_id[]")
    qtys     = f.getlist("comp_qty[]")
    for iid, qty in zip(item_ids, qtys):
        if iid.strip():
            db.execute("INSERT OR IGNORE INTO bundle_components (bundle_id,item_id,qty) VALUES (?,?,?)",
                       (bundle_id, iid.strip(), float(qty or 1)))
    db.commit()
    flash(f"Bundle {bundle_id} updated.", "success")
    return redirect(url_for("bundle_list"))

@app.route("/bundles/<bundle_id>/delete", methods=["POST"])
@login_required
@roles_required("admin","ops_manager")
def bundle_delete(bundle_id):
    db = get_db()
    db.execute("DELETE FROM bundle_components WHERE bundle_id=?", (bundle_id,))
    db.execute("DELETE FROM bundles WHERE bundle_id=?", (bundle_id,))
    db.commit()
    flash(f"Bundle {bundle_id} deleted.", "success")
    return redirect(url_for("bundle_list"))

# ── SALES UPLOADS ─────────────────────────────────────────────────────────────
import csv as _csv, io, json
try:
    import openpyxl as _xl
    _HAS_XL = True
except ImportError:
    _HAS_XL = False

SALES_ROLES = ("admin","finance_manager","finance_exec")

def next_upload_ref(db):
    yr   = date.today().year
    last = db.execute(
        "SELECT upload_ref FROM sales_uploads WHERE upload_ref LIKE ? ORDER BY id DESC LIMIT 1",
        (f"SALE-{yr}-%",)).fetchone()
    seq = int(last["upload_ref"].split("-")[-1]) + 1 if last else 1
    return f"SALE-{yr}-{seq:03d}"

def _parse_upload(file_bytes, filename):
    """Return (headers, rows-as-dicts).  rows capped at 5000."""
    rows, headers = [], []
    if filename.lower().endswith(".csv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        reader = _csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []
        for i, row in enumerate(reader):
            if i >= 5000: break
            rows.append(dict(row))
    elif _HAS_XL:
        wb = _xl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if all_rows:
            headers = [str(c) if c is not None else "" for c in all_rows[0]]
            for i, r in enumerate(all_rows[1:5001]):
                rows.append({headers[j]: (str(v) if v is not None else "") for j, v in enumerate(r)})
        wb.close()
    return headers, rows

def _guess_col(headers, candidates):
    for h in headers:
        if h and any(c.lower() in h.lower() for c in candidates):
            return h
    return None

# ── Platform categories ───────────────────────────────────────────────────────
GATEWAY_PLATFORMS = {"Airwallex", "Atome", "Stripe (AndSons)", "Stripe (GetOva)"}
ORDER_PLATFORMS   = {"Lazada (AndSons)", "Lazada (MM)", "Shopee (AndSons)", "Shopee (MM)",
                     "Tiktok (AndSons)", "Tiktok (MM)", "Shopify (MM)", "Zalora"}
ALL_PLATFORMS     = sorted(GATEWAY_PLATFORMS | ORDER_PLATFORMS)

GST_RATE = 9.0 / 109.0   # GST-inclusive: 9% of 109 = 8.257%
LARGE_TXN_THRESHOLD = 2000.0  # flag transactions above this amount

def _parse_atome(file_bytes):
    """Parse an Atome transaction Excel export.

    Returns a list of dicts, one per transaction row, with keys:
      tx_id, brand, tx_type, tx_status, tx_amount, gst_amount,
      revenue_ex_gst, amt_receivable, expenses_total,
      platform_order_id, payment_plan, is_flagged, flag_reason
    """
    if not _HAS_XL:
        raise ValueError("openpyxl is required to parse Atome files. Install it and restart.")

    wb = _xl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return []

    # Validate expected headers are present
    headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    expected = ["Outlet Name", "Transaction ID", "Transaction Type",
                "Transaction Amount", "Amount Receivable", "Transaction Status"]
    missing = [e for e in expected if e not in headers]
    if missing:
        raise ValueError(f"Atome file missing expected columns: {missing}")

    col = {h: i for i, h in enumerate(headers)}
    expense_col_names = ["MDR Fee", "Flat Fee", "MDR Rebate", "Flat Fee Rebate", "Refund Fee"]
    expense_cols = [col[c] for c in expense_col_names if c in col]

    seen_tx_ids = {}   # tx_id -> row_num, for duplicate detection
    parsed = []

    for row_num, row in enumerate(all_rows[1:], start=2):
        outlet_col = col.get("Outlet ID", 0)
        if not row or len(row) <= outlet_col or row[outlet_col] is None:
            continue  # skip blank / short rows

        def gcell(name):
            idx = col.get(name)
            return row[idx] if idx is not None else None

        def gfloat(name):
            v = gcell(name)
            if v is None: return 0.0
            try: return float(str(v).replace(",","").strip())
            except: return 0.0

        tx_id          = str(gcell("Transaction ID") or "").strip()
        brand          = str(gcell("Outlet Name") or "").strip()
        tx_type        = str(gcell("Transaction Type") or "").strip().upper()
        tx_status      = str(gcell("Transaction Status") or "").strip().upper()
        tx_amount      = gfloat("Transaction Amount")
        amt_receivable = gfloat("Amount Receivable")
        plat_order_id  = str(gcell("E-commerce Platform Order ID") or "").strip()
        payment_plan   = str(gcell("Customer's Payment Plan") or "").strip()

        gst_amount    = round(tx_amount * GST_RATE, 4)
        revenue_ex    = round(tx_amount - gst_amount, 4)
        exp_total     = sum(float(row[i] or 0) for i in expense_cols)
        exp_total     = round(exp_total, 4)

        # ── Anomaly detection ────────────────────────────────────────────────
        flags = []

        # 1. Duplicate Transaction ID
        if tx_id in seen_tx_ids:
            flags.append(f"Duplicate TX ID (also row {seen_tx_ids[tx_id]})")
        else:
            seen_tx_ids[tx_id] = row_num

        # 2. MERCHANDISE with negative Amount Receivable
        if tx_type == "MERCHANDISE" and amt_receivable < 0:
            flags.append("MERCHANDISE has negative Amount Receivable")

        # 3. REFUND with positive Amount Receivable (unusual — refunds are outflows)
        if tx_type == "REFUND" and amt_receivable > 0:
            flags.append("REFUND has positive Amount Receivable")

        # 4. Amount Receivable significantly exceeds Transaction Amount
        if tx_amount != 0 and amt_receivable > abs(tx_amount) * 1.05:
            flags.append(f"Amount Receivable ({amt_receivable}) > Transaction Amount ({tx_amount})")

        # 5. Zero-value MERCHANDISE transaction
        if tx_type == "MERCHANDISE" and tx_amount == 0:
            flags.append("Zero Transaction Amount on MERCHANDISE")

        # 6. FULLY_SETTLED with zero Amount Receivable on MERCHANDISE
        if tx_type == "MERCHANDISE" and tx_status == "FULLY_SETTLED" and amt_receivable == 0:
            flags.append("FULLY_SETTLED MERCHANDISE with zero Amount Receivable")

        # 7. Large transaction
        if abs(tx_amount) > LARGE_TXN_THRESHOLD:
            flags.append(f"Large transaction: ${tx_amount:,.2f}")

        is_flagged   = 1 if flags else 0
        flag_reason  = "; ".join(flags) if flags else None

        parsed.append({
            "row_num":        row_num - 1,   # 1-based data row number
            "tx_id":          tx_id,
            "brand":          brand,
            "tx_type":        tx_type,
            "tx_status":      tx_status,
            "tx_amount":      tx_amount,
            "gst_amount":     gst_amount,
            "revenue_ex_gst": revenue_ex,
            "amt_receivable": amt_receivable,
            "expenses_total": exp_total,
            "plat_order_id":  plat_order_id,
            "payment_plan":   payment_plan,
            "is_flagged":     is_flagged,
            "flag_reason":    flag_reason,
        })

    return parsed


def _parse_atome_payment(file_bytes):
    """Parse an Atome payout/payment Excel export.

    Returns list of dicts, one per payout row:
      payout_id, brand, payout_date, tx_status,
      total_sales, sponsored_voucher, fees_incl_gst, fees_ex_gst,
      gst_on_fees, rebates, payout_amount,
      is_flagged, flag_reason
    """
    if not _HAS_XL:
        raise ValueError("openpyxl is required to parse Atome files.")

    wb = _xl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return []

    headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    expected = ["Payout ID", "Payout Amount", "Total Sales", "All Atome Fees"]
    missing  = [e for e in expected if e not in headers]
    if missing:
        raise ValueError(f"Atome Payment file missing expected columns: {missing}. "
                         f"Make sure you are uploading the Payout List file, not the Transaction file.")

    col = {h: i for i, h in enumerate(headers)}
    parsed = []

    for row_num, row in enumerate(all_rows[1:], start=2):
        if not row or len(row) <= col.get("Payout ID", 0):
            continue
        payout_id_val = row[col.get("Payout ID", 2)]
        if payout_id_val is None:
            continue

        def gcell(name):
            idx = col.get(name)
            return row[idx] if idx is not None else None

        def gfloat(name):
            v = gcell(name)
            if v is None: return 0.0
            try: return float(str(v).replace(",", "").strip())
            except: return 0.0

        payout_id   = str(gcell("Payout ID")    or "").strip()
        brand       = str(gcell("Outlet Name")   or "").strip()
        payout_date = str(gcell("Payout Date")   or "").strip()
        tx_status   = str(gcell("Status")        or "").strip()

        total_sales       = gfloat("Total Sales")
        sponsored_voucher = gfloat("Total Sponsored Voucher Amount")  # typically negative
        fees_incl_gst     = gfloat("All Atome Fees")                  # typically negative
        rebates           = gfloat("All Rebates")
        payout_amount     = gfloat("Payout Amount")

        # Fees are stored as negative in the file — work with absolute values for accounting
        fees_abs      = abs(fees_incl_gst)
        fees_ex_gst   = round(fees_abs / 1.09, 4)
        gst_on_fees   = round(fees_abs - fees_ex_gst, 4)

        # Anomaly detection
        flags = []
        # Verify payout arithmetic: total_sales + voucher + fees + rebates ≈ payout
        expected_payout = round(total_sales + sponsored_voucher + fees_incl_gst + rebates, 2)
        if abs(expected_payout - round(payout_amount, 2)) > 0.02:
            flags.append(f"Payout arithmetic mismatch: expected {expected_payout:.2f}, got {payout_amount:.2f}")
        if payout_amount < 0:
            flags.append(f"Negative payout amount: {payout_amount:.2f}")
        if total_sales == 0 and payout_amount != 0:
            flags.append("Zero Total Sales but non-zero Payout Amount")

        is_flagged  = 1 if flags else 0
        flag_reason = "; ".join(flags) if flags else None

        parsed.append({
            "row_num":          row_num - 1,
            "payout_id":        payout_id,
            "brand":            brand,
            "payout_date":      payout_date,
            "tx_status":        tx_status,
            "total_sales":      round(total_sales, 2),
            "sponsored_voucher":round(sponsored_voucher, 2),
            "fees_incl_gst":    round(fees_incl_gst, 2),
            "fees_ex_gst":      round(fees_ex_gst, 2),
            "gst_on_fees":      round(gst_on_fees, 2),
            "rebates":          round(rebates, 2),
            "payout_amount":    round(payout_amount, 2),
            "is_flagged":       is_flagged,
            "flag_reason":      flag_reason,
        })

    return parsed



# ── Lazada fee-name → category mapping ────────────────────────────────────────
_LAZADA_REVENUE_FEES = {
    "Item Price Credit", "Reversal Item Price",
    "Shipping Fee (Paid By Customer)",
}
_LAZADA_PLATFORM_FEES = {
    "Commission", "Reversal Commission",
    "Payment Fee", "Payment Fee Credit",
    "SPA program Fee", "Reversal of SPA Program Fee",
    "LazCoins Discount Promotion Fee", "Reversal of LazCoins Discount Promotion Fee",
    "Sponsored Affiliates",
}
_LAZADA_SELLER_DISCOUNT_FEES = {
    "LazCoins Discount", "Reversal of LazCoins Discount",
    "Promotional Charges Vouchers", "Reversal Promotional Charges Vouchers",
    "Promotional Charges Flexi-Combo",
}
_LAZADA_SHIPPING_FEES = {
    "Shipping Fee Paid by Seller",
    "Reversal of Shipping Fee (Charged by Lazada)",
    "Shipping Fee Refund to Customer",
    "Shipping Fee Voucher (by Lazada)",
    "Shipping Fee Voucher Refund to Laz",
}


def _lazada_fee_category(fee_name):
    if fee_name in _LAZADA_REVENUE_FEES:        return "Revenue"
    if fee_name in _LAZADA_PLATFORM_FEES:       return "Platform Fee"
    if fee_name in _LAZADA_SELLER_DISCOUNT_FEES: return "Seller Discount"
    if fee_name in _LAZADA_SHIPPING_FEES:       return "Shipping"
    return "Other"


def _parse_lazada_income(file_bytes, period_month=None):
    """Parse a Lazada Income Details (.xlsx) export.

    period_month: 'YYYY-MM' string — used to flag deferred lines whose
    Release Date falls outside the current accounting period.

    Returns a list of dicts, one per income line, with keys:
      row_num, fee_name, fee_category,
      amount_incl, gst_amount, amount_ex_gst, vat_raw,
      order_number, seller_sku, statement_number, release_date_str,
      is_flagged, flag_reason
    """
    if not _HAS_XL:
        raise ValueError("openpyxl is required to parse Lazada files. Install it and restart.")

    # Note: read_only=True truncates rows for this Lazada file format — use data_only only
    wb = _xl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return []

    headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    expected = ["Fee Name", "Amount(Include Tax)", "VAT Amount", "Release Date", "Statement Number"]
    missing  = [e for e in expected if e not in headers]
    if missing:
        raise ValueError(
            f"Lazada Income file missing expected columns: {missing}. "
            "Make sure you are uploading the Income Details file."
        )

    col = {h: i for i, h in enumerate(headers)}
    parsed = []

    def _gcell(row, name):
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    def _gfloat(row, name):
        v = _gcell(row, name)
        if v is None: return 0.0
        try: return float(str(v).replace(",", "").strip())
        except: return 0.0

    # Determine period month number for deferred detection
    period_month_num = None
    period_year_num  = None
    if period_month:
        try:
            parts = period_month.split("-")
            period_year_num  = int(parts[0])
            period_month_num = int(parts[1])
        except Exception:
            pass

    for row_num, row in enumerate(all_rows[1:], start=2):
        if not row or len(row) <= col.get("Fee Name", 3):
            continue
        fee_name_raw = _gcell(row, "Fee Name")
        if fee_name_raw is None:
            continue
        fee_name = str(fee_name_raw).strip()
        if not fee_name:
            continue

        amount_incl = _gfloat(row, "Amount(Include Tax)")
        vat_raw     = _gfloat(row, "VAT Amount")

        # Release date — openpyxl may return datetime or string
        rel_raw = _gcell(row, "Release Date")
        release_date_str = ""
        rel_month = None
        rel_year  = None
        if rel_raw is not None:
            try:
                if hasattr(rel_raw, "strftime"):
                    release_date_str = rel_raw.strftime("%Y-%m-%d")
                    rel_month = rel_raw.month
                    rel_year  = rel_raw.year
                else:
                    s = str(rel_raw).strip()
                    if s:
                        from datetime import datetime as _dt
                        for fmt in ("%d %b %Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                            try:
                                d = _dt.strptime(s, fmt)
                                release_date_str = d.strftime("%Y-%m-%d")
                                rel_month = d.month
                                rel_year  = d.year
                                break
                            except ValueError:
                                continue
            except Exception:
                pass

        # Fee category
        fee_category = _lazada_fee_category(fee_name)

        # GST calculation
        if fee_category == "Revenue":
            # GST is embedded in the price: use 9/109 formula
            gst_amount   = round(amount_incl * GST_RATE, 4)
            amount_ex_gst = round(amount_incl - gst_amount, 4)
        elif fee_category in ("Platform Fee", "Shipping"):
            # GST is in col F (vat_raw, always positive absolute value)
            if amount_incl < 0:
                # Expense: amount_incl is negative and INCLUDES the GST
                # amount_ex_gst = amount_incl + vat_raw (less negative = base fee)
                # gst_amount stored as positive (DR input tax)
                gst_amount    = round(vat_raw, 4)
                amount_ex_gst = round(amount_incl + vat_raw, 4)
            else:
                # Reversal / credit: amount_incl is positive
                # gst_amount stored as negative (reversal reduces input tax)
                gst_amount    = round(-vat_raw, 4)
                amount_ex_gst = round(amount_incl - vat_raw, 4)
        else:
            # Seller Discount and Other: no GST
            gst_amount    = 0.0
            amount_ex_gst = round(amount_incl, 4)

        # Deferred detection
        is_deferred  = 0
        flag_reason  = None
        if period_month_num is not None and rel_month is not None:
            if rel_month != period_month_num or (period_year_num and rel_year != period_year_num):
                is_deferred = 1
                flag_reason = f"Deferred: releasing {release_date_str}"

        stmt_number  = str(_gcell(row, "Statement Number") or "").strip()
        order_number = str(_gcell(row, "Order Number")     or "").strip()
        seller_sku   = str(_gcell(row, "Seller SKU")       or "").strip()

        parsed.append({
            "row_num":          row_num - 1,
            "fee_name":         fee_name,
            "fee_category":     fee_category,
            "amount_incl":      round(amount_incl, 4),
            "gst_amount":       gst_amount,
            "amount_ex_gst":    amount_ex_gst,
            "vat_raw":          round(vat_raw, 4),
            "order_number":     order_number,
            "seller_sku":       seller_sku,
            "statement_number": stmt_number,
            "release_date_str": release_date_str,
            "is_flagged":       is_deferred,
            "flag_reason":      flag_reason,
        })

    return parsed


def _parse_lazada_wallet(file_bytes, period_month=None):
    """Parse a Lazada Wallet (.xlsx) export.

    period_month: 'YYYY-MM' string — used to flag settlements from prior
    month statements (i.e. clearing prior-period AR).

    Returns a list of dicts, one per wallet row, with keys:
      row_num, sub_type, wallet_type, amount,
      transaction_number, transaction_time, remarks,
      is_flagged, flag_reason
    """
    if not _HAS_XL:
        raise ValueError("openpyxl is required to parse Lazada files.")

    # Note: read_only=True truncates rows for this Lazada file format — use data_only only
    wb = _xl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return []

    headers = [str(c).strip() if c is not None else "" for c in all_rows[0]]
    expected = ["Sub Type", "Amount", "Remarks", "Transaction Time"]
    missing  = [e for e in expected if e not in headers]
    if missing:
        raise ValueError(
            f"Lazada Wallet file missing expected columns: {missing}. "
            "Make sure you are uploading the Wallet file, not the Income Details file."
        )

    col = {h: i for i, h in enumerate(headers)}
    parsed = []

    # Determine current period's month for prior-period detection
    period_month_num = None
    if period_month:
        try:
            period_month_num = int(period_month.split("-")[1])
        except Exception:
            pass

    for row_num, row in enumerate(all_rows[1:], start=2):
        if not row or len(row) <= col.get("Sub Type", 3):
            continue

        def _g(name):
            idx = col.get(name)
            return row[idx] if idx is not None and idx < len(row) else None

        sub_type_raw = _g("Sub Type")
        if sub_type_raw is None:
            continue
        sub_type = str(sub_type_raw).strip()
        if not sub_type:
            continue

        wallet_type = str(_g("Type") or "").strip()
        remarks     = str(_g("Remarks") or "").strip()
        txn_num     = str(_g("Transaction Number") or "").strip()
        txn_time    = str(_g("Transaction Time")   or "").strip()

        # Parse amount — may have "+" prefix and "," separators
        amt_raw = _g("Amount")
        try:
            amount = float(str(amt_raw or "0").replace(",", "").replace("+", "").strip())
        except Exception:
            amount = 0.0

        # Flag prior-month settlements
        is_prior = 0
        flag_reason = None
        if sub_type == "Settlement" and period_month_num is not None:
            # Extract statement number from remarks e.g. "Statement No. SGM5JSJ9-2026-0331"
            m = re.search(r"-(\d{4})$", remarks)
            if m:
                stmt_seq = m.group(1)      # e.g. "0331"
                stmt_mon = int(stmt_seq[:2])  # 03
                if stmt_mon != period_month_num:
                    is_prior = 1
                    flag_reason = (
                        f"Prior month settlement — stmt suffix {stmt_seq} "
                        f"(month {stmt_mon:02d} vs period month {period_month_num:02d}). "
                        f"Clear AR from previous period."
                    )

        parsed.append({
            "row_num":          row_num - 1,
            "sub_type":         sub_type,
            "wallet_type":      wallet_type,
            "amount":           round(amount, 2),
            "transaction_number": txn_num,
            "transaction_time": txn_time,
            "remarks":          remarks,
            "is_flagged":       is_prior,
            "flag_reason":      flag_reason,
        })

    return parsed


@app.route("/sales")
@login_required
@roles_required(*SALES_ROLES)
def sales_list():
    db      = get_db()
    uploads = db.execute(
        "SELECT * FROM sales_uploads ORDER BY created_at DESC").fetchall()
    return render_template("sales.html", uploads=uploads, role=session["role"])


@app.route("/sales/methodology")
@login_required
@roles_required(*SALES_ROLES)
def sales_methodology():
    return render_template("sales_methodology.html", role=session["role"])


@app.route("/sales/upload", methods=["GET","POST"])
@login_required
@roles_required(*SALES_ROLES)
def sales_upload():
    db = get_db()
    if request.method == "POST":
        f           = request.form
        platform    = (f.get("platform") or "Other").strip()
        period_val  = (f.get("period") or "").strip() or None
        period_from = f.get("period_from") or None
        period_to   = f.get("period_to")   or None
        notes_val   = f.get("notes", "")

        # ── Lazada dual-file upload ──────────────────────────────────────────
        income_file = request.files.get("income_file")
        wallet_file = request.files.get("wallet_file")

        if platform.startswith("Lazada") and (
                (income_file and income_file.filename) or
                (wallet_file and wallet_file.filename)):

            income_uid = None
            wallet_uid = None
            messages   = []
            has_warning = False

            # Process Income Details file
            if income_file and income_file.filename:
                income_bytes = income_file.read()
                try:
                    income_rows = _parse_lazada_income(income_bytes, period_month=period_val)
                except ValueError as e:
                    flash(f"Lazada Income parse error: {e}", "danger")
                    return redirect(url_for("sales_upload"))
                if not income_rows:
                    flash("No data rows found in the Lazada Income Details file.", "danger")
                    return redirect(url_for("sales_upload"))

                ref = next_upload_ref(db)
                db.execute("""INSERT INTO sales_uploads
                    (upload_ref,platform,parser_type,file_type,period,period_from,period_to,
                     filename,total_rows,uploaded_by,notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (ref, platform, "lazada", "income", period_val, period_from, period_to,
                     income_file.filename, len(income_rows), session["username"], notes_val))
                income_uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                for r in income_rows:
                    db.execute("""INSERT INTO sales_lines
                        (upload_id, row_num,
                         product_name, brand, tx_type, tx_status,
                         tx_amount, gst_amount, gross_revenue, expenses_total,
                         payout_id, payout_date,
                         is_flagged, flag_reason)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (income_uid, r["row_num"],
                         r["order_number"], r["seller_sku"],
                         r["fee_name"], r["fee_category"],
                         r["amount_incl"], r["gst_amount"],
                         r["amount_ex_gst"], r["vat_raw"],
                         r["statement_number"], r["release_date_str"],
                         r["is_flagged"], r["flag_reason"]))
                deferred_count = sum(1 for r in income_rows if r["is_flagged"])
                msg = f"Income Details: {len(income_rows)} lines parsed."
                if deferred_count:
                    msg += f" ⚠ {deferred_count} deferred line(s)."
                    has_warning = True
                messages.append(msg)

            # Process Wallet file
            if wallet_file and wallet_file.filename:
                wallet_bytes = wallet_file.read()
                try:
                    wallet_rows = _parse_lazada_wallet(wallet_bytes, period_month=period_val)
                except ValueError as e:
                    flash(f"Lazada Wallet parse error: {e}", "danger")
                    return redirect(url_for("sales_upload"))
                if not wallet_rows:
                    flash("No data rows found in the Lazada Wallet file.", "danger")
                    return redirect(url_for("sales_upload"))

                ref = next_upload_ref(db)
                db.execute("""INSERT INTO sales_uploads
                    (upload_ref,platform,parser_type,file_type,period,period_from,period_to,
                     filename,total_rows,uploaded_by,notes)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (ref, platform, "lazada", "wallet", period_val, period_from, period_to,
                     wallet_file.filename, len(wallet_rows), session["username"], notes_val))
                wallet_uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]
                for r in wallet_rows:
                    db.execute("""INSERT INTO sales_lines
                        (upload_id, row_num,
                         product_name, tx_type, tx_amount,
                         brand, payout_id, payout_date,
                         is_flagged, flag_reason)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (wallet_uid, r["row_num"],
                         r["sub_type"], r["wallet_type"], r["amount"],
                         r["remarks"], r["transaction_number"], r["transaction_time"],
                         r["is_flagged"], r["flag_reason"]))
                prior_count = sum(1 for r in wallet_rows if r["is_flagged"])
                msg = f"Wallet: {len(wallet_rows)} movements parsed."
                if prior_count:
                    msg += f" ⚠ {prior_count} prior-month settlement(s) flagged."
                    has_warning = True
                messages.append(msg)

            db.commit()
            flash(f"Lazada upload complete — " + " | ".join(messages),
                  "warning" if has_warning else "success")
            return redirect(url_for("sales_report",
                                    period=period_val or "",
                                    platform=platform))

        # ── Single-file upload (Atome, generic platforms) ────────────────────
        file      = request.files.get("sales_file")
        if not file or not file.filename:
            flash("Please select a file to upload.", "danger")
            return redirect(url_for("sales_upload"))

        file_bytes = file.read()
        ref        = next_upload_ref(db)
        file_type  = (f.get("file_type") or "transaction").strip()

        # ── Platform-specific parsers ────────────────────────────────────────
        if platform == "Atome" and file_type == "transaction":
            try:
                parsed_rows = _parse_atome(file_bytes)
            except ValueError as e:
                flash(f"Atome parse error: {e}", "danger")
                return redirect(url_for("sales_upload"))
            if not parsed_rows:
                flash("No data rows found in the Atome file.", "danger")
                return redirect(url_for("sales_upload"))

            db.execute("""INSERT INTO sales_uploads
                (upload_ref,platform,parser_type,file_type,period,period_from,period_to,
                 filename,total_rows,uploaded_by,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (ref, platform, "atome", "transaction", period_val, period_from, period_to,
                 file.filename, len(parsed_rows), session["username"], f.get("notes","")))
            uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

            for r in parsed_rows:
                db.execute("""INSERT INTO sales_lines
                    (upload_id, row_num, product_name,
                     gross_revenue, platform_fees, net_revenue,
                     brand, tx_type, tx_status, tx_amount, gst_amount,
                     amt_receivable, expenses_total, is_flagged, flag_reason)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (uid, r["row_num"], r["tx_id"],
                     r["revenue_ex_gst"], r["expenses_total"], r["amt_receivable"],
                     r["brand"], r["tx_type"], r["tx_status"],
                     r["tx_amount"], r["gst_amount"],
                     r["amt_receivable"], r["expenses_total"],
                     r["is_flagged"], r["flag_reason"]))

            db.commit()
            flag_count = sum(1 for r in parsed_rows if r["is_flagged"])
            msg = f"Atome Transaction upload complete — {len(parsed_rows)} rows parsed."
            if flag_count:
                msg += f" ⚠ {flag_count} row(s) flagged for review."
            flash(msg, "success" if not flag_count else "warning")
            return redirect(url_for("sales_detail", upload_id=uid))

        elif platform == "Atome" and file_type == "payment":
            try:
                parsed_rows = _parse_atome_payment(file_bytes)
            except ValueError as e:
                flash(f"Atome Payment parse error: {e}", "danger")
                return redirect(url_for("sales_upload"))
            if not parsed_rows:
                flash("No payout rows found in the Atome Payment file.", "danger")
                return redirect(url_for("sales_upload"))

            db.execute("""INSERT INTO sales_uploads
                (upload_ref,platform,parser_type,file_type,period,period_from,period_to,
                 filename,total_rows,uploaded_by,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (ref, platform, "atome", "payment", period_val, period_from, period_to,
                 file.filename, len(parsed_rows), session["username"], f.get("notes","")))
            uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

            for r in parsed_rows:
                db.execute("""INSERT INTO sales_lines
                    (upload_id, row_num, product_name,
                     brand, tx_type, tx_status,
                     tx_amount, amt_receivable, net_revenue,
                     expenses_total, gst_amount,
                     payout_id, payout_date,
                     sponsored_voucher, rebates,
                     fees_incl_gst, fees_ex_gst, gst_on_fees, payout_amount,
                     is_flagged, flag_reason)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (uid, r["row_num"], r["payout_id"],
                     r["brand"], "PAYOUT", r["tx_status"],
                     r["total_sales"], r["payout_amount"], r["payout_amount"],
                     r["fees_incl_gst"], r["gst_on_fees"],
                     r["payout_id"], r["payout_date"],
                     r["sponsored_voucher"], r["rebates"],
                     r["fees_incl_gst"], r["fees_ex_gst"], r["gst_on_fees"], r["payout_amount"],
                     r["is_flagged"], r["flag_reason"]))

            db.commit()
            flag_count = sum(1 for r in parsed_rows if r["is_flagged"])
            msg = f"Atome Payment upload complete — {len(parsed_rows)} payout rows parsed."
            if flag_count:
                msg += f" ⚠ {flag_count} row(s) flagged for review."
            flash(msg, "success" if not flag_count else "warning")
            return redirect(url_for("sales_detail", upload_id=uid))

        elif platform == "Atome" and file_type == "sku_sales":
            flash("SKU Sales file type for Atome is not yet configured. Please upload once the order file format is confirmed.", "warning")
            return redirect(url_for("sales_upload"))

        elif platform.startswith("Lazada") and file_type == "income":
            try:
                parsed_rows = _parse_lazada_income(file_bytes, period_month=period_val)
            except ValueError as e:
                flash(f"Lazada Income parse error: {e}", "danger")
                return redirect(url_for("sales_upload"))
            if not parsed_rows:
                flash("No data rows found in the Lazada Income Details file.", "danger")
                return redirect(url_for("sales_upload"))

            db.execute("""INSERT INTO sales_uploads
                (upload_ref,platform,parser_type,file_type,period,period_from,period_to,
                 filename,total_rows,uploaded_by,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (ref, platform, "lazada", "income", period_val, period_from, period_to,
                 file.filename, len(parsed_rows), session["username"], f.get("notes", "")))
            uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

            for r in parsed_rows:
                db.execute("""INSERT INTO sales_lines
                    (upload_id, row_num,
                     product_name, brand, tx_type, tx_status,
                     tx_amount, gst_amount, gross_revenue, expenses_total,
                     payout_id, payout_date,
                     is_flagged, flag_reason)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (uid, r["row_num"],
                     r["order_number"], r["seller_sku"],
                     r["fee_name"], r["fee_category"],
                     r["amount_incl"], r["gst_amount"],
                     r["amount_ex_gst"], r["vat_raw"],
                     r["statement_number"], r["release_date_str"],
                     r["is_flagged"], r["flag_reason"]))

            db.commit()
            deferred_count = sum(1 for r in parsed_rows if r["is_flagged"])
            msg = f"Lazada Income upload complete — {len(parsed_rows)} lines parsed."
            if deferred_count:
                msg += f" ⚠ {deferred_count} deferred line(s) (next-month release)."
            flash(msg, "success" if not deferred_count else "warning")
            return redirect(url_for("sales_detail", upload_id=uid))

        elif platform.startswith("Lazada") and file_type == "wallet":
            try:
                parsed_rows = _parse_lazada_wallet(file_bytes, period_month=period_val)
            except ValueError as e:
                flash(f"Lazada Wallet parse error: {e}", "danger")
                return redirect(url_for("sales_upload"))
            if not parsed_rows:
                flash("No data rows found in the Lazada Wallet file.", "danger")
                return redirect(url_for("sales_upload"))

            db.execute("""INSERT INTO sales_uploads
                (upload_ref,platform,parser_type,file_type,period,period_from,period_to,
                 filename,total_rows,uploaded_by,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (ref, platform, "lazada", "wallet", period_val, period_from, period_to,
                 file.filename, len(parsed_rows), session["username"], f.get("notes", "")))
            uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

            for r in parsed_rows:
                db.execute("""INSERT INTO sales_lines
                    (upload_id, row_num,
                     product_name, tx_type, tx_amount,
                     brand, payout_id, payout_date,
                     is_flagged, flag_reason)
                    VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (uid, r["row_num"],
                     r["sub_type"], r["wallet_type"], r["amount"],
                     r["remarks"], r["transaction_number"], r["transaction_time"],
                     r["is_flagged"], r["flag_reason"]))

            db.commit()
            prior_count = sum(1 for r in parsed_rows if r["is_flagged"])
            msg = f"Lazada Wallet upload complete — {len(parsed_rows)} movements parsed."
            if prior_count:
                msg += f" ⚠ {prior_count} prior-month settlement(s) flagged for AR knock-off."
            flash(msg, "success" if not prior_count else "warning")
            return redirect(url_for("sales_detail", upload_id=uid))

        else:
            # ── Generic parser for all other platforms ───────────────────────
            headers, rows = _parse_upload(file_bytes, file.filename)
            if not rows:
                flash("No data rows found in the file.", "danger")
                return redirect(url_for("sales_upload"))

            col_product = _guess_col(headers, ["product","item","sku","name","description"])
            col_qty     = _guess_col(headers, ["qty","quantity","units","sold"])
            col_price   = _guess_col(headers, ["price","unit price","selling"])
            col_gross   = _guess_col(headers, ["gross","revenue","amount","total","sales"])
            col_fees    = _guess_col(headers, ["fee","commission","platform","charge"])
            col_net     = _guess_col(headers, ["net","settlement","payout"])

            db.execute("""INSERT INTO sales_uploads
                (upload_ref,platform,parser_type,file_type,period,period_from,period_to,
                 filename,total_rows,uploaded_by,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (ref, platform, "generic", file_type, period_val, period_from, period_to,
                 file.filename, len(rows), session["username"], f.get("notes","")))
            uid = db.execute("SELECT last_insert_rowid()").fetchone()[0]

            for i, row in enumerate(rows):
                def gv(col): return row.get(col,"") if col else ""
                def gf(col):
                    v = gv(col)
                    try: return float(str(v).replace(",","").replace("$","").strip() or 0)
                    except: return 0.0
                db.execute("""INSERT INTO sales_lines
                    (upload_id,row_num,product_name,qty_sold,unit_selling_price,
                     gross_revenue,platform_fees,net_revenue)
                    VALUES (?,?,?,?,?,?,?,?)""",
                    (uid, i+1, gv(col_product), gf(col_qty), gf(col_price),
                     gf(col_gross), gf(col_fees), gf(col_net)))

            db.commit()
            flash(f"Uploaded {len(rows)} rows as {ref}.", "success")
            return redirect(url_for("sales_detail", upload_id=uid))

    return render_template("sales_upload.html",
                           role=session["role"],
                           today=date.today().isoformat(),
                           gateway_platforms=sorted(GATEWAY_PLATFORMS),
                           order_platforms=sorted(ORDER_PLATFORMS))

@app.route("/sales/<int:upload_id>")
@login_required
@roles_required(*SALES_ROLES)
def sales_detail(upload_id):
    db     = get_db()
    upload = db.execute("SELECT * FROM sales_uploads WHERE id=?", (upload_id,)).fetchone()
    if not upload:
        flash("Upload not found.", "danger"); return redirect(url_for("sales_list"))
    lines  = db.execute(
        "SELECT * FROM sales_lines WHERE upload_id=? ORDER BY row_num", (upload_id,)).fetchall()
    items  = db.execute("SELECT item_id, name FROM items WHERE status='Active' ORDER BY item_id").fetchall()

    parser_type = upload["parser_type"] or "generic"
    file_type   = upload["file_type"]   or "transaction"

    # ── Atome Transaction summary ────────────────────────────────────────────
    atome_summary         = None
    atome_payment_summary = None
    lazada_income_summary = None
    lazada_wallet_summary = None

    if parser_type == "atome" and file_type == "transaction":
        from collections import defaultdict
        brand_data = defaultdict(lambda: {
            "merch_count": 0, "refund_count": 0,
            "tx_amount": 0.0, "gst_amount": 0.0, "revenue_ex_gst": 0.0,
            "cash_received": 0.0, "receivables": 0.0, "expenses": 0.0
        })
        totals = {
            "merch_count": 0, "refund_count": 0,
            "tx_amount": 0.0, "gst_amount": 0.0, "revenue_ex_gst": 0.0,
            "cash_received": 0.0, "receivables": 0.0, "expenses": 0.0
        }
        flagged_lines = []
        for l in lines:
            b          = brand_data[l["brand"] or "Unknown"]
            txa        = l["tx_amount"]    or 0.0
            gst        = l["gst_amount"]   or 0.0
            rev        = l["gross_revenue"] or 0.0   # stored as revenue_ex_gst
            exp        = l["expenses_total"] or 0.0
            ar         = l["amt_receivable"] or 0.0
            is_merch   = (l["tx_type"]   or "").upper() == "MERCHANDISE"
            is_settled = (l["tx_status"] or "").upper() == "FULLY_SETTLED"

            for d in (b, totals):
                d["tx_amount"]      += txa
                d["gst_amount"]     += gst
                d["revenue_ex_gst"] += rev
                d["expenses"]       += exp
                if is_settled: d["cash_received"] += ar
                else:          d["receivables"]   += ar
                if is_merch: d["merch_count"] += 1
                else:        d["refund_count"] += 1

            if l["is_flagged"]:
                flagged_lines.append(l)

        for d in list(brand_data.values()) + [totals]:
            for k in ("tx_amount","gst_amount","revenue_ex_gst","cash_received","receivables","expenses"):
                d[k] = round(d[k], 2)

        # Accounting entries for transaction file
        t = totals
        gst_on_revenue      = round(t["gst_amount"], 2)
        revenue_ex_gst      = round(t["revenue_ex_gst"], 2)
        fees_abs            = round(abs(t["expenses"]), 2)
        fees_ex_gst         = round(fees_abs / 1.09, 2)
        gst_on_fees         = round(fees_abs - fees_ex_gst, 2)
        ar_confirmed        = round(t["receivables"], 2)

        atome_summary = {
            "brands":           dict(brand_data),
            "totals":           totals,
            "flagged_lines":    flagged_lines,
            "gst_on_revenue":   gst_on_revenue,
            "revenue_ex_gst":   revenue_ex_gst,
            "fees_abs":         fees_abs,
            "fees_ex_gst":      fees_ex_gst,
            "gst_on_fees":      gst_on_fees,
            "ar_confirmed":     ar_confirmed,
        }

    # ── Atome Payment summary ────────────────────────────────────────────────
    elif parser_type == "atome" and file_type == "payment":
        from collections import defaultdict
        brand_data = defaultdict(lambda: {
            "row_count": 0, "total_sales": 0.0,
            "sponsored_voucher": 0.0, "fees_incl_gst": 0.0,
            "fees_ex_gst": 0.0, "gst_on_fees": 0.0,
            "rebates": 0.0, "payout_amount": 0.0
        })
        totals = {
            "row_count": 0, "total_sales": 0.0,
            "sponsored_voucher": 0.0, "fees_incl_gst": 0.0,
            "fees_ex_gst": 0.0, "gst_on_fees": 0.0,
            "rebates": 0.0, "payout_amount": 0.0
        }
        flagged_lines = []
        for l in lines:
            b = brand_data[l["brand"] or "Unknown"]
            for d in (b, totals):
                d["row_count"]        += 1
                d["total_sales"]      += l["tx_amount"]       or 0.0
                d["sponsored_voucher"]+= l["sponsored_voucher"] or 0.0
                d["fees_incl_gst"]   += l["fees_incl_gst"]   or 0.0
                d["fees_ex_gst"]     += l["fees_ex_gst"]     or 0.0
                d["gst_on_fees"]     += l["gst_on_fees"]     or 0.0
                d["rebates"]         += l["rebates"]         or 0.0
                d["payout_amount"]   += l["payout_amount"]   or 0.0
            if l["is_flagged"]:
                flagged_lines.append(l)

        for d in list(brand_data.values()) + [totals]:
            for k in ("total_sales","sponsored_voucher","fees_incl_gst","fees_ex_gst","gst_on_fees","rebates","payout_amount"):
                d[k] = round(d[k], 2)

        atome_payment_summary = {
            "brands":        dict(brand_data),
            "totals":        totals,
            "flagged_lines": flagged_lines,
        }

    # ── Lazada Income summary ────────────────────────────────────────────────
    elif parser_type == "lazada" and file_type == "income":
        from collections import defaultdict

        fee_data = defaultdict(lambda: {
            "category": "", "amount_incl": 0.0, "gst_amount": 0.0,
            "amount_ex_gst": 0.0, "vat_raw": 0.0, "row_count": 0
        })
        cat_totals = {
            "Revenue":         {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
            "Platform Fee":    {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
            "Seller Discount": {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
            "Shipping":        {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
            "Other":           {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
        }
        # SKU summary — for Item Price Credit lines only
        sku_data   = defaultdict(lambda: {"row_count": 0, "amount_incl": 0.0,
                                           "gst_amount": 0.0, "amount_ex_gst": 0.0})
        deferred_lines  = []
        released_lines  = []

        for l in lines:
            fee_name = l["tx_type"]    or ""
            fee_cat  = l["tx_status"]  or "Other"
            amt      = l["tx_amount"]  or 0.0
            gst_val  = l["gst_amount"] or 0.0
            ex_gst   = l["gross_revenue"]  or 0.0
            vat_raw  = l["expenses_total"] or 0.0

            fd = fee_data[fee_name]
            fd["category"]     = fee_cat
            fd["amount_incl"]  += amt
            fd["gst_amount"]   += gst_val
            fd["amount_ex_gst"]+= ex_gst
            fd["vat_raw"]      += vat_raw
            fd["row_count"]    += 1

            ct = cat_totals.get(fee_cat, cat_totals["Other"])
            ct["amount_incl"]  += amt
            ct["gst_amount"]   += gst_val
            ct["amount_ex_gst"]+= ex_gst
            ct["vat_raw"]      += vat_raw

            if l["is_flagged"]:
                deferred_lines.append(l)
            else:
                released_lines.append(l)

            if fee_name == "Item Price Credit" and (l["brand"] or "").strip():
                sd = sku_data[l["brand"].strip()]
                sd["row_count"]    += 1
                sd["amount_incl"]  += amt
                sd["gst_amount"]   += gst_val
                sd["amount_ex_gst"]+= ex_gst

        # Round all accumulated values
        for fd in fee_data.values():
            for k in ("amount_incl", "gst_amount", "amount_ex_gst", "vat_raw"):
                fd[k] = round(fd[k], 2)
        for ct in cat_totals.values():
            for k in ct:
                ct[k] = round(ct[k], 2)
        for sd in sku_data.values():
            for k in ("amount_incl", "gst_amount", "amount_ex_gst"):
                sd[k] = round(sd[k], 2)

        ar_released  = round(sum(l["tx_amount"] or 0 for l in released_lines), 2)
        deferred_net = round(sum(l["tx_amount"] or 0 for l in deferred_lines), 2)

        # Build COGS lookup for known SKUs
        sku_costs = {}
        for sku_id in sku_data:
            cost = next_fifo_cost(db, sku_id)
            if cost > 0:
                sku_costs[sku_id] = round(cost, 4)

        lazada_income_summary = {
            "fee_data":       dict(sorted(fee_data.items(), key=lambda x: x[1]['category'])),
            "cat_totals":     cat_totals,
            "sku_data":       dict(sku_data),
            "sku_costs":      sku_costs,
            "deferred_lines": deferred_lines,
            "released_lines": released_lines,
            "ar_released":    ar_released,
            "deferred_net":   deferred_net,
        }

    # ── Lazada Wallet summary ────────────────────────────────────────────────
    elif parser_type == "lazada" and file_type == "wallet":
        sub_totals   = {}   # sub_type -> {count, total}
        prior_lines  = []
        current_lines = []

        for l in lines:
            sub  = l["product_name"] or "Unknown"
            amt  = l["tx_amount"] or 0.0
            if sub not in sub_totals:
                sub_totals[sub] = {"count": 0, "total": 0.0}
            sub_totals[sub]["count"] += 1
            sub_totals[sub]["total"] += amt

            if l["is_flagged"]:
                prior_lines.append(l)
            else:
                current_lines.append(l)

        for st in sub_totals.values():
            st["total"] = round(st["total"], 2)

        settlement_total  = round(sub_totals.get("Settlement", {}).get("total", 0.0), 2)
        withdrawal_total  = round(sub_totals.get("Auto Withdrawal", {}).get("total", 0.0), 2)
        topup_total       = round(sub_totals.get("Sponsored Solutions Top-up", {}).get("total", 0.0), 2)
        prior_total       = round(sum(l["tx_amount"] or 0 for l in prior_lines), 2)
        current_ar_total  = round(settlement_total - prior_total, 2)

        lazada_wallet_summary = {
            "sub_totals":       sub_totals,
            "settlement_total": settlement_total,
            "withdrawal_total": withdrawal_total,
            "topup_total":      topup_total,
            "prior_lines":      prior_lines,
            "prior_total":      prior_total,
            "current_ar_total": current_ar_total,
            "all_lines":        list(lines),
        }

    # ── Generic summary ──────────────────────────────────────────────────────
    total_revenue = sum(l["gross_revenue"] or 0 for l in lines)
    total_net     = sum(l["net_revenue"]   or 0 for l in lines)
    total_fees    = sum(l["platform_fees"] or 0 for l in lines)
    total_cogs    = 0.0
    for l in lines:
        if l["sku_id"]:
            total_cogs += (l["qty_sold"] or 0) * next_fifo_cost(db, l["sku_id"])

    return render_template("sales_detail.html",
                           upload=upload, lines=lines, items=items,
                           file_type=file_type,
                           total_revenue=total_revenue,
                           total_net=total_net, total_fees=total_fees,
                           total_cogs=total_cogs, role=session["role"],
                           atome_summary=atome_summary,
                           atome_payment_summary=atome_payment_summary,
                           lazada_income_summary=lazada_income_summary,
                           lazada_wallet_summary=lazada_wallet_summary)

@app.route("/sales/report")
@login_required
@roles_required(*SALES_ROLES)
def sales_report():
    """Combined Lazada report — shows Income Details + Wallet for a selected period/platform."""
    db       = get_db()
    period   = request.args.get("period",   "").strip()
    platform = request.args.get("platform", "").strip()

    # Build selector options — distinct period/platform combos that have lazada data
    combos = db.execute("""
        SELECT DISTINCT period, platform
        FROM sales_uploads
        WHERE parser_type='lazada' AND period IS NOT NULL AND platform IS NOT NULL
        ORDER BY period DESC, platform
    """).fetchall()

    # Available platforms for the filter (lazada only)
    lazada_platforms = ["Lazada (AndSons)", "Lazada (MM)"]

    income_upload  = None
    wallet_upload  = None
    income_summary = None
    wallet_summary = None

    if period and platform:
        income_upload = db.execute("""
            SELECT * FROM sales_uploads
            WHERE period=? AND platform=? AND parser_type='lazada' AND file_type='income'
            ORDER BY created_at DESC LIMIT 1
        """, (period, platform)).fetchone()

        wallet_upload = db.execute("""
            SELECT * FROM sales_uploads
            WHERE period=? AND platform=? AND parser_type='lazada' AND file_type='wallet'
            ORDER BY created_at DESC LIMIT 1
        """, (period, platform)).fetchone()

        # ── Build income summary ─────────────────────────────────────────────
        if income_upload:
            from collections import defaultdict
            inc_lines = db.execute(
                "SELECT * FROM sales_lines WHERE upload_id=? ORDER BY row_num",
                (income_upload["id"],)).fetchall()

            fee_data = defaultdict(lambda: {
                "category": "", "amount_incl": 0.0, "gst_amount": 0.0,
                "amount_ex_gst": 0.0, "vat_raw": 0.0, "row_count": 0
            })
            cat_totals = {
                "Revenue":         {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
                "Platform Fee":    {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
                "Seller Discount": {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
                "Shipping":        {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
                "Other":           {"amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0, "vat_raw": 0.0},
            }
            sku_data       = defaultdict(lambda: {"row_count": 0, "amount_incl": 0.0, "gst_amount": 0.0, "amount_ex_gst": 0.0})
            deferred_lines = []
            released_lines = []
            for l in inc_lines:
                fee_name = l["tx_type"]    or ""
                fee_cat  = l["tx_status"]  or "Other"
                amt      = l["tx_amount"]  or 0.0
                gst_val  = l["gst_amount"] or 0.0
                ex_gst   = l["gross_revenue"]  or 0.0
                vat_raw  = l["expenses_total"] or 0.0
                fd = fee_data[fee_name]
                fd["category"]     = fee_cat
                fd["amount_incl"]  += amt;  fd["gst_amount"]   += gst_val
                fd["amount_ex_gst"]+= ex_gst; fd["vat_raw"]    += vat_raw
                fd["row_count"]    += 1
                ct = cat_totals.get(fee_cat, cat_totals["Other"])
                ct["amount_incl"] += amt;  ct["gst_amount"]   += gst_val
                ct["amount_ex_gst"] += ex_gst; ct["vat_raw"]  += vat_raw
                if l["is_flagged"]: deferred_lines.append(l)
                else:               released_lines.append(l)
                if fee_name == "Item Price Credit" and (l["brand"] or "").strip():
                    sd = sku_data[l["brand"].strip()]
                    sd["row_count"] += 1; sd["amount_incl"] += amt
                    sd["gst_amount"] += gst_val; sd["amount_ex_gst"] += ex_gst
            for fd in fee_data.values():
                for k in ("amount_incl","gst_amount","amount_ex_gst","vat_raw"): fd[k] = round(fd[k], 2)
            for ct in cat_totals.values():
                for k in ct: ct[k] = round(ct[k], 2)
            for sd in sku_data.values():
                for k in ("amount_incl","gst_amount","amount_ex_gst"): sd[k] = round(sd[k], 2)
            ar_released  = round(sum(l["tx_amount"] or 0 for l in released_lines), 2)
            deferred_net = round(sum(l["tx_amount"] or 0 for l in deferred_lines), 2)
            sku_costs = {}
            for sku_id in sku_data:
                cost = next_fifo_cost(db, sku_id)
                if cost > 0: sku_costs[sku_id] = round(cost, 4)
            income_summary = {
                "upload": income_upload, "lines": inc_lines,
                "fee_data": dict(sorted(fee_data.items(), key=lambda x: x[1]['category'])), "cat_totals": cat_totals,
                "sku_data": dict(sku_data), "sku_costs": sku_costs,
                "deferred_lines": deferred_lines, "released_lines": released_lines,
                "ar_released": ar_released, "deferred_net": deferred_net,
            }

        # ── Build wallet summary ─────────────────────────────────────────────
        if wallet_upload:
            wal_lines    = db.execute(
                "SELECT * FROM sales_lines WHERE upload_id=? ORDER BY row_num",
                (wallet_upload["id"],)).fetchall()
            sub_totals   = {}
            prior_lines  = []
            current_lines = []
            for l in wal_lines:
                sub = l["product_name"] or "Unknown"
                amt = l["tx_amount"] or 0.0
                if sub not in sub_totals:
                    sub_totals[sub] = {"count": 0, "total": 0.0}
                sub_totals[sub]["count"] += 1
                sub_totals[sub]["total"] += amt
                if l["is_flagged"]: prior_lines.append(l)
                else:               current_lines.append(l)
            for st in sub_totals.values(): st["total"] = round(st["total"], 2)
            settlement_total = round(sub_totals.get("Settlement", {}).get("total", 0.0), 2)
            withdrawal_total = round(sub_totals.get("Auto Withdrawal", {}).get("total", 0.0), 2)
            topup_total      = round(sub_totals.get("Sponsored Solutions Top-up", {}).get("total", 0.0), 2)
            prior_total      = round(sum(l["tx_amount"] or 0 for l in prior_lines), 2)
            current_ar_total = round(settlement_total - prior_total, 2)
            wallet_summary = {
                "upload": wallet_upload, "lines": wal_lines,
                "sub_totals": sub_totals,
                "settlement_total": settlement_total, "withdrawal_total": withdrawal_total,
                "topup_total": topup_total, "prior_lines": prior_lines,
                "prior_total": prior_total, "current_ar_total": current_ar_total,
                "all_lines": list(wal_lines),
            }

    return render_template("sales_report.html",
                           period=period, platform=platform,
                           combos=combos,
                           lazada_platforms=lazada_platforms,
                           income_summary=income_summary,
                           wallet_summary=wallet_summary,
                           role=session["role"])


@app.route("/sales/<int:upload_id>/map", methods=["POST"])
@login_required
@roles_required(*SALES_ROLES)
def sales_map_line(upload_id):
    db = get_db()
    for key, val in request.form.items():
        if key.startswith("sku_"):
            line_id = int(key[4:])
            db.execute("UPDATE sales_lines SET sku_id=? WHERE id=? AND upload_id=?",
                       (val.strip() or None, line_id, upload_id))
    db.commit()
    flash("SKU mappings saved.", "success")
    return redirect(url_for("sales_detail", upload_id=upload_id))

@app.route("/sales/<int:upload_id>/delete", methods=["POST"])
@login_required
@roles_required(*SALES_ROLES)
def sales_delete(upload_id):
    db = get_db()
    db.execute("DELETE FROM sales_lines   WHERE upload_id=?", (upload_id,))
    db.execute("DELETE FROM sales_uploads WHERE id=?",        (upload_id,))
    db.commit()
    flash("Upload deleted.", "success")
    return redirect(url_for("sales_list"))


# ── COMPANY SETTINGS ─────────────────────────────────────────────────────────
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "static", "uploads")
ALLOWED_LOGO_EXT = {"png", "jpg", "jpeg", "gif", "svg", "webp"}

PO_DOCS_FOLDER = os.path.join(os.path.dirname(__file__), "static", "po_docs")
ALLOWED_DOC_EXT = {"pdf", "png", "jpg", "jpeg", "gif", "webp", "doc", "docx", "xls", "xlsx", "csv", "txt"}
os.makedirs(PO_DOCS_FOLDER, exist_ok=True)

@app.route("/settings/company", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def company_settings():
    db = get_db()
    if request.method == "POST":
        section = request.form.get("section", "branding")

        if section == "branding":
            name = request.form.get("company_name", "").strip()
            if name:
                db.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('company_name',?)", (name,))
            logo = request.files.get("company_logo")
            if logo and logo.filename:
                ext = logo.filename.rsplit(".", 1)[-1].lower()
                if ext in ALLOWED_LOGO_EXT:
                    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                    fname = f"company_logo.{ext}"
                    logo.save(os.path.join(UPLOAD_FOLDER, fname))
                    db.execute("INSERT OR REPLACE INTO settings (key,value) VALUES ('company_logo',?)", (fname,))
            flash("Branding saved.", "success")

        elif section == "info":
            for key in ("company_phone", "company_email",
                        "company_hq_address", "company_warehouse_address"):
                val = request.form.get(key, "").strip()
                db.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (key, val))
            flash("Company information saved.", "success")

        db.commit()
        return redirect(url_for("company_settings"))

    s = {r["key"]: r["value"] for r in db.execute("SELECT key,value FROM settings").fetchall()}
    return render_template("settings.html", s=s, role=session["role"])


# ─── MONTH-END CLOSE (MEC) ──────────────────────────────────────────────────

MEC_ROLES = ("admin","ops_manager","ops_planner","ops_exec","finance_exec","finance_manager")

def _period_bounds(period):
    """Return (period_start, next_period_start) as ISO date strings for YYYY-MM."""
    yr, mo = int(period[:4]), int(period[5:7])
    period_start = f"{yr}-{mo:02d}-01"
    if mo == 12:
        next_start = f"{yr+1}-01-01"
    else:
        next_start = f"{yr}-{mo+1:02d}-01"
    return period_start, next_start


def _book_balances(db, period):
    """Return dict {item_id: {opening, receipts_in, adj_in, adj_out, book}} for period."""
    period_start, next_start = _period_bounds(period)

    # Opening balances
    op_rows = db.execute(
        "SELECT item_id, opening_qty FROM period_opening_balances WHERE period=?",
        (period,)).fetchall()
    opening = {r["item_id"]: r["opening_qty"] or 0 for r in op_rows}

    if not opening:
        # First period ever — seed from all ledger activity before period start
        rows = db.execute("""
            SELECT item_id,
                   SUM(CASE WHEN qty_in  > 0 THEN qty_in  ELSE 0 END) AS ti,
                   SUM(CASE WHEN qty_out > 0 THEN qty_out ELSE 0 END) AS to_
            FROM inventory_ledger WHERE txn_date < ? GROUP BY item_id
        """, (period_start,)).fetchall()
        for r in rows:
            opening[r["item_id"]] = max(0, (r["ti"] or 0) - (r["to_"] or 0))

    # In-period movements
    mv_rows = db.execute("""
        SELECT item_id,
               SUM(CASE WHEN txn_type='PO Receipt'            THEN qty_in  ELSE 0 END) AS receipts_in,
               SUM(CASE WHEN txn_type LIKE '(+)%'           THEN qty_in  ELSE 0 END) AS adj_in,
               SUM(CASE WHEN txn_type LIKE '(-)%'           THEN qty_out ELSE 0 END) AS adj_out
        FROM inventory_ledger
        WHERE txn_date >= ? AND txn_date < ?
        GROUP BY item_id
    """, (period_start, next_start)).fetchall()
    mv = {r["item_id"]: r for r in mv_rows}

    all_ids = set(opening.keys()) | set(mv.keys())
    result = {}
    for iid in all_ids:
        op  = opening.get(iid, 0)
        m   = mv.get(iid)
        ri  = (m["receipts_in"] or 0) if m else 0
        ai  = (m["adj_in"]      or 0) if m else 0
        ao  = (m["adj_out"]     or 0) if m else 0
        result[iid] = {"opening": op, "receipts_in": ri,
                       "adj_in": ai,  "adj_out": ao,
                       "book": op + ri + ai - ao}
    return result


@app.route("/mec")
@login_required
@roles_required(*MEC_ROLES)
def mec_dashboard():
    db = get_db()
    periods = db.execute(
        "SELECT * FROM periods ORDER BY period DESC").fetchall()

    # Build status summary per period
    summary = {}
    for p in periods:
        sc = db.execute(
            "SELECT id, status FROM stockcount_uploads WHERE period=? ORDER BY id DESC LIMIT 1",
            (p["period"],)).fetchone()
        su = db.execute(
            "SELECT COUNT(*) AS c FROM sales_uploads WHERE period=?",
            (p["period"],)).fetchone()
        summary[p["period"]] = {
            "sc_id":     sc["id"]     if sc else None,
            "sc_status": sc["status"] if sc else None,
            "sales_count": su["c"]   if su else 0,
        }

    # Auto-suggest current period
    current_period = date.today().strftime("%Y-%m")
    return render_template("mec.html",
                           periods=periods, summary=summary,
                           current_period=current_period,
                           role=session["role"])


@app.route("/mec/stockcount/new", methods=["GET","POST"])
@login_required
@roles_required("admin","ops_manager","ops_planner","ops_exec")
def mec_stockcount_new():
    db = get_db()
    if request.method == "POST":
        period   = (request.form.get("period") or "").strip()
        notes    = request.form.get("notes","").strip()
        f        = request.files.get("csv_file")
        if not period:
            flash("Please select a period.", "danger")
            return redirect(url_for("mec_stockcount_new"))
        if not f or not f.filename:
            flash("Please upload a CSV file.", "danger")
            return redirect(url_for("mec_stockcount_new"))

        # Parse CSV / Excel
        raw = f.read()
        try:
            import csv as _csv, io
            text = raw.decode("utf-8-sig")
            reader = list(_csv.DictReader(io.StringIO(text)))
            headers = reader[0].keys() if reader else []
        except Exception:
            flash("Could not parse file. Please upload a valid CSV.", "danger")
            return redirect(url_for("mec_stockcount_new"))

        # Auto-detect columns
        def _gc(keys, candidates):
            for h in keys:
                for c in candidates:
                    if c.lower() in h.lower():
                        return h
            return None

        col_sku = _gc(headers, ["sku","item_id","item","product"])
        col_qty = _gc(headers, ["qty","quantity","physical","count","balance","stock"])
        if not col_sku or not col_qty:
            flash("Could not detect SKU / Quantity columns. Ensure headers contain 'sku'/'item' and 'qty'/'quantity'.", "danger")
            return redirect(url_for("mec_stockcount_new"))

        # Ensure period record exists
        db.execute("INSERT OR IGNORE INTO periods (period,status) VALUES (?,'Open')", (period,))

        # Check no approved stockcount already exists for this period
        existing = db.execute(
            "SELECT id FROM stockcount_uploads WHERE period=? AND status='Approved'",
            (period,)).fetchone()
        if existing:
            flash(f"An approved stockcount already exists for {period}. Cannot upload again.", "warning")
            return redirect(url_for("mec_stockcount_detail", upload_id=existing["id"]))

        upload_id = db.execute(
            """INSERT INTO stockcount_uploads (period,filename,uploaded_by,notes)
               VALUES (?,?,?,?)""",
            (period, f.filename, session["username"], notes)).lastrowid

        skipped = 0
        for row in reader:
            sku = str(row.get(col_sku) or "").strip().upper()
            try:
                qty = float(str(row.get(col_qty) or "0").replace(",",""))
            except ValueError:
                qty = 0
            if not sku:
                skipped += 1
                continue
            # Normalise to known items if possible
            item = db.execute("SELECT item_id FROM items WHERE item_id=?", (sku,)).fetchone()
            if not item:
                item = db.execute("SELECT item_id FROM items WHERE UPPER(item_id)=?", (sku,)).fetchone()
            final_sku = item["item_id"] if item else sku
            db.execute(
                "INSERT INTO stockcount_lines (upload_id,item_id,physical_qty) VALUES (?,?,?)",
                (upload_id, final_sku, qty))

        db.commit()
        msg = f"Stockcount uploaded for {period}."
        if skipped:
            msg += f" {skipped} blank rows skipped."
        flash(msg, "success")
        return redirect(url_for("mec_stockcount_detail", upload_id=upload_id))

    today = date.today()
    current_period = today.strftime("%Y-%m")
    months = [
        (date(today.year if today.month - i > 0 else today.year - 1,
              (today.month - i - 1) % 12 + 1, 1).strftime("%Y-%m"))
        for i in range(6)
    ]
    return render_template("mec_stockcount_new.html",
                           current_period=current_period,
                           months=months, role=session["role"])


@app.route("/mec/stockcount/<int:upload_id>")
@login_required
@roles_required(*MEC_ROLES)
def mec_stockcount_detail(upload_id):
    db = get_db()
    upload = db.execute(
        "SELECT * FROM stockcount_uploads WHERE id=?", (upload_id,)).fetchone()
    if not upload:
        flash("Stockcount not found.", "danger")
        return redirect(url_for("mec_dashboard"))

    period = upload["period"]
    lines  = db.execute(
        "SELECT * FROM stockcount_lines WHERE upload_id=? ORDER BY item_id",
        (upload_id,)).fetchall()

    # Book balances for the period
    book = _book_balances(db, period)

    # Item names
    items_map = {r["item_id"]: r["name"] for r in
                 db.execute("SELECT item_id, name FROM items").fetchall()}

    # Build variance rows — union of physical and book SKUs
    all_skus = set(l["item_id"] for l in lines) | set(book.keys())
    physical_map = {l["item_id"]: l["physical_qty"] or 0 for l in lines}

    rows = []
    for sku in sorted(all_skus):
        b    = book.get(sku, {"opening":0,"receipts_in":0,"adj_in":0,"adj_out":0,"book":0})
        phys = physical_map.get(sku, 0)
        var  = phys - b["book"]
        rows.append({
            "item_id":     sku,
            "item_name":   items_map.get(sku, "—"),
            "opening":     b["opening"],
            "receipts_in": b["receipts_in"],
            "adj_in":      b["adj_in"],
            "adj_out":     b["adj_out"],
            "book":        b["book"],
            "physical":    phys,
            "variance":    var,
        })

    period_obj = db.execute("SELECT * FROM periods WHERE period=?", (period,)).fetchone()
    return render_template("mec_stockcount_detail.html",
                           upload=upload, rows=rows,
                           period=period, period_obj=period_obj,
                           role=session["role"])


@app.route("/mec/stockcount/<int:upload_id>/approve", methods=["POST"])
@login_required
@roles_required("admin","ops_manager")
def mec_stockcount_approve(upload_id):
    db = get_db()
    upload = db.execute(
        "SELECT * FROM stockcount_uploads WHERE id=?", (upload_id,)).fetchone()
    if not upload or upload["status"] == "Approved":
        flash("Stockcount not found or already approved.", "warning")
        return redirect(url_for("mec_dashboard"))

    period = upload["period"]
    period_obj = db.execute("SELECT status FROM periods WHERE period=?", (period,)).fetchone()
    if period_obj and period_obj["status"] == "Closed":
        flash(f"Period {period} is already closed.", "danger")
        return redirect(url_for("mec_stockcount_detail", upload_id=upload_id))

    # Book balances vs physical
    book = _book_balances(db, period)
    lines = db.execute(
        "SELECT * FROM stockcount_lines WHERE upload_id=?", (upload_id,)).fetchall()
    physical_map = {l["item_id"]: l["physical_qty"] or 0 for l in lines}
    all_skus = set(book.keys()) | set(physical_map.keys())

    adj_count = 0
    rec_date  = date.today().isoformat()
    for sku in all_skus:
        b    = book.get(sku, {}).get("book", 0)
        phys = physical_map.get(sku, 0)
        var  = phys - b
        if abs(var) < 0.001:
            continue
        txn_id    = next_txn_id(db)
        txn_type  = "(+) Adjustment – Stockcount" if var > 0 else "(-) Adjustment – Stockcount"
        qty_in    = max(var, 0)
        qty_out   = max(-var, 0)
        # WACC for cost
        wacc_row  = db.execute("""
            SELECT SUM(qty_received * unit_cost) / NULLIF(SUM(qty_received),0) AS wacc
            FROM fifo_layers WHERE item_id=? AND qty_received > 0
        """, (sku,)).fetchone()
        unit_cost = (wacc_row["wacc"] or 0) if wacc_row else 0
        db.execute("""
            INSERT INTO inventory_ledger
              (txn_id,txn_date,txn_type,reference,item_id,qty_in,qty_out,unit_cost,posted_by,notes)
            VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (txn_id, rec_date, txn_type, f"STKCT-{period}",
             sku, qty_in, qty_out, unit_cost, session["username"],
             f"Stockcount variance for {period}"))
        if var > 0 and unit_cost > 0:
            existing = db.execute("SELECT COUNT(*) FROM fifo_layers").fetchone()[0]
            db.execute("""INSERT INTO fifo_layers
                (layer_id,item_id,po_reference,receipt_date,qty_received,unit_cost)
                VALUES (?,?,?,?,?,?)""",
                (f"SC-{existing+1:04d}", sku, f"STKCT-{period}", rec_date, var, unit_cost))
        adj_count += 1

    db.execute("""UPDATE stockcount_uploads
                  SET status='Approved', approved_by=?, approved_at=datetime('now')
                  WHERE id=?""", (session["username"], upload_id))
    db.commit()
    flash(f"Stockcount approved. {adj_count} variance adjustment(s) posted.", "success")
    return redirect(url_for("mec_stockcount_detail", upload_id=upload_id))


@app.route("/mec/<period>/reconciliation")
@login_required
@roles_required(*MEC_ROLES)
def mec_reconciliation(period):
    db = get_db()
    period_start, next_start = _period_bounds(period)
    period_obj = db.execute("SELECT * FROM periods WHERE period=?", (period,)).fetchone()

    # Approved stockcount for period
    sc = db.execute(
        "SELECT * FROM stockcount_uploads WHERE period=? AND status='Approved' ORDER BY id DESC LIMIT 1",
        (period,)).fetchone()
    sc_lines = []
    if sc:
        sc_lines = db.execute(
            "SELECT * FROM stockcount_lines WHERE upload_id=?", (sc["id"],)).fetchall()

    # Book balances
    book = _book_balances(db, period)
    items_map = {r["item_id"]: r["name"] for r in
                 db.execute("SELECT item_id, name FROM items").fetchall()}

    physical_map = {l["item_id"]: l["physical_qty"] or 0 for l in sc_lines}
    all_inv_skus = set(book.keys()) | set(physical_map.keys())

    inv_rows = []
    total_book = total_physical = total_variance = 0
    for sku in sorted(all_inv_skus):
        b    = book.get(sku, {"opening":0,"receipts_in":0,"adj_in":0,"adj_out":0,"book":0})
        phys = physical_map.get(sku, b["book"])  # if no physical, assume book
        var  = phys - b["book"]
        inv_rows.append({
            "item_id": sku, "item_name": items_map.get(sku,"—"),
            **b, "physical": phys, "variance": var
        })
        total_book     += b["book"]
        total_physical += phys
        total_variance += var

    # Sales uploads for this period
    sales_uploads_list = db.execute(
        "SELECT * FROM sales_uploads WHERE period=? ORDER BY created_at DESC",
        (period,)).fetchall()

    # COS per SKU: sum qty_sold across uploads × WACC
    cos_rows = db.execute("""
        SELECT sl.sku_id AS item_id,
               SUM(sl.qty_sold)         AS total_qty,
               SUM(sl.gross_revenue)    AS total_revenue,
               SUM(sl.net_revenue)      AS total_net_revenue
        FROM sales_lines sl
        JOIN sales_uploads su ON sl.upload_id = su.id
        WHERE su.period = ?
        GROUP BY sl.sku_id
    """, (period,)).fetchall()

    wacc_map = {}
    for r in db.execute("""
        SELECT item_id,
               SUM(qty_received * unit_cost) / NULLIF(SUM(qty_received),0) AS wacc
        FROM fifo_layers WHERE qty_received > 0 GROUP BY item_id
    """).fetchall():
        wacc_map[r["item_id"]] = r["wacc"] or 0

    cos_detail = []
    total_sales_qty = total_revenue = total_cos = 0
    for r in cos_rows:
        sku      = r["item_id"]
        wacc     = wacc_map.get(sku, 0)
        cos_val  = (r["total_qty"] or 0) * wacc
        cos_detail.append({
            "item_id":   sku,
            "item_name": items_map.get(sku,"—"),
            "qty":       r["total_qty"] or 0,
            "revenue":   r["total_revenue"] or 0,
            "net_revenue": r["total_net_revenue"] or 0,
            "wacc":      wacc,
            "cos":       cos_val,
        })
        total_sales_qty += r["total_qty"] or 0
        total_revenue   += r["total_revenue"] or 0
        total_cos       += cos_val

    gross_margin       = total_revenue - total_cos
    gm_pct             = (gross_margin / total_revenue * 100) if total_revenue else 0

    # Timing gap: units sold vs net inventory movement
    total_inv_movement = sum(b["receipts_in"] + b["adj_in"] - b["adj_out"] for b in book.values())
    timing_gap         = total_sales_qty - abs(total_variance)

    can_close = (
        sc is not None and
        (period_obj is None or period_obj["status"] != "Closed")
    )

    return render_template("mec_reconciliation.html",
                           period=period, period_obj=period_obj,
                           inv_rows=inv_rows, sc=sc,
                           total_book=total_book, total_physical=total_physical,
                           total_variance=total_variance,
                           sales_uploads_list=sales_uploads_list,
                           cos_detail=cos_detail,
                           total_sales_qty=total_sales_qty,
                           total_revenue=total_revenue,
                           total_cos=total_cos,
                           gross_margin=gross_margin, gm_pct=gm_pct,
                           total_inv_movement=total_inv_movement,
                           timing_gap=timing_gap,
                           can_close=can_close,
                           role=session["role"])


@app.route("/mec/<period>/close", methods=["POST"])
@login_required
@roles_required("admin","finance_manager")
def mec_close_period(period):
    db = get_db()
    period_obj = db.execute("SELECT * FROM periods WHERE period=?", (period,)).fetchone()
    if period_obj and period_obj["status"] == "Closed":
        flash(f"Period {period} is already closed.", "warning")
        return redirect(url_for("mec_reconciliation", period=period))

    sc = db.execute(
        "SELECT * FROM stockcount_uploads WHERE period=? AND status='Approved' ORDER BY id DESC LIMIT 1",
        (period,)).fetchone()
    if not sc:
        flash("Cannot close period: no approved stockcount found.", "danger")
        return redirect(url_for("mec_reconciliation", period=period))

    # Seed next period opening balances from this period's physical count
    _, next_start = _period_bounds(period)
    yr, mo = int(period[:4]), int(period[5:7])
    next_period = f"{yr+1}-01" if mo == 12 else f"{yr}-{mo+1:02d}"

    sc_lines = db.execute(
        "SELECT item_id, physical_qty FROM stockcount_lines WHERE upload_id=?",
        (sc["id"],)).fetchall()

    db.execute("INSERT OR IGNORE INTO periods (period,status) VALUES (?,'Open')",
               (next_period,))
    for line in sc_lines:
        db.execute("""
            INSERT INTO period_opening_balances (period, item_id, opening_qty)
            VALUES (?,?,?)
            ON CONFLICT(period,item_id) DO UPDATE SET opening_qty=excluded.opening_qty
        """, (next_period, line["item_id"], line["physical_qty"] or 0))

    # Close this period
    db.execute("""INSERT OR IGNORE INTO periods (period,status) VALUES (?,'Open')""",
               (period,))
    db.execute("""UPDATE periods SET status='Closed', closed_at=datetime('now'), closed_by=?
                  WHERE period=?""", (session["username"], period))
    db.commit()
    flash(f"Period {period} closed. Opening balances for {next_period} have been seeded.", "success")
    return redirect(url_for("mec_reconciliation", period=period))


# ─── DELIVERY CALENDAR ───────────────────────────────────────────────────────
import calendar as _cal

@app.route("/delivery-calendar")
@login_required
def delivery_calendar():
    from datetime import date as _date
    db = get_db()
    today = _date.today()

    # Parse requested month from query string (default = current month)
    try:
        year  = int(request.args.get("year",  today.year))
        month = int(request.args.get("month", today.month))
        if month < 1:  month = 12; year -= 1
        if month > 12: month = 1;  year += 1
    except ValueError:
        year, month = today.year, today.month

    search = request.args.get("q","").strip().lower()

    # All non-cancelled, non-draft POs with or without delivery date
    base_query = """
        SELECT po.id, po.po_number, po.status, po.expected_delivery_date,
               po.currency, s.name AS supplier_name,
               (SELECT COUNT(*) FROM po_lines WHERE po_id=po.id) AS line_count,
               (SELECT i2.name FROM po_lines pl2 JOIN items i2 ON pl2.item_id=i2.item_id
                WHERE pl2.po_id=po.id ORDER BY pl2.id LIMIT 1) AS first_item_name
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        WHERE po.status NOT IN ('Cancelled','Draft')
    """
    all_pos = db.execute(base_query).fetchall()

    # Build month calendar grid
    first_weekday, num_days = _cal.monthrange(year, month)
    month_str = f"{year:04d}-{month:02d}"

    # POs for this month
    calendar_pos = {}   # day -> list of pos
    for po in all_pos:
        d = po["expected_delivery_date"]
        if d and d.startswith(month_str):
            try:
                day = int(d[8:10])
                calendar_pos.setdefault(day, []).append(po)
            except (ValueError, IndexError):
                pass

    # POs without delivery date (unscheduled)
    no_date_pos = [po for po in all_pos if not po["expected_delivery_date"]]

    # Search filter across all POs
    search_results = []
    if search:
        search_results = [po for po in all_pos
                          if search in (po["po_number"] or "").lower()
                          or search in (po["supplier_name"] or "").lower()
                          or search in (po["first_item_name"] or "").lower()]

    # Prev / next month links
    prev_month = month - 1
    prev_year  = year
    if prev_month < 1:  prev_month = 12; prev_year  -= 1
    next_month_n = month + 1
    next_year  = year
    if next_month_n > 12: next_month_n = 1; next_year += 1

    month_name = _cal.month_name[month]

    return render_template("delivery_calendar.html",
        year=year, month=month, month_name=month_name,
        num_days=num_days, first_weekday=first_weekday,
        calendar_pos=calendar_pos,
        no_date_pos=no_date_pos,
        search_results=search_results,
        search=search,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month_n,
        today=today.isoformat())


@app.route("/pos/<int:po_id>/update_delivery_date", methods=["POST"])
@login_required
def update_delivery_date(po_id):
    from datetime import date as _date
    if session.get("role") not in ("admin","ops_planner","ops_manager"):
        return jsonify({"ok": False, "error": "Permission denied"}), 403
    db = get_db()
    new_date = request.json.get("expected_delivery_date","").strip() if request.is_json else request.form.get("expected_delivery_date","").strip()
    if not new_date:
        new_date = None
    db.execute("UPDATE purchase_orders SET expected_delivery_date=?, updated_at=datetime('now') WHERE id=?",
               (new_date, po_id))
    db.commit()
    return jsonify({"ok": True, "expected_delivery_date": new_date})


# ─── INVENTORY RUN-OUT ESTIMATOR ─────────────────────────────────────────────
@app.route("/demand-planning")
@app.route("/runout-estimator")
@login_required
def runout_estimator():
    from datetime import date as _date
    db = get_db()
    today = _date.today()

    # Parse requested period
    period = request.args.get("period", today.strftime("%Y-%m"))
    try:
        p_year, p_month = int(period[:4]), int(period[5:7])
    except (ValueError, IndexError):
        p_year, p_month = today.year, today.month
        period = today.strftime("%Y-%m")

    import calendar as _cal2
    prev_month = p_month - 1
    prev_year  = p_year
    if prev_month < 1: prev_month = 12; prev_year -= 1
    prev_period = f"{prev_year:04d}-{prev_month:02d}"

    items = db.execute("SELECT item_id, name FROM items WHERE status='Active' ORDER BY item_id").fetchall()

    # Opening balance for this period (from period_opening_balances, or FIFO fallback)
    opening_balances = {}
    for row in db.execute("SELECT item_id, opening_qty FROM period_opening_balances WHERE period=?", (period,)).fetchall():
        opening_balances[row["item_id"]] = row["opening_qty"]

    # If no MEC opening balance for this period, try prior period
    if not opening_balances:
        for row in db.execute("SELECT item_id, opening_qty FROM period_opening_balances WHERE period=?", (prev_period,)).fetchall():
            opening_balances[row["item_id"]] = row["opening_qty"]

    # Prior month actual outbound from sales uploads
    prior_outbound = {}
    for row in db.execute("""
        SELECT sl.sku_id, COALESCE(SUM(sl.qty_sold), 0) AS total_out
        FROM sales_lines sl
        JOIN sales_uploads su ON sl.upload_id = su.id
        WHERE su.period = ?
        GROUP BY sl.sku_id
    """, (prev_period,)).fetchall():
        if row["sku_id"]:
            prior_outbound[row["sku_id"]] = row["total_out"]

    # Prior month adjustment (variance from stockcount reconciliation)
    prior_adjustment = {}
    for row in db.execute("""
        SELECT item_id, COALESCE(SUM(qty_in - qty_out), 0) AS adj
        FROM inventory_ledger
        WHERE txn_type IN ('Stock Adjustment','Variance Adjustment')
          AND strftime('%Y-%m', txn_date) = ?
        GROUP BY item_id
    """, (prev_period,)).fetchall():
        prior_adjustment[row["item_id"]] = row["adj"]

    # Inbound from APPROVED POs with expected delivery in this period
    approved_inbound = {}
    _, num_days = _cal2.monthrange(p_year, p_month)
    period_start = f"{period}-01"
    period_end   = f"{period}-{num_days:02d}"
    for row in db.execute("""
        SELECT pl.item_id, COALESCE(SUM(pl.qty_ordered + COALESCE(pl.free_units,0)),0) AS qty_in
        FROM po_lines pl
        JOIN purchase_orders po ON pl.po_id = po.id
        WHERE po.status IN ('Approved','Deposit Paid','Fully Paid – Shipped','Partially Received','Fully Paid – Received')
          AND po.expected_delivery_date BETWEEN ? AND ?
        GROUP BY pl.item_id
    """, (period_start, period_end)).fetchall():
        approved_inbound[row["item_id"]] = row["qty_in"]

    # Inbound from UNAPPROVED POs with expected delivery in this period
    pending_inbound = {}
    for row in db.execute("""
        SELECT pl.item_id, COALESCE(SUM(pl.qty_ordered + COALESCE(pl.free_units,0)),0) AS qty_in
        FROM po_lines pl
        JOIN purchase_orders po ON pl.po_id = po.id
        WHERE po.status IN ('Draft','Pending Approval','Ops Approved','Returned')
          AND po.expected_delivery_date BETWEEN ? AND ?
        GROUP BY pl.item_id
    """, (period_start, period_end)).fetchall():
        pending_inbound[row["item_id"]] = row["qty_in"]

    # Load any saved manual forecast adjustments
    forecast_adj = {}
    try:
        for row in db.execute("""
            SELECT key, value FROM settings WHERE key LIKE 'forecast_adj_%'
        """).fetchall():
            # key format: forecast_adj_{period}_{item_id}
            parts = row["key"].split("_", 3)  # ['forecast', 'adj', period, item_id]
            if len(parts) == 4 and parts[2] == period:
                forecast_adj[parts[3]] = float(row["value"] or 0)
    except Exception:
        pass

    rows = []
    for item in items:
        iid = item["item_id"]
        opening     = opening_balances.get(iid, 0)
        outbound    = prior_outbound.get(iid, 0)
        adjustment  = prior_adjustment.get(iid, 0)
        total_out   = outbound + abs(adjustment) if adjustment < 0 else outbound
        manual_adj  = forecast_adj.get(iid, 0)
        inbound_app = approved_inbound.get(iid, 0)
        inbound_pen = pending_inbound.get(iid, 0)
        final_impact = inbound_app + inbound_pen - (outbound - adjustment) - manual_adj
        final_balance = opening + final_impact
        rows.append({
            "item_id":      iid,
            "name":         item["name"],
            "opening":      opening,
            "outbound":     outbound,
            "adjustment":   adjustment,
            "total_out":    outbound - adjustment,
            "manual_adj":   manual_adj,
            "inbound_app":  inbound_app,
            "inbound_pen":  inbound_pen,
            "final_impact": final_impact,
            "final_balance":final_balance,
        })

    return render_template("runout_estimator.html",
        rows=rows, period=period,
        today=today.isoformat())


@app.route("/runout-estimator/save-forecast", methods=["POST"])
@login_required
def save_forecast_adj():
    if session.get("role") not in ("admin","ops_planner","ops_manager"):
        return jsonify({"ok": False, "error": "Permission denied"}), 403
    db = get_db()
    data   = request.get_json()
    period = data.get("period","")
    adjs   = data.get("adjustments", {})  # {item_id: value}
    for item_id, val in adjs.items():
        key = f"forecast_adj_{period}_{item_id}"
        try:
            float_val = float(val)
        except (TypeError, ValueError):
            float_val = 0.0
        db.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                   (key, str(float_val)))
    db.commit()
    return jsonify({"ok": True, "saved": len(adjs)})


@app.route("/export/ap-summary.csv")
@login_required
@roles_required("admin", "finance_manager", "finance_exec")
def export_ap_summary():
    from datetime import date as _d
    db = get_db()
    today = _d.today()
    date_from = request.args.get("date_from", today.replace(day=1).isoformat())
    date_to   = request.args.get("date_to",   today.isoformat())

    rows = []

    # Section 1 — Outstanding AP (unpaid approved POs)
    outstanding = db.execute("""
        SELECT po.po_number, s.name AS supplier, po.currency,
               po.status, po.payment_due_date, po.po_date,
               COALESCE(po.amount_paid,0) AS amount_paid,
               ROUND(COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0),2) AS total_value
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id=s.supplier_id
        LEFT JOIN (SELECT po_id,SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
          ON ts.po_id=po.id
        WHERE po.status IN ('Approved','Deposit Paid','Fully Paid – Shipped','Partially Received')
        ORDER BY po.payment_due_date ASC
    """).fetchall()
    for r in outstanding:
        rows.append({
            "section":          "Outstanding AP",
            "po_number":        r["po_number"],
            "supplier":         r["supplier"],
            "currency":         r["currency"],
            "status":           r["status"],
            "date":             r["po_date"],
            "payment_due_date": r["payment_due_date"] or "",
            "total_value":      r["total_value"],
            "amount_paid":      r["amount_paid"],
            "balance":          round(r["total_value"] - r["amount_paid"], 2),
            "pay_type":         "",
            "amount":           "",
        })

    # Section 2 — Payments made in date range
    payments = db.execute("""
        SELECT po.po_number, s.name AS supplier, po.currency,
               po.deposit_date, po.full_payment_date,
               CASE WHEN po.deposit_amount_paid > 0 THEN po.deposit_amount_paid
                    ELSE po.amount_paid END AS dep_amount,
               po.deposit_amount_paid,
               ROUND(COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0),2) AS total_value,
               COALESCE(po.amount_paid,0) AS amount_paid
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id=s.supplier_id
        LEFT JOIN (SELECT po_id,SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
          ON ts.po_id=po.id
        WHERE (po.deposit_date BETWEEN ? AND ? OR po.full_payment_date BETWEEN ? AND ?)
        ORDER BY COALESCE(po.full_payment_date, po.deposit_date) DESC
    """, (date_from, date_to, date_from, date_to)).fetchall()

    for r in payments:
        if r["deposit_date"] and date_from <= r["deposit_date"] <= date_to:
            rows.append({
                "section":          f"Payments ({date_from} to {date_to})",
                "po_number":        r["po_number"],
                "supplier":         r["supplier"],
                "currency":         r["currency"],
                "status":           "Deposit",
                "date":             r["deposit_date"],
                "payment_due_date": "",
                "total_value":      r["total_value"],
                "amount_paid":      "",
                "balance":          "",
                "pay_type":         "Deposit",
                "amount":           r["dep_amount"],
            })
        if r["full_payment_date"] and date_from <= r["full_payment_date"] <= date_to:
            bal = round(r["total_value"] - (r["deposit_amount_paid"] or 0), 2)
            rows.append({
                "section":          f"Payments ({date_from} to {date_to})",
                "po_number":        r["po_number"],
                "supplier":         r["supplier"],
                "currency":         r["currency"],
                "status":           "Full Payment",
                "date":             r["full_payment_date"],
                "payment_due_date": "",
                "total_value":      r["total_value"],
                "amount_paid":      "",
                "balance":          "",
                "pay_type":         "Full Payment",
                "amount":           bal if r["deposit_amount_paid"] > 0 else r["amount_paid"],
            })

    fields = ["section","po_number","supplier","currency","status","date",
              "payment_due_date","total_value","amount_paid","balance","pay_type","amount"]
    return _csv_response(rows, fields,
                         f"ap_summary_{date_from}_to_{date_to}.csv")


@app.route("/export/stockcount-recon/<period>.csv")
@login_required
@roles_required(*MEC_ROLES)
def export_stockcount_recon(period):
    db = get_db()
    period_start, next_start = _period_bounds(period)
    book = _book_balances(db, period)
    items_map = {r["item_id"]: r["name"] for r in
                 db.execute("SELECT item_id, name FROM items").fetchall()}

    sc = db.execute(
        "SELECT * FROM stockcount_uploads WHERE period=? AND status='Approved' ORDER BY id DESC LIMIT 1",
        (period,)).fetchone()
    sc_lines = db.execute(
        "SELECT * FROM stockcount_lines WHERE upload_id=?",
        (sc["id"],)).fetchall() if sc else []
    physical_map = {l["item_id"]: l["physical_qty"] or 0 for l in sc_lines}

    all_skus = sorted(set(book.keys()) | set(physical_map.keys()))
    rows = []
    for sku in all_skus:
        b    = book.get(sku, {"opening":0,"receipts_in":0,"adj_in":0,"adj_out":0,"book":0})
        phys = physical_map.get(sku, b["book"])
        var  = phys - b["book"]
        rows.append({
            "period":      period,
            "item_id":     sku,
            "item_name":   items_map.get(sku, "—"),
            "opening_qty": b["opening"],
            "receipts_in": b["receipts_in"],
            "adj_in":      b["adj_in"],
            "adj_out":     b["adj_out"],
            "book_qty":    b["book"],
            "physical_qty":phys,
            "variance":    var,
            "variance_pct": round(var / b["book"] * 100, 2) if b["book"] else "",
        })

    fields = ["period","item_id","item_name","opening_qty","receipts_in",
              "adj_in","adj_out","book_qty","physical_qty","variance","variance_pct"]
    return _csv_response(rows, fields,
                         f"stockcount_recon_{period}.csv")



# ─── CSV EXPORT ROUTES ───────────────────────────────────────────────────────
import csv as _csv
import io as _io
from datetime import date as _date_cls

def _csv_response(rows, fieldnames, filename):
    """Build a CSV HTTP response from a list of dicts."""
    from flask import Response
    buf = _io.StringIO()
    w = _csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    w.writeheader()
    w.writerows(rows)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.route("/export/pos.csv")
@login_required
def export_pos():
    db = get_db()
    rows = db.execute("""
        SELECT po.po_number, po.po_date, s.name AS supplier,
               po.currency, po.status,
               ROUND(COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0),2) AS total_value,
               COALESCE(po.amount_paid,0) AS amount_paid,
               ROUND(COALESCE(ts.subtotal,0)+COALESCE(po.freight_costs,0)+COALESCE(po.other_costs,0)-COALESCE(po.amount_paid,0),2) AS outstanding,
               po.payment_due_date, po.expected_delivery_date,
               po.raised_by, po.ops_approved_by, po.approved_by,
               po.received_date, po.notes
        FROM purchase_orders po
        JOIN suppliers s ON po.supplier_id = s.supplier_id
        LEFT JOIN (SELECT po_id, SUM(qty_ordered*unit_price) AS subtotal FROM po_lines GROUP BY po_id) ts
          ON ts.po_id = po.id
        ORDER BY po.po_date DESC, po.po_number
    """).fetchall()
    fields = ["po_number","po_date","supplier","currency","status",
              "total_value","amount_paid","outstanding",
              "payment_due_date","expected_delivery_date",
              "raised_by","ops_approved_by","approved_by",
              "received_date","notes"]
    return _csv_response([dict(r) for r in rows], fields,
                         f"po_listing_{_date_cls.today()}.csv")


@app.route("/export/inventory.csv")
@login_required
def export_inventory():
    db = get_db()
    summary = get_wacc_summary(db)
    fields = ["item_id","name","qty_on_hand","wacc","inventory_value",
              "total_cogs","total_purchased","reorder_point"]
    rows = [{
        "item_id":       s["item_id"],
        "name":          s["name"],
        "qty_on_hand":   round(s["qty_on_hand"], 4),
        "wacc":          round(s["wacc"], 4),
        "inventory_value": round(s["inventory_value"], 2),
        "total_cogs":    round(s["total_cogs"], 2),
        "total_purchased": s["total_purchased"],
        "reorder_point": s["reorder_point"],
    } for s in summary]
    return _csv_response(rows, fields,
                         f"inventory_listing_{_date_cls.today()}.csv")


@app.route("/export/items.csv")
@login_required
def export_items():
    db = get_db()
    rows = db.execute("""
        SELECT item_id, name, description, unit_of_measure, unit_price,
               reorder_point, status, created_at
        FROM items ORDER BY item_id
    """).fetchall()
    fields = ["item_id","name","description","unit_of_measure",
              "unit_price","reorder_point","status","created_at"]
    return _csv_response([dict(r) for r in rows], fields,
                         f"items_listing_{_date_cls.today()}.csv")


@app.route("/export/suppliers.csv")
@login_required
def export_suppliers():
    db = get_db()
    rows = db.execute("""
        SELECT s.supplier_id, s.name, s.contact_person, s.email, s.phone,
               s.address, s.currency, s.payment_terms, s.status,
               COUNT(si.item_id) AS linked_skus
        FROM suppliers s
        LEFT JOIN supplier_items si ON si.supplier_id = s.supplier_id
        GROUP BY s.supplier_id ORDER BY s.name
    """).fetchall()
    fields = ["supplier_id","name","contact_person","email","phone",
              "address","currency","payment_terms","status","linked_skus"]
    return _csv_response([dict(r) for r in rows], fields,
                         f"suppliers_listing_{_date_cls.today()}.csv")


@app.route("/export/bundles.csv")
@login_required
def export_bundles():
    db = get_db()
    bundles = db.execute(
        "SELECT bundle_id, name, description, status FROM bundles ORDER BY bundle_id"
    ).fetchall()
    rows = []
    for b in bundles:
        components = db.execute("""
            SELECT bc.item_id, i.name AS item_name, bc.qty
            FROM bundle_components bc
            JOIN items i ON bc.item_id = i.item_id
            WHERE bc.bundle_id = ?
        """, (b["bundle_id"],)).fetchall()
        if components:
            for c in components:
                rows.append({
                    "bundle_id":   b["bundle_id"],
                    "bundle_name": b["name"],
                    "description": b["description"],
                    "status":      b["status"],
                    "component_item_id":   c["item_id"],
                    "component_item_name": c["item_name"],
                    "component_qty":       c["qty"],
                })
        else:
            rows.append({
                "bundle_id":   b["bundle_id"],
                "bundle_name": b["name"],
                "description": b["description"],
                "status":      b["status"],
                "component_item_id":   "",
                "component_item_name": "",
                "component_qty":       "",
            })
    fields = ["bundle_id","bundle_name","description","status",
              "component_item_id","component_item_name","component_qty"]
    return _csv_response(rows, fields,
                         f"bundles_listing_{_date_cls.today()}.csv")


@app.route("/export/ledger.csv")
@login_required
def export_ledger():
    db = get_db()
    rows = db.execute("""
        SELECT l.txn_date, l.txn_type, l.item_id, i.name AS item_name,
               l.qty_in, l.qty_out, l.unit_cost,
               ROUND(COALESCE(l.qty_in,0)*COALESCE(l.unit_cost,0) +
                     COALESCE(l.qty_out,0)*COALESCE(l.unit_cost,0), 4) AS txn_value,
               l.reference, l.notes, l.created_by
        FROM inventory_ledger l
        JOIN items i ON l.item_id = i.item_id
        ORDER BY l.txn_date DESC, l.id DESC
    """).fetchall()
    fields = ["txn_date","txn_type","item_id","item_name",
              "qty_in","qty_out","unit_cost","txn_value",
              "reference","notes","created_by"]
    return _csv_response([dict(r) for r in rows], fields,
                         f"inventory_ledger_{_date_cls.today()}.csv")


@app.route("/export/sales.csv")
@login_required
def export_sales():
    db = get_db()
    rows = db.execute("""
        SELECT su.period, su.platform, sl.row_num,
               sl.product_name, sl.sku_id,
               sl.qty_sold, sl.unit_selling_price,
               sl.gross_revenue, sl.platform_fees, sl.net_revenue,
               su.upload_ref, su.uploaded_by, su.created_at
        FROM sales_lines sl
        JOIN sales_uploads su ON sl.upload_id = su.id
        ORDER BY su.period DESC, su.platform, sl.row_num
    """).fetchall()
    fields = ["period","platform","row_num","product_name","sku_id",
              "qty_sold","unit_selling_price","gross_revenue",
              "platform_fees","net_revenue",
              "upload_ref","uploaded_by","created_at"]
    return _csv_response([dict(r) for r in rows], fields,
                         f"sales_data_{_date_cls.today()}.csv")



# ── SUPPLIER IMPORT ──────────────────────────────────────────────────────────
@app.route("/suppliers/import", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def supplier_import():
    if request.method == "GET":
        return render_template("supplier_import.html")

    f = request.files.get("supplier_file")
    if not f or not f.filename:
        flash("No file selected.", "danger")
        return render_template("supplier_import.html")

    ext = f.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx", "xls"):
        flash("Only .csv, .xlsx or .xls files are accepted.", "danger")
        return render_template("supplier_import.html")

    import io, tempfile, os as _os2
    tmp_path = None
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}")
        f.save(tmp.name)
        tmp.close()
        tmp_path = tmp.name

        rows = []
        if ext == "csv":
            import csv as _csv
            with open(tmp_path, newline="", encoding="utf-8-sig") as cf:
                reader = _csv.DictReader(cf)
                rows = [dict(r) for r in reader]
        else:
            try:
                import openpyxl as _xl2
            except ImportError:
                flash("openpyxl required for Excel uploads. Run: pip install openpyxl", "danger")
                return render_template("supplier_import.html")
            wb = _xl2.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            all_rows = list(ws.values)
            if not all_rows:
                flash("File is empty.", "danger")
                return render_template("supplier_import.html")
            headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in all_rows[0]]
            for r in all_rows[1:]:
                if not r or all(v is None for v in r):
                    continue
                rows.append({headers[i]: (str(r[i]).strip() if r[i] is not None else "")
                             for i in range(len(headers))})

        def _g(row, *keys):
            for k in keys:
                v = row.get(k, "").strip()
                if v:
                    return v
            return ""

        db = get_db()
        created = updated = skipped = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            sup_id = _g(row, "supplier_id", "id", "code")
            name   = _g(row, "name", "supplier_name", "company")
            if not sup_id or not name:
                skipped += 1
                continue
            contact = _g(row, "contact_name", "contact", "contact_person")
            email   = _g(row, "email", "email_address")
            phone   = _g(row, "phone", "telephone", "mobile")
            terms   = _g(row, "payment_terms", "terms", "payment")
            try:
                lead = int(float(_g(row, "lead_days", "lead_time", "leadtime") or 14))
            except ValueError:
                lead = 14

            existing = db.execute(
                "SELECT id FROM suppliers WHERE supplier_id=?", (sup_id,)).fetchone()
            if existing:
                db.execute("""
                    UPDATE suppliers SET name=?, contact_name=?, email=?, phone=?,
                        payment_terms=?, lead_days=?, status='Active'
                    WHERE supplier_id=?""",
                    (name, contact, email, phone, terms, lead, sup_id))
                updated += 1
            else:
                db.execute("""
                    INSERT INTO suppliers
                        (supplier_id, name, contact_name, email, phone,
                         payment_terms, lead_days, status)
                    VALUES (?,?,?,?,?,?,?,'Active')""",
                    (sup_id, name, contact, email, phone, terms, lead))
                created += 1

        db.commit()

    except Exception as e:
        flash(f"Import error: {e}", "danger")
        return render_template("supplier_import.html")
    finally:
        if tmp_path and _os2.path.exists(tmp_path):
            _os2.unlink(tmp_path)

    return render_template("supplier_import.html",
        result={"created": created, "updated": updated, "skipped": skipped, "errors": errors})


# ── TEMPLATES DOWNLOAD PAGE ───────────────────────────────────────────────────
@app.route("/templates")
@login_required
def templates_page():
    return render_template("templates_page.html")


@app.route("/templates/download/<name>")
@login_required
def template_download(name):
    """Generate and return a pre-formatted Excel template."""
    import io as _io
    try:
        import openpyxl as _xl3
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash("openpyxl required. Run: pip install openpyxl", "danger")
        return redirect(url_for("templates_page"))

    wb = _xl3.Workbook()

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    EX_FILL   = PatternFill("solid", fgColor="EBF3FB")
    EX_FONT   = Font(color="555555", italic=True, size=10)
    THIN      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    def _sheet(ws, headers, example):
        ws.sheet_view.showGridLines = True
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.font, c.fill, c.alignment, c.border = (
                HDR_FONT, HDR_FILL, Alignment(horizontal="center", vertical="center"), THIN)
            ws.column_dimensions[c.column_letter].width = max(18, len(h) + 4)
        for ci, v in enumerate(example, 1):
            c = ws.cell(2, ci, v)
            c.font, c.fill, c.border = EX_FONT, EX_FILL, THIN
        ws.row_dimensions[1].height = 22

    if name == "suppliers":
        ws = wb.active; ws.title = "Suppliers"
        _sheet(ws,
            ["supplier_id","name","contact_name","email","phone","payment_terms","lead_days"],
            ["SUP-001","Acme Trading Pte Ltd","John Tan","john@acme.com","+65 9123 4567","Net 30","14"])
        filename = "template_suppliers.xlsx"

    elif name == "items":
        ws = wb.active; ws.title = "Items_SKUs"
        _sheet(ws,
            ["item_id","name","brand","unit_price","unit_of_measure","reorder_point","lead_days","preferred_supplier"],
            ["SKU-001","Product Name","andSons","12.50","EA","10","14","SUP-001"])
        filename = "template_items_skus.xlsx"

    elif name == "inventory_count":
        ws = wb.active; ws.title = "Inventory_Count"
        _sheet(ws,
            ["item_id","physical_qty"],
            ["SKU-001","100"])
        note = ws.cell(4, 1, "Upload this file via: Inventory → Upload Physical Count")
        note.font = Font(color="888888", italic=True, size=9)
        note2 = ws.cell(5, 1, "Period is selected on the upload page — do NOT include a period column.")
        note2.font = Font(color="888888", italic=True, size=9)
        filename = "template_inventory_count.xlsx"

    elif name == "bundles":
        ws = wb.active; ws.title = "Bundles"
        _sheet(ws,
            ["bundle_id","description","brand","sku_1","sku_2","sku_3","sku_4","sku_5"],
            ["BUNDLE-001","Shampoo + Conditioner Set","andSons","SKU-001","SKU-002","","",""])
        filename = "template_bundles.xlsx"

    elif name == "master_data":
        # 3-tab format matching the SKU Cost Price Listings file
        for sheet_name in wb.sheetnames:
            del wb[sheet_name]
        for tab, brand in [("AS", "andSons"), ("OVA", "Ova")]:
            ws = wb.create_sheet(tab)
            # Rows 1-7 are header/info rows in the real file — row 8 is data start
            ws.cell(1, 1, f"{brand} SKU Cost Price Listings")
            ws.cell(7, 2, "Product SKU"); ws.cell(7, 3, "Product Name")
            ws.cell(7, 5, "Finance COGS (ex-GST)"); ws.cell(7, 6, "Remark")
            for ci in [2, 3, 5, 6]:
                c = ws.cell(7, ci)
                c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN
            ws.cell(8, 2, "SKU-001"); ws.cell(8, 3, "Example Product")
            ws.cell(8, 5, 12.50); ws.cell(8, 6, "")
        ws3 = wb.create_sheet("&Sons (Update)")
        bundle_hdrs = ["Bundle ID","","","Description","","","","","Brand",
                       "SKU 1","","SKU 2","","SKU 3","","SKU 4","","SKU 5",""]
        for ci, h in enumerate(bundle_hdrs, 1):
            if h:
                c = ws3.cell(1, ci, h)
                c.font, c.fill, c.border = HDR_FONT, HDR_FILL, THIN
        example = ["BUNDLE-001","","","Example Bundle","","","","","andSons",
                   "SKU-001","","SKU-002","","","","","","",""]
        for ci, v in enumerate(example, 1):
            if v:
                c = ws3.cell(2, ci, v)
                c.font, c.fill, c.border = EX_FONT, EX_FILL, THIN
        filename = "template_master_data.xlsx"

    else:
        flash("Unknown template.", "danger")
        return redirect(url_for("templates_page"))

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



# ── SUPPLIER REQUEST DOC UPLOAD ───────────────────────────────────────────────
@app.route("/suppliers/request/<int:req_id>/upload-doc", methods=["POST"])
@login_required
def supplier_request_doc_upload(req_id):
    db  = get_db()
    req = db.execute("SELECT * FROM supplier_requests WHERE id=?", (req_id,)).fetchone()
    if not req:
        flash("Request not found.", "danger")
        return redirect(url_for("supplier_list"))
    if req["status"] in ("Approved", "Rejected"):
        flash("Cannot upload documents to a closed request.", "warning")
        return redirect(url_for("supplier_request_detail", req_id=req_id))

    f = request.files.get("doc_file")
    if not f or not f.filename:
        flash("No file selected.", "danger")
        return redirect(url_for("supplier_request_detail", req_id=req_id))

    import os as _osd
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else "bin"
    safe = f"sreq_{req_id}_{int(__import__('time').time())}.{ext}"
    upload_dir = _osd.path.join(_osd.path.dirname(__file__), "static", "uploads")
    _osd.makedirs(upload_dir, exist_ok=True)
    f.save(_osd.path.join(upload_dir, safe))

    notes = request.form.get("doc_notes", "").strip()
    db.execute(
        """INSERT INTO supplier_request_docs
           (request_id, doc_type, filename, original_name, uploaded_by, notes)
           VALUES (?,?,?,?,?,?)""",
        (req_id, request.form.get("doc_type","Support"), safe,
         f.filename, session["username"], notes))
    db.commit()
    flash(f"Document '{f.filename}' uploaded.", "success")
    return redirect(url_for("supplier_request_detail", req_id=req_id))


# ── ITEMS FLAT IMPORT ─────────────────────────────────────────────────────────
@app.route("/items/import", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def items_import():
    if request.method == "GET":
        return render_template("items_import.html")

    f = request.files.get("items_file")
    if not f or not f.filename:
        flash("No file selected.", "danger")
        return render_template("items_import.html")

    ext = f.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx", "xls"):
        flash("Only .csv or .xlsx/.xls accepted.", "danger")
        return render_template("items_import.html")

    import io as _io2, tempfile as _tmp2, os as _os3, csv as _csv2

    tmp = _tmp2.NamedTemporaryFile(delete=False, suffix=f".{ext}")
    f.save(tmp.name); tmp.close()
    tmp_path = tmp.name

    try:
        rows = []
        if ext == "csv":
            with open(tmp_path, newline="", encoding="utf-8-sig") as cf:
                rows = [dict(r) for r in _csv2.DictReader(cf)]
        else:
            try:
                import openpyxl as _xl4
            except ImportError:
                flash("openpyxl required. Run: pip install openpyxl", "danger")
                return render_template("items_import.html")
            wb = _xl4.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            all_rows = list(ws.values)
            if not all_rows:
                flash("File is empty.", "danger")
                return render_template("items_import.html")
            headers = [str(h).strip().lower().replace(" ","_") if h else f"col{i}"
                       for i, h in enumerate(all_rows[0])]
            for r in all_rows[1:]:
                if not r or all(v is None for v in r): continue
                rows.append({headers[i]: (str(r[i]).strip() if r[i] is not None else "")
                             for i in range(min(len(headers), len(r)))})

        def _g(row, *keys):
            for k in keys:
                v = str(row.get(k, "") or "").strip()
                if v: return v
            return ""

        db = get_db()
        created = updated = skipped = 0

        for row in rows:
            item_id = _g(row, "item_id", "sku", "sku_id", "id")
            if not item_id or item_id.lower() in ("item_id","sku","example"):
                skipped += 1; continue
            name   = _g(row, "name", "item_name", "product_name") or item_id
            brand  = _g(row, "brand")
            try:    price = float(_g(row, "unit_price","cost","price") or 0)
            except: price = 0.0
            uom    = _g(row, "unit_of_measure","uom","unit") or "EA"
            try:    reorder = int(float(_g(row, "reorder_point","reorder") or 0))
            except: reorder = 0
            try:    lead = int(float(_g(row, "lead_days","lead_time","leadtime") or 14))
            except: lead = 14
            supplier = _g(row, "preferred_supplier","supplier","supplier_id")
            notes  = _g(row, "notes")

            existing = db.execute("SELECT item_id FROM items WHERE item_id=?", (item_id,)).fetchone()
            if existing:
                db.execute("""UPDATE items SET name=?,brand=?,unit_price=?,unit_of_measure=?,
                    reorder_point=?,lead_days=?,preferred_supplier=?,notes=?,status='Active'
                    WHERE item_id=?""",
                    (name,brand,price,uom,reorder,lead,supplier or None,notes,item_id))
                updated += 1
            else:
                db.execute("""INSERT INTO items
                    (item_id,name,brand,unit_price,unit_of_measure,reorder_point,
                     lead_days,preferred_supplier,notes,status)
                    VALUES (?,?,?,?,?,?,?,?,?,'Active')""",
                    (item_id,name,brand,price,uom,reorder,lead,supplier or None,notes))
                created += 1

        db.commit()
    except Exception as e:
        flash(f"Import error: {e}", "danger")
        return render_template("items_import.html")
    finally:
        import os as _osc; _osc.unlink(tmp_path) if _osc.path.exists(tmp_path) else None

    return render_template("items_import.html",
        result={"created": created, "updated": updated, "skipped": skipped})


# ── BUNDLES FLAT IMPORT ───────────────────────────────────────────────────────
@app.route("/bundles/import", methods=["GET", "POST"])
@login_required
@roles_required("admin", "ops_manager")
def bundles_import():
    if request.method == "GET":
        return render_template("bundles_import.html")

    f = request.files.get("bundles_file")
    if not f or not f.filename:
        flash("No file selected.", "danger")
        return render_template("bundles_import.html")

    ext = f.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("csv", "xlsx", "xls"):
        flash("Only .csv or .xlsx/.xls accepted.", "danger")
        return render_template("bundles_import.html")

    import tempfile as _tmp3, os as _os4, csv as _csv3

    tmp = _tmp3.NamedTemporaryFile(delete=False, suffix=f".{ext}")
    f.save(tmp.name); tmp.close()
    tmp_path = tmp.name

    try:
        rows = []
        if ext == "csv":
            with open(tmp_path, newline="", encoding="utf-8-sig") as cf:
                rows = [dict(r) for r in _csv3.DictReader(cf)]
        else:
            try:
                import openpyxl as _xl5
            except ImportError:
                flash("openpyxl required.", "danger")
                return render_template("bundles_import.html")
            wb = _xl5.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            all_rows = list(ws.values)
            if not all_rows:
                flash("File is empty.", "danger")
                return render_template("bundles_import.html")
            headers = [str(h).strip().lower().replace(" ","_") if h else f"col{i}"
                       for i, h in enumerate(all_rows[0])]
            for r in all_rows[1:]:
                if not r or all(v is None for v in r): continue
                rows.append({headers[i]: (str(r[i]).strip() if r[i] is not None else "")
                             for i in range(min(len(headers), len(r)))})

        def _g(row, *keys):
            for k in keys:
                v = str(row.get(k, "") or "").strip()
                if v and v.lower() not in ("none","n/a",""): return v
            return ""

        db  = get_db()
        user = session.get("username","system")
        now  = datetime.utcnow().isoformat(timespec="seconds")
        created = updated = skipped = 0

        for row in rows:
            bid = _g(row, "bundle_id", "id", "code")
            if not bid or bid.lower() in ("bundle_id","example"):
                skipped += 1; continue
            desc = _g(row, "description", "name", "bundle_name")
            brand = _g(row, "brand")

            skus = []
            for k in ["sku_1","sku_2","sku_3","sku_4","sku_5",
                       "component_1","component_2","component_3","component_4","component_5"]:
                s = _g(row, k)
                if s: skus.append(s)
            if not skus:
                skipped += 1; continue

            existing = db.execute("SELECT id FROM bundles WHERE bundle_id=?", (bid,)).fetchone()
            if existing:
                db.execute("UPDATE bundles SET name=?,description=?,status='Active' WHERE bundle_id=?",
                    (desc or bid, desc, bid))
                db.execute("DELETE FROM bundle_components WHERE bundle_id=?", (bid,))
                updated += 1
            else:
                db.execute("""INSERT INTO bundles (bundle_id,name,description,status,created_by,created_at)
                    VALUES (?,?,?,'Active',?,?)""",
                    (bid, desc or bid, desc, user, now))
                created += 1

            for sku in skus:
                db.execute("""INSERT OR IGNORE INTO bundle_components (bundle_id,item_id,qty)
                    VALUES (?,?,1)""", (bid, sku))

        db.commit()
    except Exception as e:
        flash(f"Import error: {e}", "danger")
        return render_template("bundles_import.html")
    finally:
        import os as _ose; _ose.unlink(tmp_path) if _ose.path.exists(tmp_path) else None

    return render_template("bundles_import.html",
        result={"created": created, "updated": updated, "skipped": skipped})



def run_migrations():
    """Apply any schema migrations to an existing database."""
    import sqlite3 as _sq
    db = _sq.connect(DB_PATH)
    db.row_factory = _sq.Row
    try:
        # Fix purchase_orders: make item_id nullable (multi-SKU POs use po_lines)
        po_cols = {r[1]: r for r in db.execute("PRAGMA table_info(purchase_orders)").fetchall()}
        if po_cols.get("item_id") and po_cols["item_id"][3] == 1:
            print("[migration] Rebuilding purchase_orders to make item_id nullable...")
            db.execute("""CREATE TABLE IF NOT EXISTS purchase_orders_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                po_number TEXT NOT NULL UNIQUE, po_date TEXT NOT NULL,
                supplier_id TEXT NOT NULL, item_id TEXT,
                qty_ordered REAL DEFAULT 0, unit_price REAL DEFAULT 0,
                free_units REAL DEFAULT 0, freight_costs REAL DEFAULT 0,
                other_costs REAL DEFAULT 0, currency TEXT DEFAULT 'SGD',
                fx_rate REAL DEFAULT 1.0, deposit_pct REAL DEFAULT 0,
                status TEXT DEFAULT 'Draft', raised_by TEXT,
                approved_by TEXT, approval_date TEXT, deposit_date TEXT,
                full_payment_date TEXT, expected_delivery_date TEXT,
                received_date TEXT, received_qty REAL, notes TEXT,
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT DEFAULT (datetime('now')),
                payment_due_date TEXT)""")
            db.execute("""INSERT INTO purchase_orders_new
                SELECT id,po_number,po_date,supplier_id,item_id,qty_ordered,unit_price,
                       free_units,freight_costs,other_costs,currency,fx_rate,deposit_pct,
                       status,raised_by,approved_by,approval_date,deposit_date,
                       full_payment_date,expected_delivery_date,received_date,received_qty,
                       notes,created_at,updated_at,payment_due_date
                FROM purchase_orders""")
            db.execute("DROP TABLE purchase_orders")
            db.execute("ALTER TABLE purchase_orders_new RENAME TO purchase_orders")
            db.commit()
            print("[migration] purchase_orders rebuilt OK")

        existing = {r[1] for r in db.execute("PRAGMA table_info(supplier_items)").fetchall()}
        if "unit_price" not in existing:
            db.execute("ALTER TABLE supplier_items ADD COLUMN unit_price REAL")
            print("[migration] Added unit_price to supplier_items")

        existing = {r[1] for r in db.execute("PRAGMA table_info(items)").fetchall()}
        if "unit_price" not in existing:
            db.execute("ALTER TABLE items ADD COLUMN unit_price REAL DEFAULT 0")
            print("[migration] Added unit_price to items")

        existing = {r[1] for r in db.execute("PRAGMA table_info(supplier_requests)").fetchall()}
        if "prop_supplier_id" not in existing:
            db.execute("ALTER TABLE supplier_requests ADD COLUMN prop_supplier_id TEXT")
            print("[migration] Added prop_supplier_id to supplier_requests")

        existing = {r[1] for r in db.execute("PRAGMA table_info(purchase_orders)").fetchall()}
        if "payment_due_date" not in existing:
            db.execute("ALTER TABLE purchase_orders ADD COLUMN payment_due_date TEXT")
            print("[migration] Added payment_due_date to purchase_orders")
        if "cancelled_by" not in existing:
            db.execute("ALTER TABLE purchase_orders ADD COLUMN cancelled_by TEXT")
            print("[migration] Added cancelled_by to purchase_orders")
        if "cancelled_at" not in existing:
            db.execute("ALTER TABLE purchase_orders ADD COLUMN cancelled_at TEXT")
            print("[migration] Added cancelled_at to purchase_orders")
        if "cancel_reason" not in existing:
            db.execute("ALTER TABLE purchase_orders ADD COLUMN cancel_reason TEXT")
            print("[migration] Added cancel_reason to purchase_orders")

        # ── sales_uploads extra columns ──────────────────────────────────────
        existing = {r[1] for r in db.execute("PRAGMA table_info(sales_uploads)").fetchall()}
        if "parser_type" not in existing:
            db.execute("ALTER TABLE sales_uploads ADD COLUMN parser_type TEXT DEFAULT 'generic'")
            print("[migration] Added parser_type to sales_uploads")
        if "file_type" not in existing:
            db.execute("ALTER TABLE sales_uploads ADD COLUMN file_type TEXT DEFAULT 'transaction'")
            print("[migration] Added file_type to sales_uploads")
        if "period" not in existing:
            db.execute("ALTER TABLE sales_uploads ADD COLUMN period TEXT")
            print("[migration] Added period to sales_uploads")

        # ── sales_lines smart-parser columns ────────────────────────────────
        existing = {r[1] for r in db.execute("PRAGMA table_info(sales_lines)").fetchall()}
        for col, defn in [
            ("brand",            "TEXT"),
            ("tx_type",          "TEXT"),
            ("tx_status",        "TEXT"),
            ("tx_amount",        "REAL DEFAULT 0"),
            ("gst_amount",       "REAL DEFAULT 0"),
            ("amt_receivable",   "REAL DEFAULT 0"),
            ("expenses_total",   "REAL DEFAULT 0"),
            ("is_flagged",       "INTEGER DEFAULT 0"),
            ("flag_reason",      "TEXT"),
            ("payout_id",        "TEXT"),
            ("payout_date",      "TEXT"),
            ("sponsored_voucher","REAL DEFAULT 0"),
            ("fees_incl_gst",    "REAL DEFAULT 0"),
            ("fees_ex_gst",      "REAL DEFAULT 0"),
            ("gst_on_fees",      "REAL DEFAULT 0"),
            ("rebates",          "REAL DEFAULT 0"),
            ("payout_amount",    "REAL DEFAULT 0"),
        ]:
            if col not in existing:
                db.execute(f"ALTER TABLE sales_lines ADD COLUMN {col} {defn}")
                print(f"[migration] Added {col} to sales_lines")

        # ── items.brand ──────────────────────────────────────────────────────
        existing = {r[1] for r in db.execute("PRAGMA table_info(items)").fetchall()}
        if "brand" not in existing:
            db.execute("ALTER TABLE items ADD COLUMN brand TEXT")
            print("[migration] Added brand to items")

        # ── settings table ───────────────────────────────────────────────────
        db.execute("""CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        )""")

        # ── users.profile_photo ───────────────────────────────────────────────
        existing = {r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()}
        if "profile_photo" not in existing:
            db.execute("ALTER TABLE users ADD COLUMN profile_photo TEXT")
            print("[migration] Added profile_photo to users")

        # ── supplier_request_docs ─────────────────────────────────────────────
        db.execute("""CREATE TABLE IF NOT EXISTS supplier_request_docs (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            request_id    INTEGER NOT NULL,
            doc_type      TEXT    NOT NULL DEFAULT 'Support',
            filename      TEXT    NOT NULL,
            original_name TEXT    NOT NULL,
            uploaded_by   TEXT    NOT NULL,
            notes         TEXT,
            uploaded_at   TEXT    DEFAULT (datetime('now'))
        )""")

        db.commit()
        print("[migration] All migrations applied.")

    except Exception as e:
        db.rollback()
        print(f"[migration] ERROR: {e}")
        raise
    finally:
        db.close()


# ── APP ENTRY POINT ───────────────────────────────────────────────────────────
# Run migrations on every startup (works both directly and via WSGI)
run_migrations()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
